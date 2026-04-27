# MAPS MONITORING WIFI PREMIUM
# Copyrigth (c) 2025 PEYCELL GROUP

import json
import os
import psutil
import time
import threading
import shutil
import requests
import subprocess
import platform
import tempfile
import random
import py_compile
from datetime import datetime, timedelta
import calendar
from flask import Flask, render_template, jsonify, request, send_file, redirect, url_for, make_response, send_from_directory
import sys
import math
import importlib.util
try:
    import openpyxl
    OPENPYXL_READY = True
except ImportError:
    openpyxl = None
    OPENPYXL_READY = False
    print("[INIT] Excel Engine : WARNING (openpyxl not installed. Import/Export limited)")

LAST_BACKUP_DATE = "" # Global variable to track last auto-backup date
LAST_BILLING_NOTIF_DATE = "" # Global variable to track last billing notification date

def get_month_name(month_num, lang='id'):
    """Helper for localized month name"""
    months_id = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", 
                 "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    months_en = ["January", "February", "March", "April", "May", "June",
                 "July", "August", "September", "October", "November", "December"]
    if lang == 'en': return months_en[month_num - 1]
    return months_id[month_num - 1]

def format_relative_time(ts):
    """Mengonversi timestamp ke format 'X menit/jam yang lalu' (Bahasa Indonesia)"""
    if not ts or ts == 0: return "n/a"
    try:
        now = time.time()
        diff = int(now - ts)
        if diff < 60: return "Baru saja"
        if diff < 3600: return f"{diff // 60} menit yang lalu"
        if diff < 86400: return f"{diff // 3600} jam yang lalu"
        return f"{diff // 86400} hari yang lalu"
    except: return "Format waktu salah"
import zipfile
import re
from db_manager import DBManager

CURRENT_VERSION = '3.6.8' # Versi Aplikasi Saat Ini
# Link Folder Pusat (Obfuscated Base64)
_B64_FOLDER = "aHR0cHM6Ly9kcml2ZS5nb29nbGUuY29tL2RyaXZlL2ZvbGRlcnMvMW5PQUFULUNiOVFGRjlpb2l6Yzl5OGtJcWxxV1ByS1FX"

def get_pusat_url():
    """Mengembalikan URL pusat yang sudah didekripsi."""
    import base64
    try: return base64.b64decode(_B64_FOLDER).decode('utf-8')
    except: return ""

def extract_gdrive_id(url):
    """Mengekstrak file ID dari berbagai format link Google Drive."""
    pats = [
        r'id=([a-zA-Z0-9_-]+)',
        r'file/d/([a-zA-Z0-9_-]+)',
        r'/d/([a-zA-Z0-9_-]+)',
        r'folders/([a-zA-Z0-9_-]+)',
        r'open\?id=([a-zA-Z0-9_-]+)'
    ]
    for p in pats:
        m = re.search(p, url)
        if m: return m.group(1)
    return None

def download_gdrive_file(file_id, timeout=30):
    """Download file dari GDrive dengan handle virus warning/large file confirmation."""
    URL = "https://docs.google.com/uc?export=download"
    session = requests.Session()
    # Use a real browser User-Agent to avoid getting blocked or different layouts
    session.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'})
    
    # Try first download attempt
    response = session.get(URL, params={'id': file_id}, timeout=timeout)
    
    def get_form_fields(html):
        fields = {}
        # Find all hidden inputs (common in GDrive confirmation pages)
        inputs = re.findall(r'<input type="hidden" name="([^"]+)" value="([^"]+)">', html)
        for name, value in inputs:
            fields[name] = value
        
        # Also check for confirm= in URL/links as backup for some layouts
        if 'confirm' not in fields:
            m = re.search(r'confirm=([a-zA-Z0-9_-]+)', html)
            if m: fields['confirm'] = m.group(1)
            
        return fields

    fields = get_form_fields(response.text)
    if 'confirm' in fields:
        # Use drive.usercontent.google.com for confirmation (standard for GDrive)
        confirm_url = "https://drive.usercontent.google.com/download"
        # Ensure we have essential fields
        fields['id'] = file_id
        response = session.get(confirm_url, params=fields, stream=True, timeout=timeout)
        
    return response
        
    return response

# CRITICAL: Add this script's directory to path FIRST
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

BASE_DIR = SCRIPT_DIR

# --- WA-BRIDGE AUTO-MOVE HOTFIX ---
# Pindahkan otomatis wa-bridge.js dari static/js ke root jika ada.
# Hal ini penting untuk proses OTA update, karena framework default
# update akan meletakkan file .js di folder static/js.
def _auto_move_wa_bridge():
    wrong_path = os.path.join(BASE_DIR, 'static', 'js', 'wa-bridge.js')
    correct_path = os.path.join(BASE_DIR, 'wa-bridge.js')
    
    if os.path.exists(wrong_path):
        import shutil
        try:
            # Pindahkan file secara aman
            if os.path.exists(correct_path):
                os.remove(correct_path)
            shutil.move(wrong_path, correct_path)
            print(f"[INIT] OTA Patch: wa-bridge.js berhasil dipindah ke root directory.")
        except Exception as e:
            print(f"[INIT] OTA Patch Error: Gagal memindah wa-bridge.js - {e}")

# Jalankan hotfix pada saat startup
_auto_move_wa_bridge()

# Try importing license_utils with detailed error reporting
try:
    # Use standard import to support both .py and .so (Nuitka)
    import license_utils
    
    verify_license = license_utils.verify_license
    get_machine_id = license_utils.get_machine_id
    
except Exception as e:
    err_msg = str(e)
    # Fallback dummy jika license_utils hilang
    def verify_license(k): return False, "Module Missing"
    def get_machine_id(): return f"IMPORT-ERR: {err_msg}"

def get_machine_id_cached():
    """Optimized: Cache machine ID as it never changes during runtime."""
    global _MACHINE_ID_CACHE
    if _MACHINE_ID_CACHE is None:
        _MACHINE_ID_CACHE = get_machine_id()
    return _MACHINE_ID_CACHE

# --- THREAD LOCKS ---
db_lock = threading.Lock()
log_lock = threading.Lock()
billing_lock = threading.Lock()
tg_lock = threading.Lock()

# DATA PATHS
LOG_FILE = os.path.join(SCRIPT_DIR, 'logs.json')
CONFIG_FILE = os.path.join(SCRIPT_DIR, 'config.json')
SETTINGS_FILE = os.path.join(SCRIPT_DIR, 'settings.json')
FINANCE_FILE = os.path.join(SCRIPT_DIR, 'finance.json')
LICENSE_FILE = os.path.join(SCRIPT_DIR, 'license.key')
BLACKLIST_FILE = os.path.join(SCRIPT_DIR, '.blacklist_cache')
BLACKLIST_URL = "https://drive.google.com/file/d/1S701eAR5OWcH_AzKuA2FEo6GWFBP6oFs/view?usp=sharing"
_cached_blacklist = []
_last_blacklist_sync = 0
_TG_LAST_UPDATE_ID = 0
_TG_UPDATE_STATES = {} # Dictionary to track last_update_id per token: {token: last_id}

# --- PERFORMANCE CACHE ---
_MACHINE_ID_CACHE = None
_LICENSE_CACHE = {"valid": False, "info": {}, "expiry": 0}
_SYSTEM_STATS_CACHE = {"data": {}, "expiry": 0}
_TOPO_CACHE = {"data": None, "expiry": 0}
_SETTINGS_CACHE = {"data": None, "expiry": 0}

def load_cached_blacklist():
    global _cached_blacklist
    if os.path.exists(BLACKLIST_FILE):
        try:
            with open(BLACKLIST_FILE, 'r') as f:
                _cached_blacklist = json.load(f)
        except: _cached_blacklist = []

def perform_blacklist_sync(force=False):
    """Melakukan sinkronisasi blacklist dari URL yang ditanam (Satu kali) dengan throttle 10 menit."""
    global _cached_blacklist, _last_blacklist_sync
    now = time.time()
    # Throttle: Jangan hajar Drive terlalu sering (maks 1x per 10 menit unless force)
    if not force and (now - _last_blacklist_sync) < 600:
        return
    _last_blacklist_sync = now
    
    try:
        f_id = extract_gdrive_id(BLACKLIST_URL)
        if f_id:
            f_url = f"https://docs.google.com/uc?export=download&id={f_id}"
            
            # 2. Download isinya
            f_resp = requests.get(f_url, timeout=15)
            if f_resp.status_code == 200:
                try:
                    data = f_resp.json()
                    if isinstance(data, list):
                        _cached_blacklist = data
                        with open(BLACKLIST_FILE, 'w') as f:
                            json.dump(data, f)
                except:
                    pass
    except:
        pass

def blacklist_sync_loop():
    """Loop background untuk sinkronisasi blacklist setiap 1 jam."""
    while True:
        perform_blacklist_sync()
        time.sleep(3600)

load_cached_blacklist()
# Jalankan loop background sekali saja saat startup
threading.Thread(target=blacklist_sync_loop, daemon=True).start()

BILLING_FILE = os.path.join(SCRIPT_DIR, 'billing.json')
PHOTO_DIR = os.path.join(SCRIPT_DIR, 'static', 'photos')
TEMP_FOLDER = os.path.join(SCRIPT_DIR, 'temp_wa')
if not os.path.exists(TEMP_FOLDER): os.makedirs(TEMP_FOLDER)
temp_folder = TEMP_FOLDER # Compatibility alias
DB_FILE = os.path.join(SCRIPT_DIR, 'topology.db')

# Global DB Manager instance
db = DBManager(DB_FILE)

def _load_settings_raw():
    """ Raw loader to avoid circular dependencies """
    if not os.path.exists(SETTINGS_FILE):
        return dict(DEFAULT_SETTINGS)
    
    # ANTI-DATA LOSS: Jika file ada tapi kosong/error, jangan langsung update default.
    # Lebih baik gagal baca daripada nimpa data lama dengan data kosong.
    try:
        data = _parse_json_file_loose(SETTINGS_FILE, None)
        if data is None:
            # Jika gagal parse, coba restore dari stable
            stable = SETTINGS_FILE + '.stable'
            if os.path.exists(stable):
                data = _parse_json_file_loose(stable, None)
        
        out = dict(DEFAULT_SETTINGS)
        if isinstance(data, dict):
            out.update(data)
            return out
        return out 
    except:
        return dict(DEFAULT_SETTINGS)

def _save_settings_raw(data):
    """ Raw saver for internal use """
    _safe_write_json(SETTINGS_FILE, data, critical=True)

def load_billing_config():
    """ 
    Use billing_lock for settings.get('billing') access.
    Includes billing_profiles from root settings for availability in all billing logic.
    """
    with billing_lock:
        settings = _load_settings_raw()
        billing = settings.get('billing', {}) or {}
        
        # FIX: Include billing_profiles & manual_arrears from root if not present in billing object
        if 'billing_profiles' not in billing:
            billing['billing_profiles'] = settings.get('billing_profiles', {})
        if 'manual_arrears' not in billing:
            billing['manual_arrears'] = settings.get('manual_arrears', [])
            
        if billing:
            # Pastikan bersih dari field lama (permintaan User)
            if isinstance(billing, dict) and 'min_payment_percentage' in billing:
                del billing['min_payment_percentage']
            return billing
            
        if os.path.exists(BILLING_FILE):
            try:
                with open(BILLING_FILE, 'r') as f:
                    legacy = json.load(f)
                
                if 'min_payment_percentage' in legacy: del legacy['min_payment_percentage']
                
                # Merge into settings.json
                settings = _load_settings_raw()
                settings['billing'] = legacy
                _save_settings_raw(settings)
                
                # Cleanup legacy file
                try: os.remove(BILLING_FILE)
                except: pass
                
                return legacy
            except:
                pass
        
        return DEFAULT_SETTINGS['billing']

def save_billing_config(data):
    """ Simpan konfigurasi billing langsung ke settings.json """
    with billing_lock:
        settings = _load_settings_raw()
        if 'billing' not in settings: settings['billing'] = {}
        settings['billing'].update(data)
        # Hapus field minimum pembayaran jika ada (Sesuai permintaan Boss)
        if 'min_payment_percentage' in settings['billing']:
            del settings['billing']['min_payment_percentage']
        return save_settings(settings)

SERVICE_NAME = os.environ.get('NMS_SERVICE', 'monitoring-wifi.service')

if not os.path.exists(PHOTO_DIR): os.makedirs(PHOTO_DIR)

# INITIALIZATION

# --- DISK SYNC HELPER ---
def force_disk_sync(file_obj=None, global_sync=False):
    """ 
    Sinkronisasi disk lintas platform (Windows/Linux).
    Gunakan global_sync=True hanya untuk operasi kritikal agar tidak membebani sistem.
    """
    try:
        if file_obj:
            file_obj.flush()
            os.fsync(file_obj.fileno())
    except: pass
    
    # Global sync for Linux/Unix (Hanya jika benar-benar perlu)
    if global_sync and hasattr(os, 'sync'):
        try: os.sync()
        except: pass

def _safe_write_json(path, data, critical=False):
    """
    Atomic write WITH fsync for disk integrity.
    """
    try:
        rid = random.randint(1000, 9999)
        tmp = f"{path}.tmp.{rid}"
        
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4)
            f.flush()
            try: os.fsync(f.fileno())
            except: pass
        
        if platform.system().lower() == 'windows':
            if os.path.exists(path): os.remove(path)
            shutil.move(tmp, path)
        else:
            os.rename(tmp, path)
            
        if critical:
            try: shutil.copy(path, path + '.stable')
            except: pass
            
        return True
    except Exception as e:
        if 'tmp' in locals() and os.path.exists(tmp):
            try: os.remove(tmp)
            except: pass
        return False
        print(f"[DISK ERROR] Gagal menyimpan {path}: {e}")
        # Hapus sisa file sementara jika gagal
        if 'tmp' in locals() and os.path.exists(tmp):
            try: os.remove(tmp)
            except: pass
        return False

# 1. Cek Library Ping (icmplib)
try:
    from icmplib import multiping as turbo_ping
    ICMP_READY = True
except ImportError:
    ICMP_READY = False
    print("[INIT] Ping Engine   : WARNING (icmplib not installed. Slow Mode active)")

# 2. Cek Library Mikrotik (routeros_api)
try:
    import routeros_api
except ImportError:
    routeros_api = None
    print("[INIT] Mikrotik API  : WARNING (routeros_api not installed)")

# 3. Cek Library Kompresi (flask_compress) - UTAMA UNTUK PERFORMA
try:
    from flask_compress import Compress
    COMPRESS_READY = True
except ImportError:
    COMPRESS_READY = False
    print("[INIT] Compression   : WARNING (flask-compress not installed. Standard Speed)")

# --- FLASK APP SETUP ---
app = Flask(__name__)
app.secret_key = 'peycell_nms_final_super_secret'
app.permanent_session_lifetime = timedelta(days=7)

# Aktifkan Kompresi (Jika Library Ada)
if COMPRESS_READY:
    Compress(app)
    app.config['COMPRESS_MIMETYPES'] = ['application/json', 'text/html', 'text/css', 'application/javascript']
    app.config['COMPRESS_LEVEL'] = 6
    app.config['COMPRESS_MIN_SIZE'] = 500

@app.context_processor
def inject_version():
    return dict(app_version=CURRENT_VERSION)
    
# --- GLOBAL STATE FOR STABILITY ---
STATUS_FAIL_TRACK = {} # {client_id: fail_count}

# --- NO CACHE ROUTE FOR LANG DICT (User Request) ---
@app.route('/static/js/lang_dict.js')
def serve_lang_dict_nocache():
    """Force no-cache for language dictionary"""
    # Use absolute path to ensure correct file serving
    js_dir = os.path.join(app.root_path, 'static', 'js')
    response = make_response(send_from_directory(js_dir, 'lang_dict.js'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# LICENSE SYSTEM
def is_licensed():
    """Check if system is licensed. Optimized: signature is cached, but file existence is real-time."""
    global _LICENSE_CACHE
    now = time.time()
    
    # --- REAL-TIME CHECK: Jika file dihapus, harus langsung tidak aktif ---
    if not os.path.exists(LICENSE_FILE): 
        _LICENSE_CACHE = {"valid": False, "info": {}, "expiry": 0}
        return False

    # Return from cache if verification is still fresh (300s = 5 mins)
    if _LICENSE_CACHE['expiry'] > now:
        return _LICENSE_CACHE['valid']

    mid = str(get_machine_id_cached()).strip().upper()
    
    # 1. Cek Blacklist Lokal (Cache)
    bl_clean = [str(x).strip().upper() for x in _cached_blacklist]
    if mid in bl_clean:
        _LICENSE_CACHE = {"valid": False, "info": {}, "expiry": now + 600}
        return False
        
    try:
        with open(LICENSE_FILE, 'r') as f:
            key = f.read().strip()
        valid, info = verify_license(key)
        
        # Cache the result of verification
        _LICENSE_CACHE = {
            "valid": valid,
            "info": info if valid else {},
            "expiry": now + 300 # 5 minutes cache for expensive verification
        }
        return valid
    except:
        return False

@app.before_request
def check_license_gate():
    """Middleware: Unlicensed = Read Only Mode"""
    # 1. Exempt static files
    if request.endpoint == 'static': return None
    
    # 2. Licensi Valid? (Optimasi: is_licensed() gunakan cache)
    if is_licensed(): return None

    # --- JIKA UNLICENSED (LOCKDOWN MODE) ---

    # Allow License Page & Activation
    if request.endpoint in ['license_page', 'activate_license']: return None

    # Block API calls (Return JSON 403)
    if request.path.startswith('/api/'):
        return jsonify({"status": "error", "msg": "LICENSE REQUIRED: System Locked. Please activate license."}), 403

    # Redirect all other pages (Dashboard, Maps, etc) to License Page
    return redirect(url_for('license_page'))

# --- LICENSE ROUTES ---
@app.route('/license')
def license_page():
    # RECOVERY: Jika user terdampar di sini (mungkin karena diblokir), 
    # paksa cek ulang ke cloud tanpa nunggu 10 menit.
    perform_blacklist_sync(force=True)
    
    # Ambil data license saat ini jika ada
    cur_data = None
    active = False
    if is_licensed():
        try:
            with open(LICENSE_FILE, 'r') as f:
                valid, info = verify_license(f.read().strip())
                if valid: 
                    cur_data = info
                    active = True
        except: pass
        
    return render_template('license.html', machine_id=get_machine_id_cached(), license_data=cur_data, active=active)

@app.route('/api/license/activate', methods=['POST'])
def activate_license():
    data = request.json or {}
    key = data.get('key', '').strip()
    
    valid, info = verify_license(key)
    if valid:
        # Simpan key ke file secara aman
        if _safe_replace_file(LICENSE_FILE, key.encode('utf-8')):
            return jsonify({"status": "ok", "msg": "Activated", "info": info})
        else:
            return jsonify({"status": "error", "msg": "Gagal menyimpan file lisensi"})
    else:
        return jsonify({"status": "error", "msg": info})

@app.route('/api/check_license')
def check_license_api():
    """API endpoint to check current license status for UI badges"""
    if not is_licensed():
        return jsonify({"active": False, "owner": "Trial Mode"})
    
    try:
        with open(LICENSE_FILE, 'r') as f:
            valid, info = verify_license(f.read().strip())
            if valid:
                return jsonify({
                    "active": True,
                    "owner": info.get('cli', 'Authorized User'),
                    "type": "LICENSED (PRO)"
                })
    except: pass
    
    return jsonify({"active": False, "owner": "Trial Mode"})
    
@app.route('/api/migrate_v2', methods=['POST'])
def api_migrate_v2():
    """Converts uploaded topology.json to SQLite database"""
    if 'file' not in request.files:
        return jsonify({"status": "error", "msg": "No file part"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"status": "error", "msg": "No selected file"}), 400
        
    try:
        # 1. Parse JSON
        raw_data = file.read().decode('utf-8')
        topology_data = json.loads(raw_data)
        
        # 2. Basic Validation
        if not isinstance(topology_data, dict) or 'server' not in topology_data:
            return jsonify({"status": "error", "msg": "Invalid topology.json format"}), 400
            
        # 3. Save to SQLite via DBManager
        success = db.save_full_topology(topology_data)
        
        if success:
            return jsonify({
                "status": "ok", 
                "msg": "Migrasi Berhasil! Data V2.9 telah dipindahkan ke SQLite.",
                "summary": {
                    "clients": len(topology_data.get('clients', [])),
                    "odps": len(topology_data.get('odps', [])),
                    "routers": len(topology_data.get('extra_routers', []))
                }
            })
        else:
            return jsonify({"status": "error", "msg": "Gagal menyimpan ke database SQLite"}), 500
            
    except Exception as e:
        return jsonify({"status": "error", "msg": f"Error: {str(e)}"}), 500



# ==============================================================================
#  KONFIGURASI OTOMATIS (AUTO-GENERATE)
# ==============================================================================
DEFAULT_CONFIG = {
    "admin_password": "admin",
    "viewer_password": "tamu",
    "telegram_bot_token": "",
    "telegram_chat_id": "",
    "service_name": "monitoring-wifi.service",
    "app_port": 5002
}

DEFAULT_SETTINGS = {
    "web_title": "MAPS MONITORING WIFI",
    "refresh_rate": 10,
    "wa_template": "Halo Kak *{name}*, tagihan internet bulan ini sudah terbit. Mohon pembayarannya ya. Terima kasih.",
    "wa_template_auto": "Yth. {name}, tagihan internet Anda sudah memasuki masa tenggang. Layanan akan dinonaktifkan otomatis dalam 3 hari jika belum ada pembayaran. Terima kasih.",
    "wa_template_payment": "Terima kasih, pembayaran wifi a.n *{name}* sebesar Rp {amount} pada {date} telah diterima. Layanan Anda tetap aktif hingga {expired}. Terima kasih.",
    "wa_template_isolir": "Yth. *{name}*, layanan internet Anda telah diisolir sementara karena keterlambatan pembayaran sebesar Rp {price}. Silakan lakukan pembayaran agar layanan kembali normal.",
    "wa_template_reactivate": "Halo *{name}*, pembayaran telah diterima dan layanan internet Anda telah diaktifkan kembali. Selamat berinternet!",
    "map_animation": True,
    "show_dashboard_clock": True,
    "dashboard_timezone": 7,
    "inventory": [],

    "billing": {
        "auto_isolir_enabled": True,
        "default_billing_day": 5,
        "grace_period_days": 3,
        "isolir_profile": "ISOLIR",
        "send_wa_notification": False,
        "billing_check_interval_hours": 24
    },
    "automation": {
        "backup": {
            "enabled": True,
            "schedule_time": "02:00",
            "keep_days": 7,
            "include_files": ["topology.db", "settings.json", "finance.json", "config.json", "license.key", "app.py", "db_manager.py", "license_utils.py", "wa-bridge.js", "package.json", "keygen.py", "CARA_PAKAI_KEYGEN.md", "DISTRIBUSI_KLIEN.md", "templates", "static"]
        },
        "telegram": {
            "enabled": False,
            "bot_token": "",
            "chat_id": "",
            "notifications": {
                "offline": True,
                "online": True,
                "isolir": True,
                "reactivate": True,
                "daily_report": False,
                "backup_report": True,
                "startup_report": True
            }
        }
    },
    "print_header": "TERIMA KASIH TELAH MENGGUNAKAN LAYANAN KAMI",
    "print_footer": "Simpan struk ini sebagai bukti pembayaran sah.",
    "print_store_name": "NMS PREMIUM WIFI",
    "print_paper": "58mm",
    "print_auto": False,
    "print_show_logo": True,
    "print_method": "browser",
    "print_template": """<div style="text-align:center; font-family:monospace;">
    <p style="margin:0; font-weight:bold; font-size:16px;">STRUK PEMBAYARAN</p>
    <p style="margin:5px 0;">{header}</p>
    <hr style="border-top:1px dashed #000;">
    <table style="width:100%; font-size:14px;">
        <tr><td style="text-align:left;">Tgl</td><td style="text-align:right;">{date}</td></tr>
        <tr><td style="text-align:left;">ID</td><td style="text-align:right;">{id}</td></tr>
        <tr><td style="text-align:left;">Nama</td><td style="text-align:right;">{name}</td></tr>
        <tr><td style="text-align:left;">Paket</td><td style="text-align:right;">{packet}</td></tr>
        <tr><td style="text-align:left; font-weight:bold;">TOTAL</td><td style="text-align:right; font-weight:bold;">Rp {amount}</td></tr>
    </table>
    <hr style="border-top:1px dashed #000;">
    <p style="margin:5px 0;">Berlaku S/D: <b>{expired}</b></p>
    <hr style="border-top:1px dashed #000;">
    <p style="margin:5px 0; font-size:12px;">{footer}</p>
</div>""",
    "print_template_text": """{store_name}
--------------------------------
Tanggal : {date}
Nama    : {name}
Paket   : {packet}
Harga   : {amount}
--------------------------------
Terima Kasih

{footer}"""
}

def deep_merge(dict1, dict2):
    """
    Recursive helper to merge dictionaries deep.
    """
    for key, value in dict2.items():
        if key in dict1 and isinstance(dict1[key], dict) and isinstance(value, dict):
            deep_merge(dict1[key], value)
        else:
            dict1[key] = value
    return dict1

def load_config():
  # : Add retry logic to bridge atomic swap windows
    retries = 3
    while retries > 0:
        if os.path.exists(CONFIG_FILE):
            break
        time.sleep(0.1)
        retries -= 1

    if not os.path.exists(CONFIG_FILE):
        # Coba restore dari .stable jika ada
        stable = CONFIG_FILE + '.stable'
        if os.path.exists(stable):
            try:
                with open(stable, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    _safe_write_json(CONFIG_FILE, data, critical=True)
                    return data
            except: pass
            
        print(f"[SYSTEM] Membuat file konfigurasi baru: {CONFIG_FILE}")
        _safe_write_json(CONFIG_FILE, DEFAULT_CONFIG, critical=True)
        return DEFAULT_CONFIG
    
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"[ERROR] Config rusak, mencoba recovery .stable: {e}")
        stable = CONFIG_FILE + '.stable'
        if os.path.exists(stable):
            try:
                with open(stable, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    _safe_write_json(CONFIG_FILE, data, critical=True)
                    return data
            except: pass
        return DEFAULT_CONFIG

def load_settings():
    """Simple cache-first loading without aggressive locking."""
    global _SETTINGS_CACHE
    now = time.time()
    
    if _SETTINGS_CACHE['data'] is not None and _SETTINGS_CACHE['expiry'] > now:
        return _SETTINGS_CACHE['data']
    
    settings_data = _load_settings_internal()
    _SETTINGS_CACHE = {"data": settings_data, "expiry": now + 30}
    return settings_data

def _load_settings_internal():
    if not os.path.exists(SETTINGS_FILE):
        _safe_write_json(SETTINGS_FILE, DEFAULT_SETTINGS)
        return DEFAULT_SETTINGS
    
    data = _parse_json_file_loose(SETTINGS_FILE, DEFAULT_SETTINGS)
    out = dict(DEFAULT_SETTINGS)
    changed = False
    if isinstance(data, dict):
        out.update(data)
    
    # Critical overrides
    try:
        out['refresh_rate'] = int(out.get('refresh_rate', 10))
    except: out['refresh_rate'] = 10
    out['map_animation'] = bool(out.get('map_animation', True))
    
    # Billing Config
    out['billing'] = load_billing_config()
    
    # Sync Config
    cfg_tmp = load_config()
    if 'automation' not in out: out['automation'] = {}
    if 'telegram' not in out['automation']: out['automation']['telegram'] = {}
    out['automation']['telegram']['bot_token'] = cfg_tmp.get('telegram_bot_token', '')
    out['automation']['telegram']['chat_id'] = cfg_tmp.get('telegram_chat_id', '')
    
    # Ensure new Telegram triggers exist
    tg_notifs = out['automation']['telegram'].setdefault('notifications', {})
    if 'isolir' not in tg_notifs: tg_notifs['isolir'] = True; changed = True
    if 'reactivate' not in tg_notifs: tg_notifs['reactivate'] = True; changed = True
    
    # Auto-enable Telegram if configured but disabled
    if out['automation']['telegram'].get('bot_token') and out['automation']['telegram'].get('chat_id'):
        if not out['automation']['telegram'].get('enabled', False):
            out['automation']['telegram']['enabled'] = True
            changed = True
    
    if changed:
        save_settings(out)
        
    return out

def save_settings(new_settings):
    try:
        current = _load_settings_raw()
        merged = dict(current)
        if isinstance(new_settings, dict):
            # Consolidate ALL settings using deep merge to prevent losing nested keys
            # EXCEPTION: Profiles mapping must be atomic replacements to allow deletions
            for k in ['billing_profiles', 'billing_profiles_meta']:
                if k in new_settings:
                    merged[k] = new_settings.pop(k)
                    
            deep_merge(merged, new_settings)

        try:
            merged['refresh_rate'] = int(merged.get('refresh_rate', 10))
        except Exception:
            merged['refresh_rate'] = 10
        merged['map_animation'] = bool(merged.get('map_animation', True))
        
        # Save ALL settings (Billing is now part of settings.json)
        core_settings = dict(merged)
        
        # BUGFIX: Exclude Infra Config from settings.json (Already in config.json)
        # Port & Service Name
        if 'app_port' in core_settings: del core_settings['app_port']
        if 'service_name' in core_settings: del core_settings['service_name']
        
        # Telegram Bot Token & Chat ID
        if 'automation' in core_settings and 'telegram' in core_settings['automation']:
            # Capture for config.json if they were updated in this payload
            changed_cfg = False
            cur_cfg = load_config()
            
            tg = core_settings['automation']['telegram']
            if 'bot_token' in tg:
                cur_cfg['telegram_bot_token'] = tg['bot_token']
                changed_cfg = True
            if 'chat_id' in tg:
                cur_cfg['telegram_chat_id'] = tg['chat_id']
                changed_cfg = True
                
            if changed_cfg:
                _safe_write_json(CONFIG_FILE, cur_cfg, critical=True)
                reload_config_globals()

        # Invalidate Settings Cache
        global _SETTINGS_CACHE
        _SETTINGS_CACHE = {"data": None, "expiry": 0}
        _safe_write_json(SETTINGS_FILE, core_settings, critical=True)
        
        # Reset billing tracking to allow immediate re-test of automation time
        global LAST_BILLING_NOTIF_DATE
        LAST_BILLING_NOTIF_DATE = None
        reload_config_globals()
        
        return merged
    except Exception:
        return load_settings()

def save_db(incoming, preserve_live=True):
    """
    Saves the full topology database while optionally preserving live status
    to prevent race conditions with monitoring threads.
    Optimized: Invalidates _TOPO_CACHE on write and uses robust string-based ID mapping.
    """
    try:
        if preserve_live:
            # Load current live data from DB to merge real-time fields (Source of Truth)
            current_db = db.load_full_topology()
            
            # Fields managed by background process (Not UI)
            live_fields = ['status', 'ping_ms', 'last_seen_ts', 'last_ping', 'last_isolir_notif_ts', 'paid_until']
            bill_live_fields = ['payment_status', 'original_profile', 'isolir_wa_sent', 'wa_sent_track', 'billing_day']

            # Create string-key lookup for robustness
            current_clients = {str(c['id']): c for c in current_db.get('clients', []) if 'id' in c}
            
            if 'clients' in incoming:
                for c in incoming['clients']:
                    orig_id = c.get('id')
                    if orig_id is None: continue
                    cid = str(orig_id)
                    
                    if cid in current_clients:
                        curr = current_clients[cid]
                        # 1. Base fields
                        for fld in live_fields:
                            if fld in curr: c[fld] = curr[fld]
                        
                        # 2. Nested billing fields
                        if 'billing' in curr and isinstance(curr['billing'], dict):
                            if 'billing' not in c: c['billing'] = {}
                            for bf in bill_live_fields:
                                if bf in curr['billing']:
                                    c['billing'][bf] = curr['billing'][bf]

        # Final Database Write (Incremental + Timeout handle)
        success = db.save_full_topology(incoming)
        
        if success:
            global _TOPO_CACHE
            _TOPO_CACHE = {"data": None, "expiry": 0} # Immediate cache invalidation
            
        return success
        
    except Exception as e:
        print(f"[DB ERROR] save_db critical failure: {e}")
        return False

def load_db(force_refresh=False):
    # Optimized Cache Logic
    global _TOPO_CACHE
    now = time.time()
    
    if not force_refresh and _TOPO_CACHE['data'] is not None and _TOPO_CACHE['expiry'] > now:
        return _TOPO_CACHE['data']

    # Retry logic for SQLite busy/locked
    retries = 5
    while retries > 0:
        try:
            data = db.load_full_topology()
            _TOPO_CACHE = {"data": data, "expiry": now + 5}
            return data
        except Exception as e:
            if "locked" in str(e).lower() or "busy" in str(e).lower():
                time.sleep(0.1)
                retries -= 1
                continue
            print(f"[DB ERROR] load_db failed: {e}")
            return {"server": {}, "odps": [], "clients": [], "extra_routers": []}
    return {"server": {}, "odps": [], "clients": [], "extra_routers": []}

def load_finance():
    if not os.path.exists(FINANCE_FILE):
        return {"transactions": []}
    try:
        with open(FINANCE_FILE, 'r') as f:
            return json.load(f)
    except:
        return {"transactions": []}

def save_finance(data):
    return _safe_write_json(FINANCE_FILE, data)

def _parse_json_file_loose(path, default_obj):
    try:
        if not os.path.exists(path):
            return default_obj
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            raw = f.read().strip()
        try:
            return json.loads(raw)
        except Exception:
            start = raw.find('{')
            end = raw.rfind('}')
            if start != -1 and end != -1 and end > start:
                return json.loads(raw[start:end+1])
            return default_obj
    except Exception:
        return default_obj

# Load Config
cfg = load_config()
ADMIN_PASSWORD = cfg.get('admin_password', 'admin')
PASSWORD_VIEWER = cfg.get('viewer_password', 'tamu')
TELEGRAM_BOT_TOKEN = cfg.get('telegram_bot_token', '')
TELEGRAM_CHAT_ID = cfg.get('telegram_chat_id', '')
SERVICE_NAME = cfg.get('service_name', os.environ.get('NMS_SERVICE', 'monitoring-wifi.service'))
APP_PORT = int(cfg.get('app_port', 5002))

print(f"[SYSTEM] MAPS MONITORING V{CURRENT_VERSION} STARTED")

_CONFIG_CACHE = {"data": None, "expiry": 0}

def reload_config_globals():
    global cfg, ADMIN_PASSWORD, PASSWORD_VIEWER, TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID, SERVICE_NAME, APP_PORT, _CONFIG_CACHE
    
    now = time.time()
    if _CONFIG_CACHE['data'] is not None and _CONFIG_CACHE['expiry'] > now:
        cfg = _CONFIG_CACHE['data']
    else:
        new_cfg = load_config()
        # Prevent wiping out passwords if disk read failed momentarily (Windows Atomic Swap)
        if new_cfg.get('admin_password') == 'admin' and cfg and cfg.get('admin_password') != 'admin':
            # Trust the old one for a bit, don't update cache.
            pass
        else:
            cfg = new_cfg
            _CONFIG_CACHE = {"data": cfg, "expiry": now + 5} # 5 second cache to bridge atomic renames
            
    ADMIN_PASSWORD = cfg.get('admin_password', 'admin')
    PASSWORD_VIEWER = cfg.get('viewer_password', 'tamu')
    TELEGRAM_BOT_TOKEN = cfg.get('telegram_bot_token', '')
    TELEGRAM_CHAT_ID = cfg.get('telegram_chat_id', '')
    SERVICE_NAME = cfg.get('service_name', 'monitoring-wifi.service')
    APP_PORT = int(cfg.get('app_port', 5002))

# --- CACHE ---
MK_RES = {} 
MK_CACHE = {} 

# DATABASE FUNCTIONS
def init_default_db():
    return {
        "server": {
            "id": "server_utama", "name": "SERVER UTAMA", 
            "coordinates": [-6.1754, 106.8272], # Default Monas, Jakarta
            "login": {"host":"", "user":"", "pass":"", "port":8728}, 
            "status": "online", "manual_wan": "", "ping_target": "8.8.8.8",
            "port_config": {"lan": "5", "sfp": "0"}
        },
        "extra_routers": [], "odps": [], "clients": []
    }

# --- TIMEZONE HELPERS ---
def get_local_now():
    """Returns server local time"""
    return datetime.now()

def add_log(name, status, msg):
  with log_lock:
    try:
        logs = []
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r') as f: logs = json.load(f)
        
        # Anti-Flood: Check if last log is same status and MESSAGE for same client
        if logs and logs[0].get('name') == name and logs[0].get('status') == status and logs[0].get('msg') == msg:
            return

        logs.insert(0, {"time": get_local_now().strftime("%Y-%m-%d %H:%M:%S"), "name": name, "status": status, "msg": msg})
        _safe_write_json(LOG_FILE, logs[:100])
    except: pass

def send_telegram_message(text, chat_id=None, token=None, reply_markup=None):
    """Send text message to Telegram with optional inline buttons (reply_markup)"""
    # 1. Try Settings First
    settings = load_settings()
    tg_conf = settings.get('automation', {}).get('telegram', {})
    
    # Use provided token or fallback to settings/env
    bot_token = token or tg_conf.get('bot_token') or TELEGRAM_BOT_TOKEN
    # If explicit chat_id is provided, use it (for responses). Otherwise use default (for alerts).
    target_id = chat_id or tg_conf.get('chat_id') or TELEGRAM_CHAT_ID
    
    if not bot_token or not target_id: return {"status":"skipped"}
    
    try:
        payload = {
            'chat_id': target_id, 
            'text': text, 
            'parse_mode': 'Markdown'
        }
        if reply_markup:
            payload['reply_markup'] = json.dumps(reply_markup)
            
        requests.post(f"https://api.telegram.org/bot{bot_token}/sendMessage", data=payload, timeout=10)
        return {"status": "ok"}
    except Exception as e: return {"status": "error", "msg": str(e)}

def dispatch_telegram_event(event_type, data):
    """
    Template-based notification dispatcher
    event_type: 'up', 'down', 'backup', 'startup'
    data: dict with placeholder values
    """
    settings = load_settings()
    tg_conf = settings.get('automation', {}).get('telegram', {})
    
    # Check if Globally Enabled
    if not tg_conf.get('enabled', False): return
    
    # Check Specific Trigger
    notifs = tg_conf.get('notifications', {})
    
    mapping = {
        'up': 'online',
        'down': 'offline',
        'backup': 'backup_report',
        'isolir': 'isolir',
        'reactivate': 'reactivate',
        'startup': 'startup_report'
    }
    
    key = mapping.get(event_type)
    if key and not notifs.get(key, False): return # Disabled by user preference
    
    # Prepare Template
    msg = ""
    dt = get_local_now()
    date_str = dt.strftime('%d-%m-%Y') # DD-MM-YYYY
    time_str = dt.strftime('%H:%M:%S')
    
    # --- AUTO CALCULATE TOTALS IF 0 ---
    if data.get('total_online', 0) == 0 and data.get('total_offline', 0) == 0:
        try:
            db_tmp = load_db()
            clients_tmp = db_tmp.get('clients', [])
            # status tracking
            t_on = sum(1 for c in clients_tmp if c.get('status') == 'online')
            t_iso = sum(1 for c in clients_tmp if c.get('status') == 'isolir')
            t_off = sum(1 for c in clients_tmp if c.get('status') == 'offline')
            data['total_online'] = t_on
            data['total_isolir'] = t_iso
            data['total_offline'] = t_off
        except Exception as e:
            print(f"[TG CALC ERROR] {e}")
    # ----------------------------------
    
    # --- CUSTOM TEMPLATE LOGIC ---
    tpl_up = tg_conf.get('template_up', '').strip()
    tpl_down = tg_conf.get('template_down', '').strip()
    
    # Prepare Template Data
    tpl_data = {
        'name': str(data.get('name', '-')),
        'ip': str(data.get('ip', '-')),
        'status': (event_type.upper() if event_type in ['isolir', 'reactivate'] else ('ONLINE' if event_type == 'up' else 'OFFLINE')),
        'date': date_str,
        'time': time_str,
        'total_online': str(data.get('total_online', 0)),
        'total_offline': str(data.get('total_offline', 0)),
        'total_isolir': str(data.get('total_isolir', 0)),
        'packet': str(data.get('packet', '-')),
        'mode': str(data.get('mode', 'PPPoE')).upper()
    }
    
    def apply_tpl(tpl, defaults):
        for k, v in defaults.items():
            tpl = tpl.replace('{' + k + '}', v)
        return tpl

    try:
        if event_type == 'up':
            if tpl_up:
                msg = apply_tpl(tpl_up, tpl_data)
            else:
                mode_label = str(data.get('mode', 'PPPoE')).upper()
                msg = f"✅ {mode_label} Connected\nTanggal: {date_str}\nJam: {time_str}\n"
                msg += f"User: {data.get('name', '-')}\n"
                msg += f"IP Client: {data.get('ip', '-')}\n"
                msg += f"Total Active: {data.get('total_online', 0)} Client\n"
                msg += f"Total Disconnected: {data.get('total_offline', 0)} Client\n"
                msg += f"Service: {data.get('packet', '-')}"
            
        elif event_type == 'down':
            if tpl_down:
                msg = apply_tpl(tpl_down, tpl_data)
            else:
                mode_label = str(data.get('mode', 'PPPoE')).upper()
                msg = f"❌ {mode_label} Disconnected\nTanggal: {date_str}\nJam: {time_str}\n"
                msg += f"User: {data.get('name', '-')}\n"
                msg += f"Total Active: {data.get('total_online', 0)} Client\n"
                msg += f"Total Disconnected: {data.get('total_offline', 0)} Client"
                
        elif event_type == 'isolir':
            tpl_iso = tg_conf.get('template_isolir', '').strip()
            if tpl_iso:
                msg = apply_tpl(tpl_iso, tpl_data)
            else:
                msg = f"🔒 Client ISOLATED\nTanggal: {date_str}\nJam: {time_str}\n"
                msg += f"User: {data.get('name', '-')}\n"
                msg += f"Packet: {data.get('packet', '-')}\n"
                msg += f"Total Isolir: {data.get('total_isolir', 0)}"

        elif event_type == 'reactivate':
            tpl_react = tg_conf.get('template_reactivate', '').strip()
            if tpl_react:
                msg = apply_tpl(tpl_react, tpl_data)
            else:
                msg = f"🔓 Client REACTIVATED\nTanggal: {date_str}\nJam: {time_str}\n"
                msg += f"User: {data.get('name', '-')}\n"
                msg += f"Packet: {data.get('packet', '-')}\n"
                msg += f"Total Active: {data.get('total_online', 0)}"
            
        elif event_type == 'backup':
            msg = f"💾 Auto Backup Success\nTanggal: {date_str}\nJam: {time_str}\n"
            msg += f"File: {data.get('filename', '-')}\n"
            msg += f"Size: {data.get('size', '-')} KB"
            
        elif event_type == 'startup':
            msg = f"⚙️ System Startup\nNMS Service Started\nTanggal: {date_str}\nJam: {time_str}"
            
        if msg:
            send_telegram_message(msg)
            
    except Exception as e:
        print(f"[TELEGRAM] Dispatch Error: {e}")

# --- TELEGRAM BOT TWO-WAY (PHASE 2: MULTI-BOT & COMMAND CONTROL) ---
def find_clients_smart(query):
    """Mencari klien dengan 'Smart Matching' (Case-insensitive & Partial)"""
    if not query: return []
    db_data = load_db()
    clients = db_data.get('clients', [])
    q = query.strip().lower()
    
    # Prioritaskan Exact Match (ID atau Nama)
    exact = [c for c in clients if q == str(c.get('id')).lower() or q == str(c.get('name', '')).lower()]
    if exact: return exact
    
    # Partial Match (Hanya yang mengandung kata tersebut)
    matches = [c for c in clients if q in str(c.get('name', '')).lower()]
    return matches

def find_router_for_client(client_id, db_data, cl_map=None, odp_map=None):
    """Helper to find which router a client belongs to by tracing parent_id (Optimized)"""
    clients = cl_map if cl_map is not None else {c['id']: c for c in db_data.get('clients', [])}
    odps = odp_map if odp_map is not None else {o['id']: o for o in db_data.get('odps', [])}
    routers = {r['id']: r for r in db_data.get('extra_routers', [])}
    if db_data.get('server'): routers['server_utama'] = db_data['server']
    
    curr_id = client_id
    for _ in range(5):
        node = clients.get(curr_id) or odps.get(curr_id)
        if not node: break
        parent_id = node.get('parent_id') or node.get('parent')
        if not parent_id: break
        if parent_id in routers: return parent_id
        curr_id = parent_id
    return 'server_utama'

def get_bot_finance_summary():
    """Helper to generate monthly finance summary text for Telegram"""
    try:
        data = load_finance()
        txs = data.get('transactions', [])
        now = get_local_now()
        cur_month = now.strftime('%Y-%m')
        income_pure = 0; expense_pure = 0
        cat_expenses = {}
        for t in txs:
            if t.get('date', '').startswith(cur_month):
                amt = int(t.get('amount', 0))
                t_type = t.get('type')
                t_cat = t.get('category', 'lain-lain')
                if t_cat == 'balance_carryover': continue
                if t_type == 'income': income_pure += amt
                elif t_type == 'expense':
                    expense_pure += amt
                    cat_expenses[t_cat] = cat_expenses.get(t_cat, 0) + amt
        cat_map = {"alat": "🛠️ Alat & Bahan", "rawat": "🔧 Pemeliharaan", "ops": "⛽ Operasional", "gaji": "👥 Gaji/Fee", "fee": "👥 Gaji/Fee", "lain-lain": "📦 Lain-lain"}
        sorted_cats = sorted(cat_expenses.items(), key=lambda x: x[1], reverse=True)[:3]
        msg = f"💰 *LAPORAN KEUANGAN ({get_month_name(now.month).upper()})*\n"
        msg += "━━━━━━━━━━━━━━━━━━\n"
        msg += f"➕ *Pemasukan* : `Rp {income_pure:,.0f}`\n"
        msg += f"➖ *Pengeluaran* : `Rp {expense_pure:,.0f}`\n"
        msg += "━━━━━━━━━━━━━━━━━━\n"
        msg += f"💳 *Saldo Bersih* : `Rp {income_pure - expense_pure:,.0f}`\n\n"
        if sorted_cats:
            msg += "📈 *Top Pengeluaran:*\n"
            for cat, val in sorted_cats:
                msg += f"• {cat_map.get(cat, cat.replace('_',' ').title())}: `Rp {val:,.0f}`\n"
        return msg
    except Exception as e: return f"❌ Gagal memproses laporan: {str(e)}"

def get_bot_active_users_summary():
    """Helper to generate active users count per router for Telegram (Optimized)"""
    try:
        db_data = load_db(); routers = []
        if db_data.get('server'): routers.append({'id': 'server_utama', 'name': db_data['server'].get('name', 'Pusat')})
        for r in db_data.get('extra_routers', []): routers.append({'id': r['id'], 'name': r.get('name', r['id'])})
        
        # Index nodes once for find_router_for_client speed
        cl_map = {c['id']: c for c in db_data.get('clients', [])}
        odp_map = {o['id']: o for o in db_data.get('odps', [])}
        
        stats = {r['id']: {'name': r['name'], 'pppoe_local': 0, 'pppoe_radius': 0, 'hotspot': 0, 'statik': 0} for r in routers}
        for rid in stats:
            r_data = MK_RES.get(rid, {})
            # Update name if identify is available (Fix Missing Name)
            if r_data.get('identity') and r_data['identity'] != '-':
                stats[rid]['name'] = r_data['identity']
            
            if not r_data.get('error'):
                # Counting with Radius Awareness
                for a in r_data.get('actives', []):
                    is_rad = (a.get('radius') == 'true' or a.get('radius') is True)
                    if is_rad: stats[rid]['pppoe_radius'] += 1
                    else: stats[rid]['pppoe_local'] += 1
                stats[rid]['hotspot'] = len(r_data.get('hotspot_actives', []))
        
        for c in db_data.get('clients', []):
            if c.get('status') == 'online':
                is_pppoe = str(c.get('credentials', {}).get('pppoe_user') or "").strip() != "" or str(c.get('service') or "").lower() == 'pppoe'
                is_hotspot = str(c.get('mode') or "").lower() == 'hotspot' or str(c.get('service') or "").lower() == 'hotspot'
                if not is_pppoe and not is_hotspot:
                    r_id = c.get('managed_by', 'server_utama')
                    if r_id in stats: stats[r_id]['statik'] += 1
        msg = "👥 *USER ONLINE PER MIKROTIK*\n"
        msg += "━━━━━━━━━━━━━━━━━━\n"
        total_global = 0
        for rid, item in stats.items():
            total_r = item['pppoe_local'] + item['pppoe_radius'] + item['hotspot'] + item['statik']
            total_global += total_r
            msg += f"📍 *{item['name']}*\n"
            if MK_RES.get(rid, {}).get('error'): msg += "   ⚠ `OFFLINE / DISCONNECTED` ❌\n"
            else:
                msg += f"   • PPPoE (Local)  : `{item['pppoe_local']}` User\n"
                msg += f"   • PPPoE (Radius) : `{item['pppoe_radius']}` User\n"
                msg += f"   • Hotspot        : `{item['hotspot']}` User\n"
                msg += f"   • Statik         : `{item['statik']}` Client\n"
            msg += "──────────────────\n"
        msg += f"🏆 *Total Global*: `{total_global}` User Online\n"
        msg += f"📋 *Total Terdaftar*: `{len(db_data.get('clients', []))}` Klien"
        return msg
    except Exception as e: return f"❌ Gagal menghitung user: {str(e)}"

def get_bot_offline_users_list():
    """Mengambil daftar klien yang statusnya offline/terputus"""
    try:
        db_data = load_db()
        # Filter clients who are definitely offline
        offline_clients = [c for c in db_data.get('clients', []) if c.get('status') == 'offline']
        
        if not offline_clients:
            return "✅ *SEMUA KLIEN TERHUBUNG*\nTidak ada klien yang terdeteksi offline saat ini."
            
        msg = f"🔴 *DAFTAR KLIEN OFFLINE ({len(offline_clients)})*\n"
        msg += "━━━━━━━━━━━━━━━━━━\n"
        
        # Limit biar tidak kepanjangan (Telegram max 4096 char)
        display_limit = 50
        for i, c in enumerate(offline_clients[:display_limit], 1):
            name = escape_telegram_markdown(c.get('name', 'Unknown'))
            msg += f"{i}. *{name}*\n"
            
        if len(offline_clients) > display_limit:
            msg += f"──────────────────\n"
            msg += f"dan `{len(offline_clients) - display_limit}` klien lainnya..."
            
        msg += "\n━━━━━━━━━━━━━━━━━━"
        return msg
    except Exception as e:
        return f"❌ Gagal mengambil daftar offline: {str(e)}"

def get_bot_isolated_users_list():
    """Mengambil daftar klien yang statusnya isolir/terkunci"""
    try:
        db_data = load_db()
        isolated_clients = [c for c in db_data.get('clients', []) if c.get('status') == 'isolir']
        
        if not isolated_clients:
            return "✅ *TIDAK ADA KLIEN TERISOLIR*\nSemua klien dalam kondisi lancar saat ini."
            
        msg = f"🟡 *DAFTAR KLIEN TERISOLIR ({len(isolated_clients)})*\n"
        msg += "━━━━━━━━━━━━━━━━━━\n"
        
        display_limit = 50
        for i, c in enumerate(isolated_clients[:display_limit], 1):
            name = escape_telegram_markdown(c.get('name', 'Unknown'))
            msg += f"{i}. *{name}*\n"
            
        if len(isolated_clients) > display_limit:
            msg += f"──────────────────\n"
            msg += f"dan `{len(isolated_clients) - display_limit}` klien lainnya..."
            
        msg += "\n━━━━━━━━━━━━━━━━━━"
        return msg
    except Exception as e:
        return f"❌ Gagal mengambil daftar isolir: {str(e)}"

def escape_telegram_markdown(text):
    """ Membersihkan karakter spesial Markdown agar pesan tidak ditolak Telegram API (HTTP 400)"""
    if not text: return ""
    return str(text).replace('*', '').replace('_', '').replace('`', '').replace('[', '').replace(']', '')

def handle_telegram_command(chat_id, text, bot_config=None, token=None, sender_name="Admin"):
    """
    Memproses perintah dari Telegram dengan dukungan Inline Buttons
    """
    if not text: return
    cmd_parts = text.split()
    if not cmd_parts: return
    
    cmd = cmd_parts[0].lower()
    args = cmd_parts[1:]
    
    # Check if command is enabled
    if bot_config and 'commands' in bot_config:
        allowed = bot_config['commands']
        clean_cmd = cmd.replace('/', '')
        if clean_cmd in allowed and not allowed[clean_cmd]: return
    
    if cmd == '/start':
        msg = f"✨ *TELEGRAM ASSISTANT NMS*\n"
        msg += "━━━━━━━━━━━━━━━━━━\n"
        msg += f"Halo *{sender_name}*! 👋 Saya adalah asisten cerdas sistem NMS Anda. Gunakan tombol atau ketik perintah di bawah:\n\n"
        msg += "📂 *MONITORING KLIEN*\n"
        msg += "🔹 `/cek [nama]` : Detail & Masa Aktif\n"
        msg += "🔹 `/bayar [nama]` : Proses Pembayaran\n"
        msg += "🔹 `/isolir [nama]` : Banned Akses Manual\n"
        msg += "🔹 `/buka [nama]` : Lepas Isolir Manual\n"
        msg += "🔹 `/bypass [nama]` : Dispensasi Isolation\n\n"
        msg += "💰 *FINANCE & LOGS*\n"
        msg += "🔹 `/laporan` : Rekap Bulanan\n"
        msg += "🔹 `/masuk` / `/keluar` : Catat Manual\n"
        msg += "🔹 `/log` : Aktivitas Terakhir\n\n"
        msg += "📊 *SISTEM*\n"
        msg += "🔹 `/status` : Kondisi MikroTik\n"
        msg += "🔹 `/users` : Daftar User Online\n"
        msg += "🔹 `/useroffline` : Daftar User Terputus\n"
        msg += "🔹 `/isolirlist` : Daftar User Terisolir\n"
        msg += "━━━━━━━━━━━━━━━━━━"
        
        # Premium Inline Buttons
        keyboard = [
            [{"text": "📊 Status", "callback_data": "/status"}, {"text": "👥 User Online", "callback_data": "/users"}],
            [{"text": "🔴 User Offline", "callback_data": "/useroffline"}, {"text": "📜 Log", "callback_data": "/log"}],
            [{"text": "💰 Laporan", "callback_data": "/laporan"}, {"text": "🔓 Bypass", "callback_data": "/bypass"}]
        ]
        send_telegram_message(msg, chat_id=chat_id, token=token, reply_markup={"inline_keyboard": keyboard})
        
    elif cmd == '/ping':
        send_telegram_message("🏓 *PONG!*\nSistem NMS merespon dengan cepat. ✅", chat_id=chat_id, token=token)

    elif cmd == '/laporan':
        summary = get_bot_finance_summary()
        send_telegram_message(summary, chat_id=chat_id, token=token)

    elif cmd == '/users':
        summary = get_bot_active_users_summary()
        send_telegram_message(summary, chat_id=chat_id, token=token)

    elif cmd == '/useroffline':
        summary = get_bot_offline_users_list()
        send_telegram_message(summary, chat_id=chat_id, token=token)

    elif cmd == '/isolirlist':
        summary = get_bot_isolated_users_list()
        send_telegram_message(summary, chat_id=chat_id, token=token)

    elif cmd == '/log':
        try:
            logs = []
            if os.path.exists(LOG_FILE):
                with open(LOG_FILE, 'r') as f: logs = json.load(f)
            
            if not logs:
                send_telegram_message("📜 *BELUM ADA LOG AKTIVITAS*", chat_id=chat_id, token=token)
            else:
                msg = "📜 *RIWAYAT AKTIVITAS TERAKHIR*\n"
                msg += "━━━━━━━━━━━━━━━━━━\n"
                for l in logs[:10]:
                    time_s = l.get('time', '').split(' ')[1] if ' ' in l.get('time','') else l.get('time','')
                    st = str(l.get('status', '')).lower()
                    icon = "✅" if st=='online' else "🔴" if st in ['offline','down'] else "🟡" if st=='isolir' else "⚙️"
                    msg += f"{icon} `[{time_s}]` *{l.get('name')}*: {l.get('msg')}\n"
                msg += "━━━━━━━━━━━━━━━━━━"
                send_telegram_message(msg, chat_id=chat_id, token=token)
        except: send_telegram_message("❌ Gagal mengambil log.", chat_id=chat_id, token=token)

    elif cmd in ['/isolir', '/buka']:
        is_isolir_cmd = (cmd == '/isolir')
        if not args:
            label = "isolir" if is_isolir_cmd else "buka akses"
            send_telegram_message(f"⚠️ *FORMAT SALAH*\nGunakan: `{cmd} [nama_klien]` untuk {label}.", chat_id=chat_id, token=token)
            return

        query = " ".join(args)
        safe_query = escape_telegram_markdown(query)
        matches = find_clients_smart(query)

        if not matches:
            send_telegram_message(f"🔍 *DATA TIDAK ADA*\nKlien *'{safe_query}'* tidak ditemukan.", chat_id=chat_id, token=token)
        elif len(matches) > 1:
            send_telegram_message(f"🤔 *HASIL GANDA*\nHarap ketik nama lebih spesifik.", chat_id=chat_id, token=token)
        else:
            client = matches[0]
            # Prepare for response
            st = load_settings()
            isolir_prof = st.get('billing', {}).get('isolir_profile', 'ISOLIR')
            label_act = isolir_prof if is_isolir_cmd else "BUKA AKSES"
            
            # Execute via CORE Functions
            if is_isolir_cmd:
                res = isolir_client_core(client['id'], processed_by=f"Bot Assistant ({sender_name})")
            else:
                # ENABLE WhatsApp Notification for /buka (Standard Profesional)
                res = reaktivasi_client_core(client['id'], send_notif=True) 
            
            if res.get('status') == 'ok':
                # Reload client data to get updated packet name
                db_data = load_db()
                client = next((c for c in db_data.get('clients', []) if c['id'] == client['id']), client)
                target_prof = client.get('packet_name', '-')
                
                msg = f"✅ *{label_act.upper()} BERHASIL*\n"
                msg += "━━━━━━━━━━━━━━━━━━\n"
                msg += f"👤 *Nama* : *{client['name']}*\n"
                msg += f"🏷️ *Profil* : `{target_prof}`\n"
                msg += "━━━━━━━━━━━━━━━━━━\n"
                msg += f"Oleh: *{sender_name}*"
                send_telegram_message(msg, chat_id=chat_id, token=token)

                # Broadcast Notification (Parity)
                d_alert = {
                    'name': client.get('name', '-'),
                    'ip': client.get('ip', '-'),
                    'packet': client.get('packet_name') or '-'
                }
                if not is_isolir_cmd:
                    dispatch_telegram_event('reactivate', d_alert)
                # isolir_client_core already dispatches the 'isolir' event
            else:
                send_telegram_message(f"❌ *GAGAL EKSEKUSI*\n{res.get('msg', 'Router tidak merespon atau data tidak valid.')}", chat_id=chat_id, token=token)


    elif cmd == '/bypass':
        if not args or len(args) < 2:
            send_telegram_message("⚠️ *FORMAT SALAH*\nGunakan: `/bypass [nama] [on/off]`\n\n*ON*: Klien tidak akan kena isolir otomatis.\n*OFF*: Kembali ke aturan normal.", chat_id=chat_id, token=token)
            return
            
        action = args[-1].lower()
        query = " ".join(args[:-1])
        if action not in ['on', 'off', 'true', 'false']:
            action = 'on' # Default if not specified clearly
            query = " ".join(args)
            
        matches = find_clients_smart(query)
        if not matches:
            send_telegram_message(f"🔍 *KLIEN TIDAK DITEMUKAN*\nKlien *'{query}'* tidak ada dalam database.", chat_id=chat_id, token=token)
        elif len(matches) > 1:
            send_telegram_message(f"🤔 *HASIL GANDA*\nHarap ketik nama lebih spesifik.", chat_id=chat_id, token=token)
        else:
            client = matches[0]
            is_on = action in ['on', 'true']
            
            db_data = load_db()
            for c in db_data.get('clients', []):
                if c['id'] == client['id']:
                    c['bypass_billing'] = is_on
                    break
            save_db(db_data, preserve_live=False)
            
            status_desc = "AKTIF (Dispensasi)" if is_on else "NONAKTIF (Normal)"
            icon = "🔓" if is_on else "🔒"
            
            msg = f"{icon} *STATUS BYPASS DIPERBARUI*\n"
            msg += "━━━━━━━━━━━━━━━━━━\n"
            msg += f"👤 *Nama* : *{client['name']}*\n"
            msg += f"🛡️ *Bypass* : `{status_desc}`\n"
            msg += "━━━━━━━━━━━━━━━━━━\n"
            if is_on:
                msg += "_Klien ini tidak akan diisolir otomatis oleh sistem._"
            else:
                msg += "_Klien ini kembali ke aturan isolir normal._"
            
            send_telegram_message(msg, chat_id=chat_id, token=token)
            add_log(client['name'], 'system', f"Bypass billing set to {is_on} oleh {sender_name} via Bot")

    elif cmd == '/cek':
        if not args:
            send_telegram_message("⚠️ *FORMAT SALAH*\nGunakan perintah: `/cek [nama_klien]`", chat_id=chat_id, token=token)
            return
            
        query = " ".join(args)
        safe_query = escape_telegram_markdown(query)
        matches = find_clients_smart(query)
        
        if not matches:
            send_telegram_message(f"🔍 *DATA TIDAK DITEMUKAN*\nKlien *'{safe_query}'* tidak terdaftar di database.", chat_id=chat_id, token=token)
        elif len(matches) > 5:
            send_telegram_message(f"⚠️ *TERLALU BANYAK HASIL*\nDitemukan {len(matches)} klien untuk *'{safe_query}'*. Harap ketik nama lebih spesifik.", chat_id=chat_id, token=token)
        elif len(matches) > 1:
            msg = f"🤔 *HASIL PENCARIAN ({len(matches)})*\n"
            msg += "━━━━━━━━━━━━━━━━━━\n"
            for i, c in enumerate(matches, 1):
                msg += f"{i}. *{c.get('name')}*\n   (ID: {str(c.get('id')).replace('client_','')})\n"
            msg += "━━━━━━━━━━━━━━━━━━\n"
            msg += "Ketik `/cek [Nama Lengkap]` atau `/cek [ID]`"
            send_telegram_message(msg, chat_id=chat_id, token=token)
        else:
            # Ketemu 1 Klien
            c = matches[0]
            status_icon = "✅" if c.get('status') == 'online' else "❌"
            if c.get('status') == 'isolir': status_icon = "🚫"
            
            p_name = c.get('packet_name', '-')
            
            # Kriteria LUNAS: Masa aktif harus melewati bulan berjalan
            now = get_local_now()
            next_month_start = (now.replace(day=1) + timedelta(days=32)).replace(day=1).date()
            paid_until_str = c.get('paid_until')
            
            is_unpaid = True
            if paid_until_str:
                try:
                    expiry_date = datetime.strptime(paid_until_str, '%Y-%m-%d').date()
                    if expiry_date >= next_month_start:
                        is_unpaid = False
                except: pass
            
            status_text = "(✅ Lunas)" if not is_unpaid else "(⚠️ Belum Bayar)"
            
            # Smart Check: Bypass Status
            is_bypassed = c.get('bypass_billing', False)
            bypass_icon = "🔓" if is_bypassed else "🔒"
            bypass_text = " (Dispensasi Aktif)" if is_bypassed else ""
            
            # Hitung Harga Paket, Tunggakan & Total
            settings_data = load_settings()
            b_profs = settings_data.get('billing_profiles', {})
            price_val = 0
            for prof_name, prof_price in b_profs.items():
                if prof_name.strip().lower() == p_name.strip().lower():
                    price_val = int(prof_price); break
                    
            if price_val == 0 and p_name:
                match = re.search(r'(\d+)rb', p_name, re.I) or re.search(r'(\d+)k', p_name, re.I)
                if match: price_val = int(match.group(1)) * 1000

            ma_list = settings_data.get('manual_arrears', [])
            arrears_val = 0
            c_name_norm = str(c.get('name', '')).strip().upper()
            for ma in ma_list:
                if (ma.get('client_name') or "").strip().upper() == c_name_norm:
                    arrears_val += int(ma.get('amount') or 0)
            
            total_bill_val = arrears_val
            if is_unpaid:
                total_bill_val += price_val

            price_str = f"Rp {price_val:,.0f}" if price_val > 0 else "-"
            arrears_str = f"Rp {arrears_val:,.0f}" if arrears_val > 0 else "Rp 0"
            total_str = f"Rp {total_bill_val:,.0f}" if total_bill_val > 0 else "Rp 0"
            
            msg = f"👤 *DETAIL KLIEN: {c.get('name')}*\n"
            msg += f"━━━━━━━━━━━━━━━━━━\n"
            msg += f"📍 *Status* : {status_icon} `{str(c.get('status', 'offline')).upper()}`\n"
            msg += f"💰 *Pembayaran* : *{status_text.replace('(','').replace(')','')}*\n"
            msg += f"{bypass_icon} *Bypass/Dispensasi* : *{'AKTIF' if is_bypassed else 'TIDAK'}*\n"
            msg += f"🌐 *IP Address* : `{c.get('ip', '-')}`\n"
            msg += f"📦 *Profil Paket* : `{p_name}`\n"
            msg += f"💵 *Harga Paket* : `{price_str}`\n"
            msg += f"🏷️ *Tunggakan* : `{arrears_str}`\n"
            msg += f"🧾 *Total Tagihan* : `{total_str}`\n"
            msg += f"📅 *Masa Aktif* : *{c.get('paid_until', 'N/A')}*\n"
            msg += f"━━━━━━━━━━━━━━━━━━\n"
            # Smart button for detail
            kb = [[{"text": "💸 Bayar Sekarang", "callback_data": f"/bayar {c['name']}"}]]
            if is_bypassed:
                kb.append([{"text": "🔒 Matikan Bypass", "callback_data": f"/bypass {c['name']} off"}])
            else:
                kb.append([{"text": "🔓 Aktifkan Bypass", "callback_data": f"/bypass {c['name']} on"}])
                
            send_telegram_message(msg, chat_id=chat_id, token=token, reply_markup={"inline_keyboard": kb})

    elif cmd == '/bayar':
        if not args:
            send_telegram_message("⚠️ *FORMAT SALAH*\nGunakan: `/bayar [nama]` untuk cek harga/ID.", chat_id=chat_id, token=token)
            return
        
        # Cerdas: Cek apakah argumen terakhir adalah angka (nominal)
        has_amount = False
        amount = 0
        query = ""
        
        try:
            # Peningkatan deteksi nominal: Cek jika argumen terakhir adalah murni angka
            last_arg = args[-1].replace('.', '').replace(',', '')
            if last_arg.isdigit() and len(last_arg) >= 3:
                amount = int(last_arg)
                query = " ".join(args[:-1])
                has_amount = True
            else:
                query = " ".join(args)
        except:
            query = " ".join(args)
 
        if not query:
            send_telegram_message("⚠️ *NAMA KLIEN KOSONG*\nGunakan perintah: `/bayar [nama/ID] [nominal]`", chat_id=chat_id, token=token)
            return
 
        safe_query = escape_telegram_markdown(query)
        matches = find_clients_smart(query)
        if not matches:
            send_telegram_message(f"🔍 *DATA TIDAK DITEMUKAN*\nKlien *'{safe_query}'* tidak terdaftar di database.", chat_id=chat_id, token=token)
        elif len(matches) > 1:
            msg = f"🤔 *HASIL PENCARIAN ({len(matches)})*\n"
            msg += "━━━━━━━━━━━━━━━━━━\n"
            for i, c in enumerate(matches, 1):
                msg += f"{i}. *{c.get('name')}*\n   (ID: {str(c.get('id')).replace('client_','')})\n"
            msg += "━━━━━━━━━━━━━━━━━━\n"
            msg += "Contoh: `/bayar NamaLengkap [Nominal]`"
            send_telegram_message(msg, chat_id=chat_id, token=token)
        elif not has_amount:
            # Ketemu 1 Klien tapi belum input nominal
            client = matches[0]
            p_name = client.get('packet_name') or client.get('service_plan', 'Standard')
            
            # Cari harga paket di settings
            st = load_settings()
            b_profs = st.get('billing_profiles', {})
            price_val = b_profs.get(p_name, 0)
            
            msg = "📋 *INFO TAGIHAN*\n"
            msg += "━━━━━━━━━━━━━━━━━━\n"
            msg += f"👤 *Nama* : *{client['name']}*\n"
            msg += f"📦 *Paket* : `{p_name}`\n"
            msg += f"💰 *Harga Bulanan* : `Rp {price_val:,.0f}`\n"
            msg += "━━━━━━━━━━━━━━━━━━\n"
            msg += f"Gunakan perintah di bawah untuk bayar:\n"
            msg += f"`/bayar {client['name']} {price_val}`"
            # Button for quick action
            kb = [[{"text": f"✅ Bayar Rp {price_val:,.0f}", "callback_data": f"/bayar {client['name']} {price_val}"}]]
            send_telegram_message(msg, chat_id=chat_id, token=token, reply_markup={"inline_keyboard": kb})
        else:
            # Ketemu 1 Klien
            client = matches[0]
            res = bot_execute_payment(client['id'], amount, processed_by=sender_name)
            
            if res.get('status') == 'ok':
                dur_txt = f" ({res.get('duration')} Bulan)" if res.get('duration', 1) > 1 else ""
                msg = f"✅ *PEMBAYARAN BERHASIL{dur_txt}*\n"
                msg += "━━━━━━━━━━━━━━━━━━\n"
                msg += f"👤 *Nama* : *{client['name']}*\n"
                msg += f"💰 *Nominal* : `Rp {amount:,.0f}`\n"
                msg += f"📅 *Masa Aktif Baru* : *{res.get('new_expiry')}*\n"
                msg += f"📍 *Status* : `{'ONLINE' if res.get('was_reactivated') else 'AKTIF'}` ✅\n"
                msg += "━━━━━━━━━━━━━━━━━━\n"
                msg += f"Oleh: *{sender_name}*"
                send_telegram_message(msg, chat_id=chat_id, token=token)
            else:
                send_telegram_message(f"❌ *GAGAL PROSES*\n{res.get('msg', 'Terjadi kesalahan internal.')}", chat_id=chat_id, token=token)
 
    elif cmd == '/batal':
        if not args:
            send_telegram_message("⚠️ *FORMAT SALAH*\nGunakan: `/batal [nama_klien]` untuk membatalkan pembayaran terakhir.", chat_id=chat_id, token=token)
            return
        
        query = " ".join(args)
        safe_query = escape_telegram_markdown(query)
        matches = find_clients_smart(query)
        if not matches:
            send_telegram_message(f"🔍 *KLIEN TIDAK DITEMUKAN*\nKlien *'{safe_query}'* tidak ada.", chat_id=chat_id, token=token)
        elif len(matches) > 1:
            send_telegram_message("🤔 *HASIL GANDA*\nHarap ketik nama klien lebih spesifik agar tidak salah hapus.", chat_id=chat_id, token=token)
        else:
            client = matches[0]
            res = bot_cancel_last_payment(client['id'], sender_name)
            if res['status'] == 'ok':
                msg = "🗑️ *PEMBAYARAN DIBATALKAN*\n"
                msg += "━━━━━━━━━━━━━━━━━━\n"
                msg += f"👤 *Klien* : *{client['name']}*\n"
                msg += f"💰 *Nominal* : `Rp {res['amount']:,.0f}`\n"
                msg += f"📅 *Masa Aktif Mundur Ke* : `{res['new_expiry']}`\n"
                msg += "━━━━━━━━━━━━━━━━━━\n"
                msg += f"Dibatalkan oleh: *{sender_name}*"
                send_telegram_message(msg, chat_id=chat_id, token=token)
            else:
                send_telegram_message(f"❌ *GAGAL BATALKAN*\n{res['msg']}", chat_id=chat_id, token=token)
 
    elif cmd in ['/masuk', '/keluar']:
        tx_type = "income" if cmd == '/masuk' else "expense"
        if len(args) < 3:
            label = "pemasukan" if tx_type=="income" else "pengeluaran"
            msg = f"⚠️ *FORMAT /{cmd.replace('/','')}:*\n"
            msg += f"`{cmd} [kategori] [nominal] [keterangan]`\n\n"
            msg += "*Kategori:*\n"
            if tx_type=="income": msg += "• `investasi`, `hibah`, `lain-lain`"
            else: msg += "• `alat`, `rawat`, `ops`, `gaji`, `lain-lain`"
            send_telegram_message(msg, chat_id=chat_id, token=token)
            return
        
        cat_input = args[0].lower(); raw_amt = re.sub(r'\D', '', args[1])
        try:
            amt = int(raw_amt); note = " ".join(args[2:])
            res = bot_add_manual_tx(tx_type, cat_input, amt, note, sender_name)
            if res['status'] == 'ok':
                cat_display_map = {"investasi": "Investasi", "hibah": "Hibah", "alat": "Alat & Bahan", "rawat": "Pemeliharaan", "ops": "Operasional", "gaji": "Gaji / Fee", "lain-lain": "Lain-lain"}
                cat_label = cat_display_map.get(cat_input.lower(), "Lain-lain")
                icon = "💰" if tx_type == "income" else "💸"
                label = "PEMASUKAN" if tx_type == "income" else "PENGELUARAN"
                msg = f"{icon} *{label} TERCATAT*\n"
                msg += "━━━━━━━━━━━━━━━━━━\n"
                msg += f"💰 *Nominal* : `Rp {amt:,.0f}`\n"
                msg += f"🏷️ *Kategori* : {cat_label}\n"
                msg += f"📝 *Ket* : {note}\n"
                msg += "━━━━━━━━━━━━━━━━━━\n"
                msg += f"Oleh: *{sender_name}*"
                send_telegram_message(msg, chat_id=chat_id, token=token)
            else: send_telegram_message(f"❌ *GAGAL CATAT*\n{res['msg']}", chat_id=chat_id, token=token)
        except: send_telegram_message("❌ *ERROR*\nPastikan nominal adalah angka.", chat_id=chat_id, token=token)
            
    elif cmd == '/status':
        try:
            stats = get_system_stats_cached()
            msg = "🖥️ *STATUS SISTEM NMS*\n"
            msg += f"━━━━━━━━━━━━━━━━━━\n"
            msg += f"⚡ *CPU Usage* : `{stats.get('cpu_usage', 0)}%`\n"
            msg += f"📟 *RAM Usage* : `{stats.get('ram_usage', 0)}%`\n"
            msg += f"💾 *Disk Space* : `{stats.get('disk_usage', 0)}%`\n"
            msg += f"⏱️ *System Uptime* : {get_system_uptime()}\n"
            msg += f"━━━━━━━━━━━━━━━━━━\n"
            
            db_snap = load_db(); routers = []
            if db_snap.get('server'): routers.append(('server_utama', db_snap['server']))
            for r in db_snap.get('extra_routers', []): routers.append((r['id'], r))
            
            if routers:
                msg += "\n📡 *STATUS MIKROTIK*\n"
                msg += "━━━━━━━━━━━━━━━━━━\n"
                for i, (rid, rcfg) in enumerate(routers, 1):
                    r_data = MK_RES.get(rid, {})
                    is_online = rcfg.get('status') == 'online'
                    status_icon = "🟢" if is_online else "🔴"
                    identify = r_data.get('identity', rcfg.get('name', 'Unknown'))
                    msg += f"{i}. *{identify}*\n"
                    msg += f"   📍 Status : {status_icon} `{str(rcfg.get('status', 'offline')).upper()}`\n"
                    if is_online:
                        msg += f"   ⚡ CPU : `{r_data.get('cpu', 0)}%`\n"
                        msg += f"   ⏱️ Uptime : `{r_data.get('uptime', '-')}`\n"
                    msg += "──────────────────\n"
            send_telegram_message(msg, chat_id=chat_id, token=token)
        except: send_telegram_message("❌ Gagal mengambil status sistem.", chat_id=chat_id, token=token)

def telegram_listener_loop():
    """Refactored Listener Loop to support Multi-Bot and command control"""
    global _TG_UPDATE_STATES
    print("[TELEGRAM] Multi-Bot Listener Started")
    
    while True:
        try:
            settings = load_settings()
            automation = settings.get('automation', {})
            
            # --- 1. COLLECT ALL ACTIVE BOTS ---
            active_bots = {} # { token: { "chats": { chat_id: bot_config } } }
            
            # B. Telegram Assistant Bots (Multi-Bot)
            tg_assist = automation.get('tg_assist', {})
            if tg_assist.get('enabled', False):
                raw_bots = tg_assist.get('bots', [])
                assist_cmds = tg_assist.get('commands', { "cek": True, "status": True, "bayar": True })
                
                for entry in raw_bots:
                    if ':' in entry:
                        parts = entry.rsplit(':', 1)
                        if len(parts) == 2:
                            token = parts[0].strip()
                            chat_id = parts[1].strip()
                            
                            if token and chat_id:
                                if token not in active_bots: active_bots[token] = {"chats": {}}
                                active_bots[token]["chats"][chat_id] = {"commands": assist_cmds}

            if not active_bots:
                time.sleep(15)
                continue
                
            # --- 2. POLL EACH UNIQUE TOKEN ---
            for token, bot_data in active_bots.items():
                try:
                    last_id = _TG_UPDATE_STATES.get(token, 0)
                    url = f"https://api.telegram.org/bot{token}/getUpdates"
                    params = {'offset': last_id + 1, 'timeout': 20}
                    
                    resp = requests.get(url, params=params, timeout=25)
                    if resp.status_code == 200:
                        data = resp.json()
                        if data.get('ok'):
                            results = data.get('result', [])
                            for update in results:
                                _TG_UPDATE_STATES[token] = update['update_id']
                                
                                msg_obj = update.get('message')
                                cb_obj = update.get('callback_query')
                                if not msg_obj and not cb_obj: continue
                                
                                if msg_obj:
                                    chat_id = str(msg_obj.get('chat', {}).get('id'))
                                    text = msg_obj.get('text')
                                    sender = msg_obj.get('from', {})
                                else:
                                    chat_id = str(cb_obj.get('message', {}).get('chat', {}).get('id'))
                                    text = cb_obj.get('data')
                                    sender = cb_obj.get('from', {})
                                    try: requests.post(f"https://api.telegram.org/bot{token}/answerCallbackQuery", data={'callback_query_id': cb_obj['id']}, timeout=5)
                                    except: pass
                                
                                # Check if sender is authorized for THIS token
                                allowed_chats = bot_data.get("chats", {})
                                if chat_id in allowed_chats:
                                    f_name = sender.get('first_name', '')
                                    l_name = sender.get('last_name', '')
                                    s_name = (f_name + " " + l_name).strip() or "Admin"
                  # : Panggil fungsi handle secara Asynchronous menggunakan Thread 
                                    # Agar proses koneksi Mikrotik yang lambat TIDAK memblokir (nge-lag-in) chat yang lain
                                    threading.Thread(target=handle_telegram_command, args=(chat_id, text), kwargs={"bot_config": allowed_chats[chat_id], "token": token, "sender_name": s_name}, daemon=True).start()
                                else:
                                    if text and text.startswith('/'):
                                        print(f"[TELEGRAM] Unauthorized {text} from {chat_id} on Bot {token[:8]}...")
                    elif resp.status_code == 401:
                        print(f"[TELEGRAM] Invalid Token detected: {token[:8]}...")
                        time.sleep(2)
                except Exception as ex:
                    print(f"[TELEGRAM] Bot Polling Error ({token[:8]}): {ex}")
                    time.sleep(2)

        except Exception as e:
            print(f"[TELEGRAM] Global Listener Error: {e}")
            time.sleep(5)
        
        time.sleep(0.5)

def send_telegram_file(filepath, caption, token=None):
    settings = load_settings()
    tg_conf = settings.get('automation', {}).get('telegram', {})
    
    bot_token = token or tg_conf.get('bot_token') or TELEGRAM_BOT_TOKEN
    chat_id = tg_conf.get('chat_id') or TELEGRAM_CHAT_ID

    if not bot_token or not chat_id: 
        return {"status":"error", "msg": "Token atau Chat ID kosong"}
    try:
        with open(filepath, 'rb') as f:
            response = requests.post(
                f"https://api.telegram.org/bot{bot_token}/sendDocument", 
                data={'chat_id': chat_id, 'caption': caption}, 
                files={'document': f}, 
                timeout=30
            )
        
        if response.status_code == 200:
            result = response.json()
            if result.get('ok'):
                return {"status": "ok", "msg": "File berhasil dikirim ke Telegram"}
            else:
                return {"status": "error", "msg": f"Telegram API Error: {result.get('description', 'Unknown')}"}
        else:
            return {"status": "error", "msg": f"HTTP {response.status_code}: {response.text[:200]}"}
    except Exception as e: 
        return {"status": "error", "msg": f"Exception: {str(e)}"}

# ==============================================================================
#  BILLING HELPER FUNCTIONS
# ==============================================================================
def get_router_data(router_id, db=None):
    """ Ambil kredensial router berdasarkan ID """
    if db is None:
        db = load_db()
    
    if router_id == 'server_utama':
        return load_db().get("server", {})
    else:
        return next((r for r in load_db().get("extra_routers", []) if r['id'] == router_id), None)

def ensure_isolir_profile(router_id):
    """ Buat profil ISOLIR jika belum ada """
    conn = None
    try:
        router_data = get_router_data(router_id)
        if not router_data or not router_data.get('login', {}).get('host'):
            return False
        
        settings = load_settings()
        isolir_prof_name = settings.get('billing', {}).get('isolir_profile', 'ISOLIR')
        
        login = router_data['login']
        conn = routeros_api.RouterOsApiPool(
            login['host'], 
            username=login['user'], 
            password=login['pass'],
            port=int(login.get('port', 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        
        # Check if profile exists
        profiles = api.get_resource('/ppp/profile').get()
        isolir_exists = any(p.get('name') == isolir_prof_name for p in profiles)
        
        if not isolir_exists:
            # Create profile with 1kbps bandwidth
            api.get_resource('/ppp/profile').add(
                name=isolir_prof_name,
                **{'rate-limit': '1k/1k'}
            )
            print(f"[BILLING] Created profile {isolir_prof_name} on {router_id}")
        
        return True
    except Exception as e:
        print(f"[ERROR] Failed to ensure ISOLIR profile on {router_id}: {e}")
        return False
    finally:
        if conn: conn.disconnect()

def get_pppoe_current_profile(username, router_id):
    """ Cek profil user PPPoE saat ini dari Mikrotik """
    conn = None
    try:
        router_data = get_router_data(router_id)
        if not router_data or not router_data.get('login', {}).get('host'):
            return None
        
        login = router_data['login']
        conn = routeros_api.RouterOsApiPool(
            login['host'], 
            username=login['user'], 
            password=login['pass'],
            port=int(login.get('port', 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        secrets = api.get_resource('/ppp/secret')
        user = secrets.get(name=username)
        
        if user and len(user) > 0:
            profile = user[0].get('profile', 'default')
            return profile
        return None
    except Exception as e:
        print(f"[ERROR] get_pppoe_current_profile: {e}")
        return None
    finally:
        if conn: conn.disconnect()

def change_pppoe_profile(username, new_profile, router_id):
    """ Ganti profil secret PPPoE """
    conn = None
    try:
        router_data = get_router_data(router_id)
        if not router_data or not router_data.get('login', {}).get('host'):
            return {"status": "error", "msg": "Router not found"}
        
        login = router_data['login']
        conn = routeros_api.RouterOsApiPool(
            login['host'], 
            username=login['user'], 
            password=login['pass'],
            port=int(login.get('port', 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        
        secrets = api.get_resource('/ppp/secret')
        user = secrets.get(name=username)
        
        # Fallback: Case-insensitive search
        if not user:
            all_secrets = secrets.get()
            user = [s for s in all_secrets if s.get('name', '').lower() == username.lower()]

        if user:
            # Use .id or id for compatibility
            target_id = user[0].get('.id') or user[0].get('id')
            secrets.set(id=target_id, profile=new_profile)
            return {"status": "ok"}
        else:
            return {"status": "error", "msg": f"User '{username}' no match in MikroTik"}
    except Exception as e:
        return {"status": "error", "msg": str(e)}
    finally:
        if conn: conn.disconnect()

def kick_pppoe_user(username, router_id):
    """ Kick user PPPoE dari active connections """
    conn = None
    try:
        router_data = get_router_data(router_id)
        if not router_data or not router_data.get('login', {}).get('host'):
            return {"status": "error", "msg": "Router not found"}
        
        login = router_data['login']
        conn = routeros_api.RouterOsApiPool(
            login['host'], 
            username=login['user'], 
            password=login['pass'],
            port=int(login.get('port', 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        
        actives = api.get_resource('/ppp/active')
        session = actives.get(name=username)
        
        if session:
            # Use .id or id for compatibility
            target_id = session[0].get('.id') or session[0].get('id')
            actives.remove(id=target_id)
        
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "msg": str(e)}
    finally:
        if conn: conn.disconnect()

def add_to_address_list(address, list_name, router_id, comment=""):
    """ Tambahkan entry ke IP Firewall Address-list """
    conn = None
    try:
        router_data = get_router_data(router_id)
        if not router_data or not router_data.get('login', {}).get('host'):
            return {"status": "error", "msg": "Router not found"}
        
        login = router_data['login']
        conn = routeros_api.RouterOsApiPool(
            login['host'], 
            username=login['user'], 
            password=login['pass'],
            port=int(login.get('port', 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        
        resource = api.get_resource('/ip/firewall/address-list')
        
        # SMART CLEANUP: If comment is provided, remove old entries with SAME comment
        # This handles IP changes (removes old IP block, adds new IP block)
        if comment:
            old_entries = resource.get(list=list_name, comment=comment)
            for entry in old_entries:
                target_id = entry.get('.id') or entry.get('id')
                resource.remove(id=target_id)

        # Double check by address to avoid duplicates
        existing = resource.get(address=address, list=list_name)
        if not existing:
            resource.add(address=address, list=list_name, comment=comment)
        
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "msg": str(e)}
    finally:
        if conn: conn.disconnect()

def remove_from_address_list(address, list_name, router_id, comment=None):
    """ Hapus entry dari IP Firewall Address-list """
    conn = None
    try:
        router_data = get_router_data(router_id)
        if not router_data or not router_data.get('login', {}).get('host'):
            return {"status": "error", "msg": "Router not found"}
        
        login = router_data['login']
        conn = routeros_api.RouterOsApiPool(
            login['host'], 
            username=login['user'], 
            password=login['pass'],
            port=int(login.get('port', 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        
        resource = api.get_resource('/ip/firewall/address-list')
        
        # Cleanup by Comment (Name) - Precise
        if comment:
            entries = resource.get(list=list_name, comment=comment)
            for e in entries:
                target_id = e.get('.id') or e.get('id')
                resource.remove(id=target_id)
        
        # Cleanup by specific IP - Aggressive (Remove from ANY list to ensure activation)
        if address and address != '-':
            # Find IP in any list to prevent stale isolation
            all_entries = resource.get(address=address)
            for e in all_entries:
                target_id = e.get('.id') or e.get('id')
                if target_id: resource.remove(id=target_id)
        
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "msg": str(e)}
    finally:
        if conn: conn.disconnect()

def calculate_due_date(year, month, billing_day):
    """ Hitung tanggal jatuh tempo (dengan handle tanggal 31/kabisat) """
    import calendar
    
    # Get last day of the month
    last_day = calendar.monthrange(year, month)[1]
    
    # Use the smaller value (handles date 31 in Feb, Apr, etc)
    safe_day = min(billing_day, last_day)
    
    return datetime(year, month, safe_day)

# ==============================================================================
#  HELPER & MONITORING ENGINE
# ==============================================================================
def parse_size(size_str):
    try: return float(size_str) if size_str else 0.0
    except: return 0.0

def parse_hotspot_limit_bytes(datalimit):
    if datalimit is None: return None
    raw_data = str(datalimit).upper().strip()
    if raw_data == "" or raw_data == "0": return "0"
    try:
        num_str = ''.join([c for c in raw_data if c.isdigit() or c == '.'])
        num = float(num_str) if num_str else 0
        if 'G' in raw_data: bytes_total = int(num * 1024 * 1024 * 1024)
        elif 'M' in raw_data: bytes_total = int(num * 1024 * 1024)
        elif 'K' in raw_data: bytes_total = int(num * 1024)
        else: bytes_total = int(num)
        return str(bytes_total)
    except:
        return "0"

def format_speed(bps):
    try:
        val = float(bps)
        if val >= 1000000000: return f"{val/1000000000:.1f}G"
        if val >= 1000000: return f"{val/1000000:.1f}M"
        if val >= 1000: return f"{val/1000:.0f}k"
        return f"{val:.0f}"
    except: return "0"

def get_cpu_temp():
    try:
        with open("/sys/class/thermal/thermal_zone0/temp", "r") as f: return int(int(f.read().strip()) / 1000)
    except: return 0

def apply_bulk_updates(updates):
    """ Terapkan update status secara atomik lewat SQLite """
    if not updates: return
    try:
        # Invalidate Topology Cache to ensure next load gets pings/status
        global _TOPO_CACHE
        _TOPO_CACHE['expiry'] = 0
        
        db.apply_bulk_updates(updates)
    except Exception as e:
        print(f"[DB ERROR] apply_bulk_updates failed: {e}")

def ping_ip_manual(ip):
    if not ip or ip in ["0.0.0.0", "Dynamic"]: return False
    try:
        param_c = '-n' if platform.system().lower() == 'windows' else '-c'
        param_w = '-w' if platform.system().lower() == 'windows' else '-W'
        wait_v = '1000' if platform.system().lower() == 'windows' else '1'
        subprocess.check_output(['ping', param_c, '1', param_w, wait_v, ip], stderr=subprocess.STDOUT)
        return True
    except: return False

def ping_ip_linux(ip):
    """Ping menggunakan /usr/bin/ping (Linux asli) untuk stabilitas"""
    if not ip or ip in ["0.0.0.0", "Dynamic"]: return False
    # Hanya untuk Linux
    if platform.system().lower() == 'windows': return ping_ip_manual(ip)
    try:
        subprocess.check_output(['/usr/bin/ping', '-c', '1', '-W', '1', ip], stderr=subprocess.STDOUT)
        return True
    except: return False

def fetch_single_router_data(router_id, login_data, router_config):
    if not routeros_api or not login_data.get("host"): 
        MK_RES[router_id] = {"error": True}
        return

    conn = None
    try:
        api_port = int(login_data.get("port", 8728))
        conn = routeros_api.RouterOsApiPool(
            login_data["host"], username=login_data["user"], password=login_data["pass"], 
            port=api_port, plaintext_login=True
        )
        api = conn.get_api()

        # Load settings and current DB for profile sync
        settings = load_settings()
        db_now = load_db()

        res = api.get_resource('/system/resource').get()
        ident = api.get_resource('/system/identity').get()
        
        det_lan = 0; det_sfp = 0
        try:
            all_ifaces = api.get_resource('/interface').get()
            for i in all_ifaces:
                itype = i.get('type', '').lower()
                idefault = i.get('default-name', '').lower()
                if (itype == 'ether') or ('ether' in idefault) or ('sfp' in idefault):
                    if 'sfp' in i.get('name', '').lower() or 'sfp' in idefault: det_sfp += 1
                    else: det_lan += 1
            if det_lan == 0 and det_sfp == 0: det_lan = 1
        except: det_lan = 1

        wan_rx_tot = 0; wan_tx_tot = 0; wan_name = "Scanning..."
        manual_wan = router_config.get('manual_wan', '')
        target_ifaces = []
        if manual_wan:
            target_ifaces = [x.strip() for x in manual_wan.split(',') if x.strip()]
            wan_name = manual_wan
        else:
            try:
                routes = api.get_resource('/ip/route').get(dst_address='0.0.0.0/0', active='true')
                if routes:
                    gw = routes[0].get('gateway')
                    arp = api.get_resource('/ip/arp').get(address=gw)
                    detected = arp[0].get('interface') if arp else gw
                    target_ifaces = [detected]; wan_name = detected
            except: pass
            
        for iface in target_ifaces:
            try:
                traf = api.get_resource('/interface').call('monitor-traffic', {'interface': iface, 'once': 'true'})
                if traf:
                    wan_rx_tot += parse_size(traf[0].get('rx-bits-per-second', '0'))
                    wan_tx_tot += parse_size(traf[0].get('tx-bits-per-second', '0'))
            except: pass

        ping_tgt = router_config.get('ping_target', '8.8.8.8')
        ext_ping_res = "Wait..."
        try:
            p_res = api.get_resource('/').call('ping', {'address': ping_tgt, 'count': '1'})
            ext_ping_res = p_res[0].get('time', 'RTO') if p_res else "RTO"
        except: ext_ping_res = "Error"

        try:
            arps = api.get_resource('/ip/arp').get()
            arp_map = {a.get('address'): True for a in arps if a.get('complete')=='true' or a.get('mac-address')}
        except: arp_map = {}

        try:
            leases = api.get_resource('/ip/dhcp-server/lease').get()
            lease_map = {l.get('address'): l.get('status') for l in leases}
        except: lease_map = {}

        try:
            nws = api.get_resource('/tool/netwatch').get()
            netwatch_map = {str(n.get('host')).strip(): str(n.get('status')).lower() for n in nws}
        except: netwatch_map = {}

        isolir_prof_name = settings.get('billing', {}).get('isolir_profile', 'ISOLIR')
        
        # ISOLIR MAP (Normalize IPs by stripping subnet masks like /32)
        try:
            i_list = api.get_resource('/ip/firewall/address-list').get(list=isolir_prof_name)
            isolir_map = {str(i.get('address')).split('/')[0]: True for i in i_list}
        except: isolir_map = {}

        secrets = api.get_resource('/ppp/secret').get()
        actives = api.get_resource('/ppp/active').get()
        
        # HOTSPOT SUPPORT
        hotspot_actives = []
        try:
            hotspot_actives = api.get_resource('/ip/hotspot/active').get()
        except: pass

        active_map = {a.get('name'): a for a in actives}
        secret_map = {s.get('name'): s for s in secrets}
        # Local Secrets Cache
        MK_CACHE[router_id] = {
            "secrets": [{"name":s['name'], "profile":s.get('profile','')} for s in secrets],
            "radius_candidates": [] 
        }

        # Filter RADIUS Candidates (Active but not in Secret)
        radius_candidates = []
        for a in actives:
            if a.get('name') not in secret_map:
                radius_candidates.append({
                    "name": a.get('name'), "service": a.get('service', 'pppoe'),
                    "ip": a.get('address'), "uptime": a.get('uptime'),
                    "profile": a.get('profile', ''), "type": "pppoe_radius"
                })
        for h in hotspot_actives:
            radius_candidates.append({
                "name": h.get('user'), "service": "hotspot",
                "ip": h.get('address'), "uptime": h.get('uptime'),
                "profile": h.get('profile', ''), "type": "hotspot"
            })
        MK_CACHE[router_id]["radius_candidates"] = radius_candidates

        # Queue Traffic Map (Critical for Dashboard Traffic)
        traffic_map = {}
        try:
            all_queues = api.get_resource('/queue/simple').get()
            for q in all_queues:
                tgt = q.get('target','').split('/')[0]
                rate = q.get('rate','0/0').split('/')
                traffic_map[tgt] = int(rate[0]) + int(rate[1])
        except: pass

        # MikroTik Ping Map
        mikrotik_ping_map = {}
        try:
            db_now = load_db()
            mikrotik_ping_clients = [
                c for c in db_now.get('clients', [])
                if c.get('managed_by', 'server_utama') == router_id
                and not c.get('credentials', {}).get('pppoe_user')
                and c.get('monitor_mode') == 'mikrotik_ping'
                and c.get('ip') and c.get('ip') not in ["0.0.0.0", "Dynamic", "-", ""]
            ]
            
            for client in mikrotik_ping_clients:
                try:
                    ping_res = api.get_resource('/').call('ping', {'address': client['ip'], 'count': '2'})
                    is_up = any((p.get('received') == '1') or (p.get('ttl') and p.get('size')) for p in ping_res)
                    mikrotik_ping_map[client['ip']] = is_up
                except:
                    mikrotik_ping_map[client['ip']] = False
        except Exception as e:
            print(f"[WARN] MK Ping Loop Error: {e}")

        # Simpan ke Global Cache untuk API
        MK_RES[router_id] = {
            "cpu": int(res[0].get('cpu-load', 0)) if res else 0,
            "uptime": res[0].get('uptime', '00:00:00') if res else "-",
            "board": res[0].get('board-name', '-') if res else "-",
            "version": res[0].get('version', '-') if res else "-",
            "identity": ident[0].get('name', '-') if ident else "-",
            "wan_name": wan_name,
            "wan_rx": format_speed(wan_rx_tot),
            "wan_tx": format_speed(wan_tx_tot),
            "wan_rx_raw": wan_rx_tot,
            "wan_tx_raw": wan_tx_tot,
            "port_lan": det_lan,
            "port_sfp": det_sfp,
            "ext_ping": ext_ping_res,
            "ping_target": ping_tgt,
            "actives": actives,
            "hotspot_actives": hotspot_actives
        }

        # Race Condition Fix: Don't load-modify-save the whole DB.
        # Instead, calculate updates locally and apply atomically.
        db_snap = load_db()
        updates = []
        pending_notifs = []
        
        for c in db_snap.get('clients', []):
            if c.get('managed_by', 'server_utama') == router_id:
                monitor_mode = c.get('monitor_mode', 'default')
                c_ip = c.get('ip')
                pppoe_u = c.get('credentials', {}).get('pppoe_user')
                c_id = c.get('id')
                
                upd_entry = {'id': c_id}
                has_upd = False

                if pppoe_u and pppoe_u.strip() != "":
                    is_active = pppoe_u in active_map
                    is_disabled = secret_map.get(pppoe_u,{}).get('disabled')=='true'
                    
                    isolir_prof = settings.get('billing', {}).get('isolir_profile', 'ISOLIR')
                    cur_p = secret_map.get(pppoe_u, {}).get('profile') or active_map.get(pppoe_u, {}).get('profile')
                    
                    if cur_p == isolir_prof: new_stat = 'isolir'
                    elif is_active: new_stat = 'online'
                    elif is_disabled: new_stat = 'isolir'
                    else: new_stat = 'offline'

                    if isolir_map.get(str(c_ip).strip()): new_stat = 'isolir'
                    
                    if c.get('status') != new_stat:
                        upd_entry['status'] = new_stat; has_upd = True
                        cl_mode = c.get('mode', 'pppoe')
                        mode_label = "PPPoE" if cl_mode == 'pppoe' else ("Radius" if cl_mode == 'pppoe_radius' else "Client")
                        log_msg = f"{mode_label} Connected" if new_stat == 'online' else (f"{mode_label} Isolir" if new_stat == 'isolir' else f"{mode_label} Disconnected")
                        add_log(c['name'], new_stat, log_msg)
                        
                        evt = 'up' if new_stat == 'online' else ('down' if new_stat == 'offline' else None)
                        if evt:
                            if evt == 'up' and c.get('status') == 'isolir': pass
                            else:
                                pending_notifs.append((evt, {
                                    'name': c['name'], 'ip': active_map[pppoe_u].get('address') if is_active else c.get('ip', '-'),
                                    'packet': c.get('packet_name') or '-', 'mode': mode_label, 'total_online': 0, 'total_offline': 0
                                }))
                    
                    if is_active: 
                        new_ip = active_map[pppoe_u].get('address')
                        if c.get('ip') != new_ip: upd_entry['ip'] = new_ip; has_upd = True
                    else:
                        if c.get('ip') != '-': upd_entry['ip'] = '-'; has_upd = True
                    
                    if cur_p and cur_p != isolir_prof and c.get('packet_name') != cur_p:
                        upd_entry['packet_name'] = cur_p; has_upd = True
                    if c.get('ping_ms') != -1: upd_entry['ping_ms'] = -1; has_upd = True
                
                elif c_ip and c_ip not in ["0.0.0.0", "Dynamic", "-", ""]:
                    is_online = False
                    if monitor_mode == 'netwatch':
                        if netwatch_map.get(str(c_ip).strip()) == 'up': is_online = True
                    elif monitor_mode == 'mikrotik_ping':
                        if mikrotik_ping_map.get(c_ip, False): is_online = True
                    elif monitor_mode == 'api':
                        in_arp = c_ip in arp_map
                        is_bound = lease_map.get(c_ip) == 'bound'
                        if in_arp or is_bound:
                            if traffic_map.get(c_ip, 0) > 100: is_online = True
                            else:
                                try:
                                    p_res = api.get_resource('/').call('ping', {'address': c_ip, 'count': '3'})
                                    is_online = any((p.get('received') == '1') or (p.get('ttl') and p.get('size')) for p in p_res)
                                except: is_online = False
                        else:
                            if traffic_map.get(c_ip, 0) > 1000: is_online = True
                    
                    router_sync_modes = ['api', 'mikrotik_ping', 'netwatch']
                    if monitor_mode in router_sync_modes:
                        global STATUS_FAIL_TRACK
                        f_count = STATUS_FAIL_TRACK.get(c_id, 0)
                        if is_online:
                            new_stat = 'online'
                            STATUS_FAIL_TRACK[c_id] = 0
                        else:
                            if c.get('status') == 'online':
                                f_count += 1
                                STATUS_FAIL_TRACK[c_id] = f_count
                                new_stat = 'offline' if f_count >= 3 else 'online'
                            else:
                                new_stat = 'offline'
                    else:
                        new_stat = c.get('status', 'offline')
                    
                    clean_ip = str(c_ip).strip().split('/')[0]
                    if isolir_map.get(clean_ip): new_stat = 'isolir'
                    elif new_stat == 'isolir': new_stat = 'online'
                    
                    if c.get('status') != new_stat:
                        upd_entry['status'] = new_stat; has_upd = True
                        src_label = "Router API"
                        if monitor_mode == 'mikrotik_ping': src_label = "MikroTik Ping" 
                        elif monitor_mode == 'netwatch': src_label = "Netwatch"
                        add_log(c['name'], new_stat, f"Connected ({src_label})" if new_stat == 'online' else f"Disconnected ({src_label})")
                        
                        evt = 'up' if new_stat == 'online' else 'down'
                        pending_notifs.append((evt, {
                            'name': c['name'], 'ip': c_ip, 'packet': c.get('packet_name') or '-',
                            'mode': 'Static', 'total_online': 0, 'total_offline': 0
                        }))
                
                # RADIUS ENFORCEMENT LOOP
                if c.get('status') == 'isolir' and c.get('mode') == 'pppoe_radius':
                    curr_ip = None
                    if pppoe_u in active_map: curr_ip = active_map[pppoe_u].get('address')
                    
                    if curr_ip:
                        # User is active! Check if this IP is already blocked
                        # We use a cache or just do it blindly? Doing it blindly might spam API.
                        # Optimization: Check if IP matches what we think is isolated?
                        # Or better: check address-list presence.
                        try:
                            fw_list = api.get_resource('/ip/firewall/address-list')
                            # Check if current IP is in ISOLIR list
                            isolir_prof = settings.get('billing', {}).get('isolir_profile', 'ISOLIR')
                            is_blocked = fw_list.get(address=curr_ip, list=isolir_prof)
                            
                            if not is_blocked:
                                # Not in list, add it
                                # Remove anything with same comment first
                                old_entries = fw_list.get(list=isolir_prof, comment=f"{isolir_prof}_{c['name']}")
                                for o in old_entries: fw_list.remove(id=o['.id'])
                                
                                add_to_address_list(curr_ip, isolir_prof, router_id, comment=f"{isolir_prof}_{c['name']}")
                                
                                # 3. Kick
                                kick_pppoe_user(pppoe_u, router_id)
                        except: pass
                
                if has_upd: updates.append(upd_entry)

        if updates: 
            apply_bulk_updates(updates)
            # Dispatch Notifications AFTER updates applied to DB (for accurate counts)
            for p_evt, p_data in pending_notifs:
                dispatch_telegram_event(p_evt, p_data)



    except Exception as e:
        print(f"[ERROR] MikroTik Fetch ({router_id}): {e}")
        MK_RES[router_id] = {"error": True}
    finally:
        if conn: conn.disconnect()



# --- THREAD LOOPS ---
def monitor_mikrotik_loop():
    while True:
        try:
            data = load_db()
            srv = data.get("server", {})
            if srv.get("login", {}).get("host"): fetch_single_router_data("server_utama", srv["login"], srv)
            for rtr in data.get("extra_routers", []):
                if rtr.get("login", {}).get("host"): fetch_single_router_data(rtr["id"], rtr["login"], rtr)
        except: pass
        time.sleep(10)


def turbo_ping_loop():
    time.sleep(5)
    while True:
        try:
            data_db = load_db()
            updates = []
            now = time.time()
            
            # 1. Separate Clients by Mode
            linux_clients = []
            python_clients = []
            
            for c in data_db.get('clients', []):
                # Skip PPPoE or non-IP users
                if c.get('credentials', {}).get('pppoe_user'): continue
                if not c.get('ip') or c.get('ip') in ["0.0.0.0", "Dynamic", "-", ""]: continue
                
                mode = c.get('monitor_mode', 'python_ping') # Default to python_ping ( behavior)
                if mode == 'default': mode = 'python_ping'
                
                if mode == 'linux_ping':
                    linux_clients.append(c)
                elif mode == 'python_ping':
                    python_clients.append(c)
            
            pending_notifs = []

            # Create update helper
            def prepare_update(cl, is_alive, ms_val, src_name):
                # Helper to create atomic update dict
                upd = {'id': cl['id']}
                changed = False
                
                if cl.get('ping_ms') != ms_val:
                    upd['ping_ms'] = ms_val; changed = True
                
                last_ok = cl.get('last_seen_ts', 0)
                curr_stat = cl.get('status', 'offline')
                
                if is_alive:
                    upd['last_seen_ts'] = now; changed = True
                    
                    # PROTECTION: If client is currently 'isolir' in DB, do NOT change to 'online'
                    # Allow the main sync thread to handle isolation -> online transition
                    if curr_stat != 'online' and curr_stat != 'isolir':
                        upd['status'] = 'online'; changed = True
                        add_log(cl['name'], 'online', f"Connected ({src_name} OK)")
                        
                        # Automation Alert: Suppress UP if previous was ISOLIR
                        if curr_stat != 'isolir':
                            data_alert = {
                                'name': cl['name'], 
                                'ip': cl.get('ip', '-'), 
                                'total_online': 0, 
                                'total_offline': 0,
                                'packet': cl.get('packet_name') or cl.get('service_plan', '-')
                            }
                            pending_notifs.append(('up', data_alert))
                else:
                    # Debounce 20s
                    if (now - last_ok) > 20:
                        is_isolir_db = cl.get('billing', {}).get('payment_status') == 'overdue'
                        target_stat = 'isolir' if is_isolir_db else 'offline'
                        
                        if curr_stat != target_stat:
                            upd['status'] = target_stat; changed = True
                            add_log(cl['name'], target_stat, f"Disconnected ({src_name} RTO)")
                            
                            # Automation Alert: Suppress DOWN if going to ISOLIR
                            if target_stat != 'isolir':
                                data_alert = {'name': cl['name'], 'total_online': 0, 'total_offline': 0}
                                pending_notifs.append(('down', data_alert))
                
                if changed: updates.append(upd)

            # 2. Process LINUX Ping (Sequential System Ping)
            for cl in linux_clients:
                try:
                    alive = ping_ip_linux(cl['ip']) # Uses absolute path /usr/bin/ping
                    ms = 10 if alive else -1
                    prepare_update(cl, alive, ms, "Linux Ping")
                except: pass

            # 3. Process PYTHON Ping (Turbo / Batch)
            if python_clients:
                if ICMP_READY:
                    targets = [c['ip'] for c in python_clients]
                    if targets:
                        try:
                            results = turbo_ping(targets, count=3, interval=0.5, timeout=1.5, privileged=True)
                            results_map = {r.address: r for r in results}
                            
                            for cl in python_clients:
                                res = results_map.get(cl['ip'])
                                if not res: continue
                                
                                alive = res.is_alive
                                val = int(res.min_rtt)
                                ms = 1 if val < 1 else val
                                if not alive: ms = -1
                                
                                # Fallback if permission error
                                if not alive and ping_ip_manual(cl['ip']):
                                    alive = True; ms = 10
                                    
                                prepare_update(cl, alive, ms, "Turbo Ping")
                        except Exception as e:
                            print(f"[TURBO ERROR] {e}")
                else:
                    # Fallback to manual if ICMP lib not installed
                    for cl in python_clients:
                        alive = ping_ip_manual(cl['ip'])
                        ms = 10 if alive else -1
                        prepare_update(cl, alive, ms, "System Ping")

            # 4. Apply Atomic Updates
            if updates: 
                apply_bulk_updates(updates)
                # Dispatch Notifications AFTER updates applied to DB (for accurate counts)
                for p_evt, p_data in pending_notifs:
                    dispatch_telegram_event(p_evt, p_data)
            
        except Exception as e:
            print(f"[PING LOOP ERROR] {e}")
        time.sleep(5)

def cleanup_old_safety_backups(keep_days=7):
    """Menghapus file .bak.* (safety backup) yang lebih lama dari N hari."""
    try:
        now_ts = time.time()
        cutoff = now_ts - (keep_days * 86400)
        target_dirs = [BASE_DIR, os.path.join(BASE_DIR, 'templates')]
        
        # Tambahan folder statis jika ada update OTA sebelumnya
        for sub in ['js', 'css']:
            p = os.path.join(BASE_DIR, 'static', sub)
            if os.path.exists(p): target_dirs.append(p)

        count = 0
        for d in target_dirs:
            if not os.path.exists(d): continue
            for f in os.listdir(d):
                if '.bak.' in f:
                    fp = os.path.join(d, f)
                    if os.path.isfile(fp):
                        try:
                            if os.path.getmtime(fp) < cutoff:
                                os.remove(fp)
                                count += 1
                        except: pass
        if count > 0:
            print(f"[CLEANUP] Berhasil menghapus {count} file backup lama (.bak)")
    except Exception as e:
        print(f"[CLEANUP ERROR] {e}")

def auto_backup_logic(force=False):
    """Logic Backup Utama"""
    try:
        settings = load_settings()
        bk_conf = settings.get('automation', {}).get('backup', {})
        
        # Check Time (HH:MM)
        now = get_local_now()
        sched_time = bk_conf.get('schedule_time', '02:00')
        hour = int(sched_time.split(':')[0])
        minute = int(sched_time.split(':')[1])
        
        is_time = (now.hour == hour and now.minute == minute)
        
        # CLEANUP: Always run daily cleanup at scheduled time even if backup is disabled
        if force or is_time:
            keep_days = int(bk_conf.get('keep_days', 7))
            threading.Thread(target=cleanup_old_safety_backups, kwargs={'keep_days': keep_days}, daemon=True).start()

        if not force and not bk_conf.get('enabled', False): return
        
        if force or (now.hour == hour and now.minute == minute):
            global LAST_BACKUP_DATE
            today_str = now.strftime('%d-%m-%y') # Format request: 30-01-26
            if LAST_BACKUP_DATE == today_str: return
            
            # 1. Prepare Zip
            BACKUP_DIR = os.path.join(SCRIPT_DIR, 'backups')
            if not os.path.exists(BACKUP_DIR): os.makedirs(BACKUP_DIR)
            
            # Filename for internal storage
            filename = f"BACKUP_{now.strftime('%d-%m-%y_%H%M')}.zip"
            zip_path = os.path.join(BACKUP_DIR, filename)
            
            includes = bk_conf.get('include_files', ['topology.db', 'settings.json', 'finance.json', 'config.json', 'license.key', 'app.py', 'db_manager.py', 'license_utils.py', 'wa-bridge.js', 'package.json', 'keygen.py', 'CARA_PAKAI_KEYGEN.md', 'DISTRIBUSI_KLIEN.md', 'templates', 'static'])
            
            # Safety: Ensure app.py is ALWAYS included even if removed from settings
            if 'app.py' not in includes:
                includes.append('app.py')
            if 'static' not in includes:
                includes.append('static')
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for f in includes:
                    # Use absolute path for robustness
                    f_path = os.path.abspath(os.path.join(SCRIPT_DIR, f))
                    if not os.path.exists(f_path): continue
                    
                    if os.path.isfile(f_path):
                        zipf.write(f_path, f)
                    elif os.path.isdir(f_path):
                        for root, _, files in os.walk(f_path):
                            for file in files:
                                full_p = os.path.join(root, file)
                                rel_p = os.path.relpath(full_p, SCRIPT_DIR)
                                zipf.write(full_p, rel_p)
            
            LAST_BACKUP_DATE = today_str # Mark as done
            
            # 2. Send Telegram
            tg_conf = settings.get('automation', {}).get('telegram', {})
            if tg_conf.get('notifications', {}).get('backup_report', True):
                if tg_conf.get('enabled', False):
                    caption = f"💾 AUTO BACKUPS NMS V3 {now.strftime('%d-%m-%y')}"
                    send_telegram_file(zip_path, caption)
            
            # 3. Clean Old Backups (Redundant since moved to start of function, but kept for force flow)
            keep_days = int(bk_conf.get('keep_days', 7))
            cutoff = time.time() - (keep_days * 86400)
            
            for f in os.listdir(BACKUP_DIR):
                fp = os.path.join(BACKUP_DIR, f)
                if os.path.getmtime(fp) < cutoff:
                    try: os.remove(fp)
                    except: pass
            
            # Cleanup safety backups (.bak.*) as well
            cleanup_old_safety_backups(keep_days=keep_days)
            
            time.sleep(70) # Prevent multiple execution in same minute

    except Exception as e:
        print(f"[BACKUP ERROR] {e}")

def auto_backup_loop():
    while True:
        try:
            auto_backup_logic()
        except: pass
        time.sleep(10)

def sync_billing_from_finance(client=None, db_data=None):
    """
    Proactively syncs a client's billing status/paid_until from finance history.
    Useful for legacy data or when finance entries were added without automatic triggers.
    """
    if not client: return False
    
    fin = load_finance()
    client_id = client.get('id')
    if not client_id: return False
    
    # Filter transactions for this client, sorted by date DESC
    txs = [t for t in fin.get('transactions', []) if t.get('client_id') == client_id 
           and t.get('category') in ['wifi_payment', 'Pembayaran WiFi']]
    
    changed = False
    if txs:
        txs.sort(key=lambda x: x.get('date', ''), reverse=True)
        latest_tx = txs[0]
        
        import re
        note = latest_tx.get('note', '')
        # date_ = latest_tx.get('date', '') # Unused
        
        # 1. Look for explicit "Lunas s/d YYYY-MM-DD" in note
        match = re.search(r'Lunas s/d (\d{4}-\d{2}-\d{2})', note)
        if match:
            new_expiry = match.group(1)
            # Only sync if current date is empty or NEWER than database (Don't revert manual updates)
            curr_p = client.get('paid_until')
            if not curr_p or new_expiry > curr_p:
                client['paid_until'] = new_expiry
                changed = True
    
    # 2. Update status based on paid_until vs NOW or History
    now = get_local_now()
    paid_until_str = client.get('paid_until')
    
    if 'billing' not in client: client['billing'] = {}
    old_status = client['billing'].get('payment_status')
    
    # Logic: Penentuan Status
    new_status = 'unpaid'
    
    if paid_until_str:
        # --- SUMBER KEBENARAN UTAMA: paid_until ---
        try:
            expiry_dt = datetime.strptime(paid_until_str, '%Y-%m-%d').date()
            if expiry_dt >= now.date():
                new_status = 'paid'
            else:
                # Meskipun hari ini ada transaksi, kalau paid_until masih di masa lalu,
                # berarti dia masih nunggak (mungkin bayar maret tapi cuma buat nutup februari).
                new_status = 'unpaid'
        except: 
            new_status = 'unpaid'
    else:
        # --- FALLBACK: Jika tidak ada paid_until, cek history transaksi bulan ini ---
        cur_month = now.strftime('%Y-%m')
        has_current_payment = any(t.get('date', '').startswith(cur_month) and float(t.get('amount', 0)) > 0 for t in txs)
        if has_current_payment:
            new_status = 'paid'

    if old_status != new_status:
        # Jangan menimpa status 'overdue' (manual/auto isolir) menjadi 'paid' 
        # jika client saat ini sedang dalam status 'isolir' dan masa aktif belum benar-benar diperpanjang.
        if client.get('status') == 'isolir' and new_status == 'paid' and (old_status == 'overdue' or old_status == 'unpaid'):
             # If isolated, only allow switch to 'paid' if there is a CURRENT month payment 
             # OR the expiry is actually in the future.
             # (This is already handled by the logic above, but we keep the guard for safety)
             pass
        
        client['billing']['payment_status'] = new_status
        changed = True
            
    return changed

def run_billing_check(notify_only=False, target_user=None, force=False, template_mode='auto'):
    global GLOBAL_BILLING_HEARTBEAT, _TOPO_CACHE
    GLOBAL_BILLING_HEARTBEAT = get_local_now().strftime('%d %b %Y, %H:%M:%S')
    try:
        # Load directly from billing.json
        billing_config = load_billing_config()
        
        if not force:
            # Check if isolation is disabled globally
            auto_isolir = billing_config.get('auto_isolir_enabled', True)
        else:
            auto_isolir = True
        
        # Use DBManager instance for updates, keep db_data for dict reading
        from db_manager import DBManager
        db_mgr = DBManager(DB_FILE)
        db_data = load_db()
        now = get_local_now()
        grace_period = billing_config.get('grace_period_days', 3)
        isolir_profile = billing_config.get('isolir_profile', 'ISOLIR')
        
        changed = False
        processed_count = 0
        isolir_count = 0
        reactivate_count = 0
        wa_queue = [] # Queue for WhatsApp messages
        deferred_isolations = [] # NEW: Queue for MikroTik isolation commands
        client_updates = {} # Track ONLY changed clients to prevent data wipe races
        
        # Localized month for templates
        settings = load_settings()
        pref_lang = billing_config.get('language') or settings.get('language', 'id')
        curr_month_id = get_month_name(now.month, pref_lang)
        
        db_data = load_db()
        target_found = False
        target_skip_reason = None
        
        if target_user:
            pass

        # Iterating through clients
        for idx, client in enumerate(db_data.get('clients', [])):
            try:
                # Get PPPoE user
                pppoe_user = client.get('credentials', {}).get('pppoe_user')
                
                # FILTERING LOGIC
                if target_user:
                    # Match against PPPoE User OR IP address OR Client Name
                    pp_user = (pppoe_user and pppoe_user.strip().lower() == target_user.strip().lower())
                    c_ip_match = (client.get('ip') and client.get('ip').strip() == target_user.strip())
                    c_name_match = (client.get('name') and client.get('name').strip().lower() == target_user.strip().lower())
                    
                    if not (pp_user or c_ip_match or c_name_match):
                        continue
                    target_found = True

                # FORCE MODE must be BEFORE Bypass/Enabled checks for manual testing
                if force and target_user and target_found:
                    router_id = client.get('managed_by', 'server_utama')
                    success = False
                    
                    # Save original_profile BEFORE change
                    current_profile_real = client.get('packet_name', 'default')
                    if pppoe_user:
                        try:
                            # Try to get live profile from Mikrotik
                            live_prof = get_pppoe_current_profile(pppoe_user, router_id)
                            if live_prof: current_profile_real = live_prof
                        except: pass
                    
                    if isolir_profile.upper() not in current_profile_real.upper():
                        if 'billing' not in client: client['billing'] = {}
                        client['billing']['original_profile'] = current_profile_real

                    # Execute Isolation
                    if pppoe_user:
                        is_radius = client.get('mode') == 'pppoe_radius'
                        if is_radius:
                            # PPPoE RADIUS Isolation (Address List method)
                            conn_pool = None
                            try:
                                conn_pool = get_router_connection(router_id)
                                if conn_pool:
                                    api = conn_pool.get_api()
                                    target_ip = None
                                    try:
                                        ppp_act = api.get_resource('/ppp/active').get(name=pppoe_user)
                                        if ppp_act: target_ip = ppp_act[0].get('address')
                                    except: pass
                                    
                                    if not target_ip: target_ip = client.get('ip')
                                    
                                    if target_ip and target_ip != '-':
                                        ensure_isolir_profile(router_id)
                                        res_add = add_to_address_list(target_ip, isolir_profile, router_id, comment=f"{isolir_profile}_{client['name']}")
                                        if res_add.get('status') == 'ok':
                                            kick_pppoe_user(pppoe_user, router_id)
                                            success = True
                                        else:
                                            target_skip_reason = f"Failed to add to address-list: {res_add.get('msg', 'Unknown')}"
                                    else:
                                        target_skip_reason = "User offline and no static IP found in DB"
                                else:
                                    target_skip_reason = "Could not connect to router"
                            except Exception as e:
                                target_skip_reason = f"Radius isolation error: {str(e)}"
                            finally:
                                if conn_pool: conn_pool.disconnect()
                        else:
                            # Local Secret Mode
                            ensure_isolir_profile(router_id)
                            res1 = change_pppoe_profile(pppoe_user, isolir_profile, router_id)
                            if res1.get('status') == 'ok':
                                time.sleep(1)
                                kick_pppoe_user(pppoe_user, router_id)
                                success = True
                            else:
                                # Use 'msg' instead of 'error' to match helper return
                                target_skip_reason = f"Failed to change profile: {res1.get('msg', 'Unknown')}"
                    elif client.get('ip') and client.get('ip') != '-':
                        # Static IP Mode
                        res1 = add_to_address_list(client['ip'], isolir_profile, router_id, comment=f"{isolir_profile}_{client['name']}")
                        if res1.get('status') == 'ok':
                            success = True
                        else:
                            target_skip_reason = f"Failed to add to address-list: {res1.get('msg', 'Unknown')}"
                    else:
                        target_skip_reason = "No PPPoE user or valid IP for isolation"

                    if success:
                        client['status'] = 'isolir'
                        if 'billing' not in client or not isinstance(client.get('billing'), dict):
                            client['billing'] = {}
                        client['billing']['payment_status'] = 'overdue'
                        client['billing']['isolir_date'] = now.strftime('%Y-%m-%d')
                        changed = True
                        isolir_count += 1
                        add_log(client['name'], 'isolir', f'Manual test isolation (FORCE) to profile: {isolir_profile}')
                        
                        # Telegram Notification
                        d_alert = {
                            'name': client.get('name', '-'),
                            'ip': client.get('ip', '-'),
                            'packet': client.get('packet_name') or '-'
                        }
                        dispatch_telegram_event('isolir', d_alert)
                        
                        # [NEW] Manual Force WA Notification
                        wa_enabled = billing_config.get('send_wa_notification', False)
                        phone_number = client.get('phone') or client.get('whatsapp_number') or client.get('wa_number') or client['billing'].get('wa_number')
                        if wa_enabled and phone_number:
                            msg_tpl = settings.get('wa_template_isolir', "Yth. {name}, layanan internet Anda diisolir karena tunggakan. Silakan lakukan pembayaran segera.")
                            # Dynamic Expired Date for Manual Isolir
                            manual_exp = now.strftime('%d-%m-%Y')
                            if client.get('paid_until'):
                                try:
                                    dt_me = datetime.strptime(client['paid_until'], '%Y-%m-%d')
                                    m_n_me = get_month_name(dt_me.month, pref_lang)
                                    manual_exp = f"{dt_me.day} {m_n_me} {dt_me.year}"
                                except: pass

                            wa_msg = msg_tpl.replace("{name}", client.get('name', 'Pelanggan'))\
                                           .replace("{id}", str(client.get('id', '')).replace('client_', ''))\
                                           .replace("{price}", "0")\
                                           .replace("{expired}", manual_exp)
                            wa_queue.append({"to": phone_number, "msg": wa_msg})
                            client['billing']['isolir_wa_sent'] = True

                        # Use targeted updates instead of full client object to prevent overwriting manual changes
                        client_updates[client['id']] = {
                            'id': client['id'],
                            'status': client['status'],
                            'billing': client['billing']
                        }
                        db_mgr.apply_bulk_updates([client_updates[client['id']]])
                        _TOPO_CACHE = {"data": None, "expiry": 0}
                        break
                    else:
                        # Log failure to console for admin
                        add_log(client['name'], 'error', f"Gagal Isolir Manual: {target_skip_reason}")
                        break

                # BYPASS CHECK
                if client.get('bypass_billing', False):
                    if target_found: target_skip_reason = "Client has BILLING BYPASS enabled"
                    continue
                
                # NORMAL MODE: Check business rules
                # Skip if billing not enabled for client
                if 'billing' not in client or not isinstance(client['billing'], dict):
                    client['billing'] = {}
                billing = client['billing']
                
                if not billing.get('enabled', False):
                    if target_found: target_skip_reason = "Billing feature disabled for this client"
                    continue
                
                # Reset notification flag if NOT isolated
                if client.get('status') != 'isolir' and billing.get('isolir_wa_sent'):
                    billing['isolir_wa_sent'] = False
                    changed = True
                    # Initialize update dict if not exists
                    if client['id'] not in client_updates: client_updates[client['id']] = {'id': client['id']}
                    client_updates[client['id']]['billing'] = billing
                
                # NEW: Bypass Billing (Gratis Selamanya) check
                bypass_list = billing_config.get('bypass_list', [])
                if client.get('name') in bypass_list or str(client.get('id')) in bypass_list:
                    if target_found: target_skip_reason = "Client is in Bypass List (Gratis Selamanya)"
                    continue
                
                # Proactive Sync from Finance
                if sync_billing_from_finance(client, db_data):
                    changed = True
                    if client['id'] not in client_updates: client_updates[client['id']] = {'id': client['id']}
                    client_updates[client['id']]['paid_until'] = client.get('paid_until')
                    client_updates[client['id']]['billing'] = client.get('billing')
                    billing = client.get('billing', {}) # Re-fetch updated object
                
                if (not pppoe_user or pppoe_user.strip() == "") and (not client.get('ip') or client.get('ip') == '-'):
                    if target_found: target_skip_reason = "No PPPoE username or valid IP"
                    continue
                
                # NEW LOGIC: Use paid_until as primary source of truth
                paid_until_str = client.get('paid_until')
                
                due_date = None # Initialize due_date
                if paid_until_str:
                    try:
                        due_date = datetime.strptime(paid_until_str, '%Y-%m-%d')
                    except:
                        # Fallback: Pakai billing_day (di level client atau global)
                        # If paid_until_str was invalid, use billing_day logic
                        pass # Let the next block handle it if due_date is still None
                
                if not due_date: # If paid_until was not present or invalid
                    # Legacy: Pakai billing_day (di level client atau global)
                    b_day = billing.get('billing_day') or billing_config.get('default_billing_day', 5)
                    # Prioritize client-specific billing_day
                    client_billing = client.get('billing', {})
                    if client_billing and isinstance(client_billing, dict):
                        b_day = client_billing.get('billing_day') or b_day

                    # Jika hari ini belum sampai tanggal tagihan bulan ini, berarti yang dicek adalah jatuh tempo bulan lalu
                    if now.day < b_day:
                        prev_m = now.month - 1
                        prev_y = now.year
                        if prev_m == 0: prev_m = 12; prev_y -= 1
                        due_date = calculate_due_date(prev_y, prev_m, b_day)
                    else:
                        due_date = calculate_due_date(now.year, now.month, b_day)
                
                days_overdue = (now - due_date).days
                payment_status = billing.get('payment_status', 'unpaid')
                
                # Dynamic Expiry Check: Khusus pengiriman WA agar H-0 tetap terkirim meskipun status masih 'paid'
                if paid_until_str:
                    try:
                        exp_date = datetime.strptime(paid_until_str, '%Y-%m-%d').date()
                        # Kriteria LUNAS: Masa aktif harus melewati bulan berjalan
                        next_month_start = (now.replace(day=1) + timedelta(days=32)).replace(day=1).date()
                        if exp_date < next_month_start: payment_status = 'unpaid'
                    except: pass
                
                processed_count += 1

                # --- PRE-CALCULATE VARIABLES FOR NOTIFICATIONS ---
                phone_number = client.get('phone') or client.get('whatsapp_number') or client.get('wa_number') or billing.get('wa_number')
                
                # Variabel Dasar (Price & Expired)
                billing_profiles = settings.get('billing_profiles', {})
                packet_name = (client.get('packet_name') or "").strip()
                
                price_val = 0
                for prof_name, prof_price in billing_profiles.items():
                    if prof_name.strip().lower() == packet_name.lower():
                        price_val = prof_price; break
                
                manual_arr_val = 0
                manual_arrears_list = settings.get('manual_arrears', [])
                for ma in manual_arrears_list:
                    if ma.get('client_name') == client['name']:
                        manual_arr_val += int(ma.get('amount', 0))

                unpaid_months_wa = 1
                if paid_until_str:
                    try:
                        dt_exp = datetime.strptime(paid_until_str, '%Y-%m-%d')
                        unpaid_months_wa = (now.year - dt_exp.year) * 12 + (now.month - dt_exp.month)
                        if unpaid_months_wa < 1: unpaid_months_wa = 1
                    except:
                        if days_overdue > 0: unpaid_months_wa = math.ceil(days_overdue / 30)
                elif days_overdue >= 0:
                    unpaid_months_wa = math.ceil(days_overdue / 30)
                else:
                    # New Client or recently ON: Default to current month only
                    unpaid_months_wa = 1
                
                # Professional Debt Calculation
                # Fallback Harga Paketan jika billing_profiles tidak cocok
                if price_val == 0 and packet_name:
                    match = re.search(r'(\d+)rb', packet_name, re.I) or re.search(r'(\d+)k', packet_name, re.I)
                    if match: price_val = int(match.group(1)) * 1000

                # Jika LUNAS, Harga Paket (actual_package_debt) harus 0
                if payment_status == 'paid':
                    actual_package_debt = 0.0
                else:
                    actual_package_debt = float(price_val) * unpaid_months_wa
                
                total_debt_num = actual_package_debt + manual_arr_val
                price_str = "{:,.0f}".format(total_debt_num).replace(",", ".")
                
                paid_until_val = client.get('paid_until')
                if paid_until_val and '-' in paid_until_val:
                    try:
                        dt_exp = datetime.strptime(paid_until_val, '%Y-%m-%d')
                        exp_m_name = get_month_name(dt_exp.month, pref_lang)
                        expired_date = f"{dt_exp.day} {exp_m_name} {dt_exp.year}"
                    except: expired_date = paid_until_val
                else:
                    last_day_val = calendar.monthrange(now.year, now.month)[1]
                    expired_date = f"{last_day_val} {curr_month_id} {now.year}"

                qris_path = os.path.join(SCRIPT_DIR, 'static', 'photos', 'qris.jpg')
                has_qris = os.path.exists(qris_path)
                wa_auto_qris = settings.get('wa_auto_qris', True)
                wa_manual_qris = settings.get('wa_manual_qris', True)
                wa_isolir_qris = settings.get('wa_isolir_qris', True)

                # --- NEW: WHATSAPP NOTIFICATION LOGIC ---
                wa_enabled = billing_config.get('send_wa_notification', False)

                if notify_only:
                     # Check extraction logic preview
                     _append_wa_log(f"[DEBUG] Client: {client.get('name')} | Overdue: {days_overdue} | Status: {payment_status} | Phone: {phone_number} | WA Enabled: {wa_enabled}")
                
                # Allow if enabled OR if manually triggered
                wa_auto_enabled = settings.get('wa_auto_enabled', True)
                wa_manual_enabled = settings.get('wa_manual_enabled', True)
                
                # Logic: If manual mode (button click), check wa_manual_enabled. If auto loop, check wa_auto_enabled.
                is_manual = template_mode == 'manual'
                can_send = (wa_enabled and not is_manual and wa_auto_enabled) or (is_manual and wa_manual_enabled)

                if can_send and phone_number:
                    # Logic Flags
                    wa_end_month_active = billing_config.get('wa_end_month_enabled', False)
                    wa_pre_isolir_active = billing_config.get('wa_pre_isolir_enabled', False)
                    wa_pre_isolir_days = int(billing_config.get('wa_pre_isolir_days', 2))

                    # One-time notification tracking (Month-Year period)
                    curr_period = now.strftime('%m-%Y')
                    if 'wa_sent_track' not in billing or not isinstance(billing.get('wa_sent_track'), dict):
                        billing['wa_sent_track'] = {}

                    # Helper untuk replace semua variabel
                    def format_wa_msg(tpl):
                        # Logic pintar: Jika paid_until masih aktif, status = LUNAS
                        msg_status = "LUNAS" if payment_status == 'paid' else "BELUM DIBAYAR"
                        
                        return tpl.replace("{name}", client['name'])\
                                  .replace("{id}", str(client.get('id', '')).replace('client_', ''))\
                                  .replace("{month}", curr_month_id)\
                                  .replace("{price}", "{:,.0f}".format(price_val).replace(",", "."))\
                                  .replace("{bill}", price_str)\
                                  .replace("{amount}", price_str)\
                                  .replace("{expired}", expired_date)\
                                  .replace("{profile}", packet_name or "-")\
                                  .replace("{time}", now.strftime("%H:%M"))\
                                  .replace("{status}", msg_status)\
                                  .replace("{x}", str(wa_pre_isolir_days))

                    # 1. REMOVED: H-3 REMINDER (Per User Request)
                    
                    # 2. END-OF-MONTH / DUE-DATE REMINDER
                    # Allow run if active OR manual mode
                    # FIX: Prevent scheduled messages from being sent during routine isolir checks (off-schedule).
                    # Only send if notify_only is True (from 9 AM scheduler) or it's triggered manually.
                    if (notify_only or template_mode == 'manual') and wa_end_month_active and (payment_status == 'unpaid' or manual_arr_val > 0) and not client.get('bypass_billing'):
                        b_mode = billing_config.get('billing_mode', 'monthly')
                        should_trigger_wa = False
                        
                        # REFINED: Use scheduling logic by default, only bypass if force=True
                        if force:
                            should_trigger_wa = True
                        elif b_mode == 'cyclic':
                            # Mode 30 Hari: Trigger pas HARI-H Jatuh Tempo
                            if days_overdue == 0:
                                should_trigger_wa = True
                        elif b_mode == 'fixed':
                            # [MODE FIXED]: Trigger pas Fixed Day (H-0)
                            # Ambil Fixed Day dari client, fallback ke default_billing_day
                            b_day = billing.get('billing_day') or billing_config.get('default_billing_day', 5)
                            if now.day == b_day:
                                should_trigger_wa = True
                            else:
                                # Handle tgl 31 di bulan pendek (30)
                                last_day_in_mo = calendar.monthrange(now.year, now.month)[1]
                                if b_day > last_day_in_mo and now.day == last_day_in_mo:
                                    should_trigger_wa = True
                        else:
                            # Mode Global: Trigger pas HARI TERAKHIR BULAN
                            last_day = calendar.monthrange(now.year, now.month)[1]
                            if now.day == last_day:
                                should_trigger_wa = True
                                

                        if should_trigger_wa:
                            # Check track STRICTLY (No manual bypass)
                            # User Request: "proteksi pasti hanya bisa kirim 1x"
                            if billing['wa_sent_track'].get('eom') != curr_period:
                                msg_tpl = settings.get('wa_template', "Halo {name}, tagihan anda belum terbayar. Mohon segera diselesaikan. Terima kasih.")
                                
                                wa_msg = format_wa_msg(msg_tpl)
                                attach_qris = has_qris and wa_manual_qris
                                wa_queue.append({"to": phone_number, "msg": wa_msg, "image": qris_path if attach_qris else None})
                                billing['wa_sent_track']['eom'] = curr_period
                                changed = True
                                if client['id'] not in client_updates: client_updates[client['id']] = {'id': client['id']}
                                client_updates[client['id']]['billing'] = billing
                                if notify_only: _append_wa_log(f"[TRACE] Queued Manual/EOM for {client['name']}")
                            else:
                                if notify_only: _append_wa_log(f"[INFO] Skip EOM {client['name']}: Already sent for {curr_period}")
                        else:
                             if notify_only: _append_wa_log(f"[DEBUG] Skip EOM {client['name']}: Trigger condition not met (EOM/Cyclic/Manual)")
                    
                    elif notify_only:
                        _append_wa_log(f"[DEBUG] Skip Notifications for {client['name']}: can_send=False (WA Enabled: {wa_enabled}, Manual Mode: {is_manual})")

                    # 3. PRE-ISOLATION WARNING (H-X before Grace ends)
                    # Trigger only on AUTO mode
                    # FIX: Only trigger during the scheduled 9 AM notification cycle (notify_only).
                    # [ANTI-SPAM]: Nonaktifkan Pre-Isolir jika mode Fixed aktif
                    b_mode_tmp = billing_config.get('billing_mode', 'monthly')
                    if (notify_only) and wa_pre_isolir_active and payment_status == 'unpaid' and template_mode != 'manual':
                        trigger_day = (grace_period + 1) - wa_pre_isolir_days
                        if days_overdue == trigger_day:
                             # Check track
                            if billing['wa_sent_track'].get('pre_isolir') != curr_period:
                                msg_tpl = settings.get('wa_template_auto', "Yth. {name}, layanan internet Anda akan dinonaktifkan dalam {x} hari karena tunggakan. Mohon segera melakukan pembayaran.")
                                wa_msg = format_wa_msg(msg_tpl)
                                attach_qris = has_qris and wa_auto_qris
                                wa_queue.append({"to": phone_number, "msg": wa_msg, "image": qris_path if attach_qris else None})
                                billing['wa_sent_track']['pre_isolir'] = curr_period
                                changed = True
                                if client['id'] not in client_updates: client_updates[client['id']] = {'id': client['id']}
                                client_updates[client['id']]['billing'] = billing
                
                # ISOLIR & REACTIVATION: Skip if notify_only is True
                if not notify_only:
                    # ISOLIR CONDITION: Priority to paid_until
                    should_isolate = False
                    
                    # SMART ISOLIR: Also isolate if there is manual debt and it's past cycle/eom
                    manual_arr_val = 0 # Re-calculate just in case for isolation logic
                    manual_arrears_list = settings.get('manual_arrears', [])
                    for ma in manual_arrears_list:
                        if ma.get('client_name') == client['name']:
                            manual_arr_val += int(ma.get('amount', 0))
                    has_manual_debt = manual_arr_val > 0
                    
                    if paid_until_str:
                        # If paid_until exists, check if date is expired OR if it has manual debt past due
                        # FIX: Don't isolate if ALREADY PAID manually
                        if payment_status == 'paid':
                            should_isolate = False
                        elif days_overdue > grace_period:
                            should_isolate = True
                        elif has_manual_debt:
                            # FIX: Gunakan perhitungan kalender asli (timedelta) untuk menghindari bug "Tanggal 33"
                            try:
                                b_day = billing.get('billing_day') or billing_config.get('default_billing_day', 5)
                                # Target isolir bulan ini = tgl tagihan bulan ini + grace
                                isolir_deadline = datetime(now.year, now.month, 1) + timedelta(days=b_day - 1 + grace_period)
                                if now > isolir_deadline:
                                    # Still check status for manual debt case
                                    if payment_status != 'paid':
                                        should_isolate = True
                            except:
                                # Fallback simple addition if somehow date constructor fails
                                if now.day > (billing.get('billing_day') or 5) + grace_period and payment_status != 'paid':
                                    should_isolate = True
                    else:
                        # Fallback for legacy clients without paid_until
                        if (days_overdue > grace_period and payment_status == 'unpaid') or has_manual_debt:
                            should_isolate = True

                    if should_isolate and auto_isolir:
                        # Only isolir if not already isolated
                        if client.get('status') != 'isolir':
                            # DEFER ISOLATION: Add to deferred list to execute after WA notifications
                            deferred_isolations.append({
                                'client_idx': idx,
                                'client_name': client['name'],
                                'router_id': client.get('managed_by', 'server_utama'),
                                'pppoe_user': pppoe_user,
                                'ip_addr': client.get('ip'),
                                'days_overdue': days_overdue,
                                'packet_name': client.get('packet_name', 'default'),
                                'isolir_profile': isolir_profile  # <--- FIX: Simpan profil spesifik router saat dimasukkan
                            })

                            # QUEUE ISOLATION NOTIFICATION NOW
                            wa_isolir_enabled = settings.get('wa_isolir_enabled', True)
                            if wa_enabled and wa_isolir_enabled and phone_number:
                                # Only send isolation WA once per arrear period
                                if not billing.get('isolir_wa_sent'):
                                    msg_tpl = settings.get('wa_template_isolir', "Yth. {name}, layanan internet Anda diisolir karena tunggakan sebesar Rp {price}. Silakan lakukan pembayaran segera.")
                                    
                                    # Use pre-calculated price_str and expired_date
                                    # NEW LOGIC: Use helper or manual replace for consistency
                                    wa_msg = format_wa_msg(msg_tpl, client_status="ISOLIR")
                                    
                                    attach_qris = has_qris and wa_isolir_qris
                                    wa_queue.append({"to": phone_number, "msg": wa_msg, "image": qris_path if attach_qris else None})
                                    billing['wa_sent_track']['isolir'] = curr_period # Use 'isolir' key for tracking
                                    billing['isolir_wa_sent'] = True
                                    changed = True
                                    if client['id'] not in client_updates: client_updates[client['id']] = {'id': client['id']}
                                    client_updates[client['id']]['billing'] = billing
                                else:
                                    if notify_only: _append_wa_log(f"[INFO] Skip Isolir WA {client['name']}: Already sent (isolir_wa_sent=True)")
                            elif notify_only:
                                _append_wa_log(f"[DEBUG] Skip Isolir WA {client['name']}: wa_isolir_enabled={settings.get('wa_isolir_enabled', True)}, wa_enabled={wa_enabled}, phone={phone_number}")
                    
                    # REACTIVATION CONDITION: Previously isolated but now paid
                    elif client.get('status') == 'isolir' and payment_status == 'paid':
                        router_id = client.get('managed_by', 'server_utama')
                        success = False
                        
                        if pppoe_user:
                            is_radius = client.get('mode') == 'pppoe_radius'
                            if is_radius:
                                # PPPoE RADIUS Reactivation (Remove from Address List)
                                conn_ra = None
                                try:
                                    conn_ra = get_router_connection(router_id)
                                    if conn_ra:
                                        api = conn_ra.get_api()
                                        fw_list = api.get_resource('/ip/firewall/address-list')
                                        entries = fw_list.get(list=isolir_profile, comment=f"{isolir_profile}_{client['name']}")
                                        for e in entries: fw_list.remove(id=e['id'])
                                except: pass
                                finally:
                                    if conn_ra: conn_ra.disconnect()
                                kick_pppoe_user(pppoe_user, router_id)
                                success = True
                            else:
                                # Local Secret Reactivation
                                current_profile = billing.get('original_profile') or client.get('packet_name', 'default')
                                res = change_pppoe_profile(pppoe_user, current_profile, router_id)
                                if res.get('status') == 'ok':
                                    time.sleep(1)
                                    kick_pppoe_user(pppoe_user, router_id)
                                    success = True
                        elif client.get('ip') and client.get('ip') != '-':
                            res = remove_from_address_list(client['ip'], isolir_profile, router_id, comment=f"{isolir_profile}_{client['name']}")
                            if res.get('status') == 'ok':
                                success = True
                        
                        if success:
                            client['status'] = 'online'
                            billing['isolir_wa_sent'] = False # Reset flag for next arrear period
                            changed = True
                            if client['id'] not in client_updates: client_updates[client['id']] = {'id': client['id']}
                            client_updates[client['id']]['status'] = 'online'
                            client_updates[client['id']]['billing'] = billing
                            reactivate_count += 1
                            add_log(client['name'], 'online', 'Auto-reactivation after payment')

                            # 3. REACTIVATION NOTIFICATION (AUTO)
                            wa_react_enabled = settings.get('wa_reactivate_enabled', True)
                            if wa_enabled and wa_react_enabled and phone_number:
                                msg_tpl = settings.get('wa_template_reactivate', "Halo {name}, pembayaran telah diterima dan layanan internet Anda telah diaktifkan kembali. Terima kasih.")
                                wa_msg = format_wa_msg(msg_tpl)
                                wa_queue.append({"to": phone_number, "msg": wa_msg, "image": None})
                    else:
                        if target_found: 
                            if payment_status == 'paid': target_skip_reason = "Status is PAID"
                            elif days_overdue <= grace_period: target_skip_reason = f"Not overdue yet ({days_overdue} days overdue, grace {grace_period})"
                            else: target_skip_reason = "Conditions not met"

            except Exception as e:
                print(f"[BILLING] Error processing client {client.get('name')}: {e}")
                if target_found: target_skip_reason = f"Error: {str(e)}"
                continue

        # --- AFTER MAIN LOOP: SEND NOTIFICATIONS & EXECUTE DEFERRED ISOLATIONS ---

        # 1. Trigger WhatsApp Batch first
        if wa_queue:
            try:
                import tempfile, json
                with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, dir=SCRIPT_DIR) as tf:
                    json.dump(wa_queue, tf); tf.flush()
                    temp_path = tf.name
                # Ensure file is closed before spawning worker
                tf.close() 
                spawn_wa_worker(mode="batch", task_file=temp_path)
                
                # Delay isolation by 10 seconds to allow WA delivery while internet IS STILL ON
                if deferred_isolations:
                    time.sleep(10)
            except Exception as e:
                pass

        # 2. Process Deferred MikroTik Isolation Commands
        if deferred_isolations:
            for item in deferred_isolations:
                try:
                    client = db_data['clients'][item['client_idx']]
                    pppoe_user = item['pppoe_user']
                    router_id = item['router_id']
                    ip_addr = item['ip_addr']
                    client_name = item['client_name']
                    days_overdue = item['days_overdue']
                    item_isolir_profile = item.get('isolir_profile', 'ISOLIR')
                    
                    success = False
                    if pppoe_user:
                        # PPPoE / RADIUS Isolation
                        is_radius = client.get('mode') == 'pppoe_radius'
                        if is_radius:
                            # Address List method for Radius
                            conn = None
                            try:
                                conn = get_router_connection(router_id)
                                if conn:
                                    api = conn.get_api()
                                    # Cleanup & Re-add logic
                                    try:
                                        fw_list = api.get_resource('/ip/firewall/address-list')
                                        old = fw_list.get(list=item_isolir_profile, comment=f"{item_isolir_profile}_{client_name}")
                                        for o in old: fw_list.remove(id=o['id'])
                                    except: pass
                                    
                                    target_ip = None
                                    ppp_act = api.get_resource('/ppp/active').get(name=pppoe_user)
                                    if ppp_act: target_ip = ppp_act[0].get('address')
                                    
                                    if not target_ip: target_ip = ip_addr
                                    if target_ip and target_ip != '-':
                                        # Save original profile for Radius
                                        cur_prof = client.get('packet_name', 'default')
                                        if item_isolir_profile.upper() not in cur_prof.upper():
                                            if 'billing' not in client: client['billing'] = {}
                                            client['billing']['original_profile'] = cur_prof

                                        res_add = add_to_address_list(target_ip, item_isolir_profile, router_id, comment=f"{item_isolir_profile}_{client_name}")
                                        if res_add.get('status') == 'ok':
                                            kick_pppoe_user(pppoe_user, router_id)
                                            success = True
                                    else:
                                        # User Offline? Mark success anyway so db updates status to ISOLIR (Parity with Original)
                                        success = True
                            except Exception as e:
                                print(f"[BILLING] Radius isolation error for {client_name}: {e}")
                            finally:
                                if conn: conn.disconnect()
                        else:
                            # Local Secret Mode
                            ensure_isolir_profile(router_id)
                            
                            # Get real current profile from Mikrotik, fallback to DB packet_name, then 'default'
                            real_prof_mk = get_pppoe_current_profile(pppoe_user, router_id)
                            cur_prof = real_prof_mk if (real_prof_mk and real_prof_mk != "") else (item.get('packet_name') or 'default')
                            
                            res1 = change_pppoe_profile(pppoe_user, item_isolir_profile, router_id)
                            if res1.get('status') == 'ok':
                                time.sleep(1)
                                kick_pppoe_user(pppoe_user, router_id)
                                if 'billing' not in client: client['billing'] = {}
                                
                                # Only overwrite original_profile if the current one isn't ALREADY isolir
                                if item_isolir_profile.upper() not in cur_prof.upper():
                                    client['billing']['original_profile'] = cur_prof
                                    
                                success = True
                    elif ip_addr and ip_addr != '-':
                        # Static IP Mode
                        # Save original profile for Static IP
                        cur_prof = client.get('packet_name', 'default')
                        if item_isolir_profile.upper() not in cur_prof.upper():
                            if 'billing' not in client: client['billing'] = {}
                            client['billing']['original_profile'] = cur_prof
                            
                        res1 = add_to_address_list(ip_addr, item_isolir_profile, router_id, comment=f"{item_isolir_profile}_{client_name}")
                        if res1.get('status') == 'ok': success = True
                    
                    if success:
                        client['status'] = 'isolir'
                        if 'billing' not in client: client['billing'] = {}
                        client['billing']['payment_status'] = 'overdue'
                        client['billing']['isolir_date'] = now.strftime('%Y-%m-%d')
                        # Sync packet_name to ISOLIR for dashboard consistency
                        client['packet_name'] = item_isolir_profile
                        changed = True
                        
                        if client['id'] not in client_updates: client_updates[client['id']] = {'id': client['id']}
                        client_updates[client['id']]['status'] = 'isolir'
                        client_updates[client['id']]['packet_name'] = item_isolir_profile
                        client_updates[client['id']]['billing'] = client['billing']
                        
                        isolir_count += 1
                        add_log(client_name, 'isolir', f"Auto-isolir (Scheduled): Tunggakan {days_overdue} hari")
                        add_log("SYSTEM", "system", f"Isolir Otomatis: {client_name} (Tunggakan {days_overdue} hari)")
                        
                        # Telegram Notification
                        d_alert = {
                            'name': client_name,
                            'ip': client.get('ip', '-'),
                            'packet': client.get('packet_name') or '-'
                        }
                        dispatch_telegram_event('isolir', d_alert)
                    else:
                        print(f"[BILLING] Skip DB update for {client_name}: Router command failed/skipped.")
                except Exception as ex:
                    print(f"[BILLING] Critical error processing deferred isolation for {item.get('client_name')}: {ex}")

        # Summarize results
        # Improved Save logic using apply_bulk_updates to prevent race conditions
        if changed:
            # 1. Collect all clients that were modified in the main loop or deferred isolation
            # (Note: client objects were modified in-place in db_data['clients'])
            # Since we have client_updates={} initialized but some logic just sets 'changed=True'
            # we do a final sweep for changed clients if not already captured.
            # But the most reliable way is to just keep the original 'changed' check 
            # and collect all modified client dicts.
            
            # Identify all modified clients (marked by isolation/reactivation logic)
            final_upds = list(client_updates.values())
            
            # If we used the local 'changed' variable but skipped adding to client_updates (safety fallback)
            if not final_upds and changed:
                # This should rarely happen with the new granular logic, but we keep it for extreme safety.
                # However, we only pass fields that were actually intended to change.
                # Let's just trust client_updates for now as it's more precise.
                pass
            
            if final_upds:
                db_mgr.apply_bulk_updates(final_upds)

            # Invalidate Cache
            _TOPO_CACHE = {"data": None, "expiry": 0}
            
        # HEARTBEAT LOG
        wa_sent_count = len(wa_queue) if wa_queue else 0
        # Summary log (Heartbeat)
        g_wa_en = billing_config.get('send_wa_notification', False)
        summary_msg = f"Billing Heartbeat: {processed_count} checked, {isolir_count} isolated, {reactivate_count} active, {len(wa_queue)} notifications (WA: {g_wa_en})."
        add_log("SYSTEM_BILLING", "system", summary_msg)

        if target_user:
            if target_found:
                return f"Target {target_user} scan finished. Result: {target_skip_reason or 'Processed'}"
            else:
                return f"Target user '{target_user}' not found in DB"
        
        return f"Routine check finished. {processed_count} checked, {isolir_count} isolated, {reactivate_count} active."
    except Exception as e:
        print(f"[ERROR] Billing check error: {e}")
        return f"Error: {e}"

# --- WHATSAPP BRIDGE LOGIC ---
WA_LOG_FILE = os.path.join(SCRIPT_DIR, 'wa_logs.json')
wa_log_lock = threading.Lock() # Lock for thread-safe file access
active_wa_workers = [] # Track running processes

def spawn_wa_worker(mode="test", target="", message="", task_file=None, image=None):
    """
    [Singleton Standby] Menambahkan pesan WhatsApp ke dalam antrean (queue).
    Jika service wa-bridge.js belum menyala, maka akan dinyalakan.
    Jika sudah menyala, cukup menambahkan antrean saja.
    """
    import time, random, platform
    try:
        bridge_path = os.path.join(SCRIPT_DIR, 'wa-bridge.js')
        nm_path = os.path.join(SCRIPT_DIR, 'node_modules')
        queue_dir = os.path.join(SCRIPT_DIR, 'wa_queue')
        
        if not os.path.exists(bridge_path):
            return False, "Bridge file missing"
            
        if not os.path.exists(queue_dir):
            try: os.makedirs(queue_dir)
            except: pass

        now_ts = get_local_now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 1. Tulis Job ke Folder Antrean
        job_id = f"job_{int(time.time()*100)}_{random.randint(100,999)}.json"
        job_path = os.path.join(queue_dir, job_id)
        
        job_data = {
            "mode": mode,
            "to": target,
            "msg": message,
            "file": task_file,
            "image": image,
            "timestamp": now_ts
        }
        
        with open(job_path, 'w') as f:
            json.dump(job_data, f)
            
        _append_wa_log(f"[{now_ts}] QUEUE: Job {job_id} ditambahkan (Mode: {mode})")

        # 2. Periksa apakah wa-bridge.js sudah jalan (Singleton Check)
        global active_wa_workers
        # Bersihkan process yang sudah mati dari list lokal
        active_wa_workers = [p for p in active_wa_workers if p.poll() is None]
        
        # Cross-process check (Global Singleton):
        # Mencegah Gunicorn spawning multiple bridge processes
        pid_file = os.path.join(SCRIPT_DIR, "wa_bridge.pid")
        bridge_is_running = False
        if os.path.exists(pid_file):
            try:
                with open(pid_file, 'r') as f:
                    old_pid = int(f.read().strip())
                # Sinyal 0 mencek apakah proses masih hidup
                os.kill(old_pid, 0)
                bridge_is_running = True
            except:
                bridge_is_running = False
        
        if len(active_wa_workers) > 0 or bridge_is_running:
            _append_wa_log(f"[{now_ts}] INFO: Worker sudah Standby secara global. Job diantrekan.")
            return True, "Job queued (Worker already running)"

        # 3. Lakukan Spawn Baru karena belum ada yang jalan
        import shutil
        node_exe = shutil.which("node") or shutil.which("nodejs")
        if not node_exe:
            # IMPROVED: Search in user Linux home (NVM)
            home_dir = os.path.expanduser("~")
            common_paths = [
                "/usr/bin/node", "/usr/local/bin/node", "/bin/node", 
                "/usr/bin/nodejs", "/usr/local/bin/nodejs",
                "/opt/node/bin/node", "/snap/bin/node",
                os.path.join(home_dir, ".nvm/versions/node/v*/bin/node")
            ]
            import glob
            for p_pattern in common_paths:
                found_paths = glob.glob(p_pattern)
                if found_paths:
                    node_exe = found_paths[0]
                    break
        
        if not node_exe:
            _append_wa_log(f"[{now_ts}] ERROR: Node.js tidak ditemukan.")
            return False, "Node.js not found"
            
        if not os.path.exists(nm_path):
            _append_wa_log(f"[{now_ts}] ERROR: Folder 'node_modules' tidak ada.")
            return False, "node_modules missing."

        _append_wa_log(f"[{now_ts}] START: Membangunkan wa-bridge.js (Standby Mode)...")

        import subprocess
        try:
            # Panggil dengan mode default (standby)
            cmd = [node_exe, bridge_path, "--mode", "standby"]
            process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, cwd=SCRIPT_DIR)
            active_wa_workers.append(process)
            
            # Catat PID untuk singleton check lintas-proses
            try:
                with open(pid_file, 'w') as f:
                    f.write(str(process.pid))
            except: pass
            
            # Cek status dalam 2 detik pertama (Deteksi Crash Startup)
            time.sleep(2)
            if process.poll() is not None:
                stdout, stderr = process.communicate()
                _append_wa_log(f"[{now_ts}] CRITICAL: Script berhenti mendadak ({process.returncode}). Error: {stderr or stdout}")
                return False, f"Startup Error: {stderr or stdout}"
            
            def monitor_output(p):
                # Tangkap STDOUT
                for line in p.stdout:
                    if not line: break
                    msg = line.strip()
                    # Capture all standard bridge tags
                    tags = ["[STATUS]", "[SUCCESS]", "[ERROR]", "[FAILED]", "[SEND]", "[QR]", "[QUEUE]", "[CLEANUP]"]
                    if any(t in msg for t in tags):
                        _append_wa_log(f"[{get_local_now().strftime('%H:%M:%S')}] {msg}")
                        
                        # NEW: Update Broadcast UI Progress if it's a success or failure log
                        if "[SUCCESS]" in msg or "[FAILED]" in msg:
                            try:
                                # Parse: [SUCCESS] [ID:bc_123_0] ...
                                status = "success" if "[SUCCESS]" in msg else "failed"
                                id_tag = ""
                                if "[ID:" in msg:
                                    id_tag = msg.split("[ID:")[1].split("]")[0]
                                
                                if id_tag:
                                    # Update log_wa_broadcast.json
                                    bc_log_file = os.path.join(SCRIPT_DIR, 'log_wa_broadcast.json')
                                    bc_logs = []
                                    if os.path.exists(bc_log_file):
                                        try:
                                            with open(bc_log_file, 'r', encoding='utf-8') as f:
                                                bc_logs = json.load(f)
                                        except: bc_logs = []
                                    
                                    # Add new entry with ID for FE to match
                                    new_entry = {
                                        "id": id_tag,
                                        "status": status,
                                        "timestamp": get_local_now().strftime('%Y-%m-%d %H:%M:%S'),
                                        "msg": "Terkirim" if status == "success" else "Gagal"
                                    }
                                    bc_logs.append(new_entry)
                                    if len(bc_logs) > 500: bc_logs = bc_logs[-500:]
                                    
                                    with open(bc_log_file, 'w', encoding='utf-8') as f:
                                        json.dump(bc_logs, f, indent=2)
                            except Exception as e:
                                print(f"[DEBUG] Failed to update BC progress: {e}")
                
                # Tangkap STDERR
                for line in p.stderr:
                    if not line: break
                    _append_wa_log(f"[{get_local_now().strftime('%H:%M:%S')}] node-err: {line.strip()}")
                
                return_code = p.wait()
                # Hapus PID file setelah proses mati
                try:
                    if os.path.exists(pid_file): os.remove(pid_file)
                except: pass
                _append_wa_log(f"[{get_local_now().strftime('%H:%M:%S')}] SHUTDOWN: Service wa-bridge dimatikan otomatis (Code: {return_code})")
            
            threading.Thread(target=monitor_output, args=(process,), daemon=True).start()
            
            _append_wa_log(f"[{now_ts}] INFO: Singleton Worker berjalan (PID: {process.pid})")
            return True, "Worker started"
            
        except Exception as e:
            _append_wa_log(f"[{now_ts}] EXCEPTION: {str(e)}")
            return False, str(e)
    except Exception as e:
        return False, str(e)

def _append_wa_log(text):
    """Simpan log khusus WhatsApp ke wa_logs.json dengan metode atomic, thread-safe, dan retry logic"""
    global wa_log_lock
    import time
    import random
    
    # Gunakan lock khusus agar tidak bentrok antar thread di dalam satu proses
    with wa_log_lock:
        # Retry mechanism (5 kali percobaan dengan backoff acak)
        # Sangat penting untuk Gunicorn dengan banyak worker
        max_retries = 5
        for attempt in range(max_retries):
            try:
                logs = []
                # 1. Safe Read: Handle missing, 0-byte, atau corrupted JSON
                if os.path.exists(WA_LOG_FILE):
                    try:
                        # Linux uses utf-8-sig fine for compatibility
                        with open(WA_LOG_FILE, 'r', encoding='utf-8-sig') as f:
                            content = f.read().strip()
                            if content:
                                logs = json.loads(content)
                    except (json.JSONDecodeError, ValueError):
                        # Hanya reset jika memang isi file bukan JSON yang valid (corrupt)
                        print(f"[WARN] WA Log JSON Corrupt, resetting.")
                        logs = []
                    except Exception as read_err:
                        # Jika gagal baca karena file sibuk/lock, angkat exception untuk retry
                        raise read_err
                
                # 2. Append & Limit
                logs.append(text)
                if len(logs) > 50: logs = logs[-50:] # Simpan 50 log terakhir
                
                # 3. Atomic Write on Linux: Write to temp file then replace
                tmp_file = WA_LOG_FILE + ".tmp"
                try:
                    with open(tmp_file, 'w', encoding='utf-8') as f:
                        json.dump(logs, f, indent=2)
                        f.flush()
                        os.fsync(f.fileno()) 
                    
                    # os.replace adalah atomic di Linux/Unix systems
                    os.replace(tmp_file, WA_LOG_FILE)
                    
                    # Jika sampai sini, berarti sukses
                    return True
                except Exception as write_err:
                    if os.path.exists(tmp_file): os.remove(tmp_file)
                    raise write_err
                    
            except Exception as e:
                # Jika ini percobaan terakhir, baru log error
                if attempt == max_retries - 1:
                    print(f"[ERROR] _append_wa_log Gagal total setelah {max_retries} percobaan: {e}")
                else:
                    # Tunggu sebentar sebelum coba lagi (backoff acak 0.1 - 0.5 detik)
                    time.sleep(random.uniform(0.1, 0.5))
        
        return False

@app.route('/api/whatsapp/logs')
def api_wa_logs():
    """Ambil logs khusus WhatsApp (Unified & Self-Healing)"""
    if not check_auth(request): return jsonify({"error":"Unauthorized"}), 401
    try:
        if os.path.exists(WA_LOG_FILE):
            with open(WA_LOG_FILE, 'r', encoding='utf-8-sig') as f:
                content = f.read().strip()
                if content:
                    logs = json.loads(content)
                    return jsonify(logs)
    except Exception as e:
        print(f"[ERROR] api_wa_logs read error: {e}")
    return jsonify([])

# Alias for backward compatibility if needed, but unified logic
@app.route('/api/logs/wa')
def get_wa_logs_route():
    return api_wa_logs()

@app.route('/api/whatsapp/test', methods=['POST'])
def api_wa_test():
    """Trigger tes kirim WA instan"""
    data = request.json or {}
    target = data.get('target', '').strip()
    client_name = data.get('name', 'Pelanggan')
    t_type = data.get('type', 'manual') # 'manual' or 'auto'
    
    settings = load_settings()
    now = get_local_now()
    pref_lang = settings.get('language', 'id')
    curr_month_id = get_month_name(now.month, pref_lang)

    if t_type == 'auto':
        template = settings.get('wa_template_auto', "Halo {name}, (Auto Response)")
    else:
        template = settings.get('wa_template', "Halo {name}, ini adalah pesan tagihan/tes dari NMS.")
        
    # Tambahkan Tunggakan Manual jika ada
    manual_arr_val = 0
    manual_arrears_list = settings.get('manual_arrears', [])
    for ma in manual_arrears_list:
        if ma.get('client_name') == client_name:
            manual_arr_val += int(ma.get('amount', 0))

    # Professional Billing Logic: Separate Master Price and Total Bill
    billing_profiles = settings.get('billing_profiles', {})
    
    # 1. Cari Harga Paket Master ({price}) - Cari dari DB atau settings
    master_price_val = 0
    p_name = "-"
    c_obj = {}
    
    # Pre-fetch client object once
    db_test = load_db()
    for c in db_test.get('clients', []):
        if c.get('name') == client_name:
            c_obj = c; break
            
    if c_obj:
        p_name = c_obj.get('packet_name', '-')
        # Match against profiles
        for prof_name, prof_price in billing_profiles.items():
            if prof_name.strip().lower() == p_name.lower():
                master_price_val = prof_price; break
                
        # Fallback regex if not found in profiles
        if master_price_val == 0:
            match = re.search(r'(\d+)rb', p_name, re.I) or re.search(r'(\d+)k', p_name, re.I)
            if match: master_price_val = int(match.group(1)) * 1000

    # 2. Hitung Nominal Tagihan ({bill}/{amount})
    final_bill_val = 0
    is_paid_status = False
    
    # Periksa status pembayaran klien di DB
    if c_obj:
        p_until = c_obj.get('paid_until')
        if p_until:
            try:
                expiry = datetime.strptime(p_until, '%Y-%m-%d')
                # Kriteria LUNAS: Masa aktif harus melewati bulan berjalan (bulan depan atau lebih)
                next_month_start = (now.replace(day=1) + timedelta(days=32)).replace(day=1)
                if expiry.date() >= next_month_start.date(): is_paid_status = True
            except: pass
            
    # Jika BELUM bayar, tambahkan harga paket ke tagihan
    if not is_paid_status:
        final_bill_val += master_price_val
        
    # Tambahkan tunggakan manual (selalu ditambahkan jika ada)
    final_bill_val += manual_arr_val
    
    # Jika FE mengirimkan bill_data (misal dari tombol Manual Arrears yang spesifik), gunakan itu
    bill_data = data.get('bill')
    if bill_data is not None and float(bill_data) > 0:
        # Jika tombol manual WA di klik, FE biasanya kirim 'totalForWa' yang sudah dihitung
        final_bill_val = float(bill_data)

    try:
        bill_str = "{:,.0f}".format(final_bill_val).replace(",", ".")
        price_master_str = "{:,.0f}".format(master_price_val).replace(",", ".")
    except:
        bill_str = "0"
        price_master_str = "0"

    # Clean ID
    raw_id = str(data.get('id', ''))
    if not raw_id:
        # Cari ID dari DB jika tidak dikirim dari FE
        db = load_db()
        for c in db.get('clients', []):
            if c.get('name') == client_name:
                raw_id = str(c.get('id', ''))
                break
    clean_id = raw_id.replace('client_', '')

    # Smart Expired Date for Test
    test_expired = f"28 {curr_month_id} {now.year}" # Default EOM Feb
    db_test = load_db()
    for c in db_test.get('clients', []):
        if c.get('name') == client_name:
            p_until = c.get('paid_until')
            if p_until and '-' in p_until:
                try:
                    dt_exp = datetime.strptime(p_until, '%Y-%m-%d')
                    m_name = get_month_name(dt_exp.month, pref_lang)
                    test_expired = f"{dt_exp.day} {m_name} {dt_exp.year}"
                except: pass
            break

    curr_time = now.strftime('%H:%M')
    p_name = "-"
    c_obj = next((c for c in db_test.get('clients', []) if c.get('name') == client_name), {})
    if c_obj: p_name = c_obj.get('packet_name', '-')

    # Replace placeholders
    test_price = price_master_str # Master Price ({price})
    test_bill = bill_str          # Total Bill ({bill} & {amount})
    status_label = "LUNAS" if is_paid_status else "BELUM DIBAYAR"
    
    message = template.replace("{name}", client_name)\
                        .replace("{id}", clean_id)\
                        .replace("{month}", curr_month_id)\
                        .replace("{price}", test_price)\
                        .replace("{amount}", test_bill)\
                        .replace("{bill}", test_bill)\
                        .replace("{expired}", test_expired)\
                        .replace("{profile}", p_name)\
                        .replace("{time}", curr_time)\
                        .replace("{status}", status_label)\
                        .replace("{x}", str(settings.get('wa_pre_isolir_days', 2)))
    
    # Path QRIS (Jika ada)
    qris_path = os.path.join(SCRIPT_DIR, 'static', 'photos', 'qris.jpg')
    has_qris = os.path.exists(qris_path)
    
    # Check if QRIS should be attached for this type
    wa_auto_qris = settings.get('wa_auto_qris', True)
    wa_manual_qris = settings.get('wa_manual_qris', True)
    attach_qris = False
    if t_type == 'auto':
        attach_qris = has_qris and wa_auto_qris
    else:
        attach_qris = has_qris and wa_manual_qris

    if not target: return jsonify({"status": "error", "msg": "Nomor tujuan harus diisi"}), 400
    
    ok, msg = spawn_wa_worker(mode="test", target=target, message=message, image=qris_path if attach_qris else None)
    if ok: return jsonify({"status": "ok", "msg": "Worker dijalankan. Pantau log di halaman Settings -> WhatsApp (Log Aktivitas di bagian bawah)."})
    else: return jsonify({"status": "error", "msg": msg})

@app.route('/api/whatsapp/init', methods=['POST'])
def api_wa_init():
    """Bangunkan worker WA untuk pairing atau sinkronisasi status"""
    ok, msg = spawn_wa_worker(mode="test", target="", message="") # Empty target = init mode
    if ok: return jsonify({"status": "ok", "msg": "Menghubungkan ke WhatsApp Service... Mohon tunggu beberapa detik."})
    else: return jsonify({"status": "error", "msg": msg})

# Removed api_wa_logs duplicate (unified above)

@app.route('/api/whatsapp/logs/clear', methods=['POST'])
def api_wa_logs_clear():
    """Hapus log aktivitas WhatsApp"""
    if not check_auth(request): return jsonify({"error":"Unauthorized"}), 401
    global wa_log_lock
    with wa_log_lock:
        try:
            # Berikan isi minimal [] agar UI tidak error saat fetch ulang
            with open(WA_LOG_FILE, 'w', encoding='utf-8') as f:
                f.write("[]")
            return jsonify({"status": "ok"})
        except Exception as e:
            return jsonify({"status": "error", "msg": str(e)})

@app.route('/api/whatsapp/reset', methods=['POST'])
def api_wa_reset():
    """Hapus sesi WhatsApp"""
    global active_wa_workers
    try:
        # 1. Matikan semua worker yang masih jalan (biar tidak ada zombie)
        for p in active_wa_workers:
            if p.poll() is None:
                try:
                    if platform.system().lower() == 'windows':
                        subprocess.run(['taskkill', '/F', '/T', '/PID', str(p.pid)], capture_output=True)
                    else:
                        p.terminate()
                        p.wait(timeout=2)
                except:
                    try: p.kill()
                    except: pass
        active_wa_workers = []
        
        # 2. Hapus sisa file QR (biar UI bersih)
        qr_file = os.path.join(SCRIPT_DIR, 'wa_qr.txt')
        if os.path.exists(qr_file): os.remove(qr_file)

        # 3. Hapus folder sesi
        session_dir = os.path.join(SCRIPT_DIR, 'wa_session')
        import shutil
        if os.path.exists(session_dir): shutil.rmtree(session_dir)
        
        # 4. Hapus folder antrean agar tidak nge-blast sisa pesan lama saat login baru
        queue_dir = os.path.join(SCRIPT_DIR, 'wa_queue')
        if os.path.exists(queue_dir):
            for file_name in os.listdir(queue_dir):
                file_path = os.path.join(queue_dir, file_name)
                try: os.unlink(file_path)
                except: pass
        
        _append_wa_log(f"[{get_local_now().strftime('%H:%M:%S')}] SYSTEM: Sesi dan proses direset total. Silakan scan ulang.")
        return jsonify({"status": "ok", "msg": "Sesi WA berhasil dihapus dan proses dimatikan."})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})

@app.route('/api/whatsapp/qr')
def api_wa_qr_status():
    """Cek status koneksi dan ambil QR jika ada"""
    qr_file = os.path.join(SCRIPT_DIR, 'wa_qr.txt')
    session_file = os.path.join(SCRIPT_DIR, 'wa_session', 'creds.json')
    
    # 1. Cek Sesi Utama
    if os.path.exists(session_file):
        return jsonify({"status": "connected", "qr": None})

    # 2. Cek apakah ada worker yang sedang 'pairing'
    is_pairing = False
    global active_wa_workers
    active_wa_workers = [p for p in active_wa_workers if p.poll() is None]
    if active_wa_workers:
        is_pairing = True

    # 3. Ambil data QR jika ada file-nya
    qr_data = None
    if os.path.exists(qr_file):
        try:
            with open(qr_file, 'r') as f:
                qr_data = f.read().strip()
                # Jika ada QR tapi worker mati, berarti worker itu zombie/lama
                if not is_pairing:
                     try: os.remove(qr_file)
                     except: pass
                     qr_data = None
        except: pass
    
    status = "pairing" if (is_pairing or qr_data) else "disconnected"
    return jsonify({"status": status, "qr": qr_data})


LAST_BILLING_NOTIF_DATE = None
GLOBAL_BILLING_HEARTBEAT = "Menunggu Sinkronisasi..."
LAST_BILLING_CHECK_TS = 0  # unix timestamp
LAST_WA_CHECK_TS = 0       # unix timestamp

def auto_billing_loop():
    """Auto-isolir clients based on configurable interval or specific time"""
    global GLOBAL_BILLING_HEARTBEAT
    time.sleep(15)  # Wait for system init
    add_log("SYSTEM", "system", "Background Thread: Auto-Billing Loop Started")
    
    last_check_time = None
    while True:
        try:
            now = get_local_now()
            GLOBAL_BILLING_HEARTBEAT = now.strftime('%d %b %Y, %H:%M:%S')
            
            # Load interval configuration (in hours)
            b_cfg = load_billing_config()
            interval_hours = int(b_cfg.get('billing_check_interval_hours', 24))
            
            # 1. Logic Interval (Routine Isolation)
            should_run_interval = False
            if last_check_time is None:
                should_run_interval = True
            else:
                elapsed_seconds = (now - last_check_time).total_seconds()
                interval_seconds = interval_hours * 3600
                if elapsed_seconds >= interval_seconds:
                    should_run_interval = True
            
            # 2. Logic Specific Time (Notifications) - PROFESSIONAL LOGIC
            wa_time_str = b_cfg.get('wa_notif_time', '09:00')
            try:
                wa_time_clean = wa_time_str.lower().replace('am', '').replace('pm', '').strip()
                t_hour, t_min = map(int, wa_time_clean.split(':'))
                if 'pm' in wa_time_str.lower() and t_hour < 12: t_hour += 12
                if 'am' in wa_time_str.lower() and t_hour == 12: t_hour = 0
                
                sched_today = now.replace(hour=t_hour, minute=t_min, second=0, microsecond=0)
                today_str = now.strftime('%Y-%m-%d')
                
                if now >= sched_today:
                    _sys_sett = _load_settings_raw()
                    _last_run_date = _sys_sett.get('last_auto_billing_date')
                    
                    if _last_run_date != today_str:
                        now_ts = now.strftime('%H:%M:%S')
                        _append_wa_log(f"[{now_ts}] INFO: Mengecek jadwal kirim hari ini ({today_str})...")
                        _append_wa_log(f"[{now_ts}] INFO: Menemukan jadwal {wa_time_str}. Membangunkan gateway WA...")
                        
                        # Executing
                        global LAST_WA_CHECK_TS
                        LAST_WA_CHECK_TS = int(now.timestamp())
                        run_billing_check(notify_only=True)

                        # Only mark as DONE after successful run_billing_check
                        # Update last_auto_billing_date
                        _fresh_sett = _load_settings_internal() 
                        _fresh_sett['last_auto_billing_date'] = today_str
                        _safe_write_json(SETTINGS_FILE, _fresh_sett, critical=True)
                        _SETTINGS_CACHE = {"data": _fresh_sett, "expiry": time.time() + 30}
                        reload_config_globals()
                        now_ts = get_local_now().strftime('%H:%M:%S')
                        _append_wa_log(f"[{now_ts}] INFO: Tugas selesai. Mematikan gateway WA (Standby).")
                    else:
                        # Already done today. Silently standby.
                        pass
            except Exception as e:
                print(f"[BILLING AUTO] Time parse error '{wa_time_str}': {e}")

            if should_run_interval:
                global LAST_BILLING_CHECK_TS
                LAST_BILLING_CHECK_TS = int(now.timestamp())
                run_billing_check(notify_only=False) # Routine Isolation (No WA log)
                last_check_time = now

        except Exception as e:
            add_log("SYSTEM", "error", f"Billing Loop Error: {str(e)}")
            print(f"[BILLING LOOP ERROR] {e}")
            
        time.sleep(60)  # Check every 1 minute

# --- START BACKGROUND THREADS (GLOBAL SCOPE PRODUCTION) ---
# Startup Cleanup (Safety Backups) - Works in Gunicorn
threading.Thread(target=cleanup_old_safety_backups, kwargs={'keep_days': 7}, daemon=True).start()

t1 = threading.Thread(target=turbo_ping_loop, daemon=True); t1.start()
t2 = threading.Thread(target=monitor_mikrotik_loop, daemon=True); t2.start()
t3 = threading.Thread(target=auto_backup_loop, daemon=True); t3.start()
t4 = threading.Thread(target=auto_billing_loop, daemon=True); t4.start()

# --- AUTH CHECK ---
def check_auth(req):
    reload_config_globals()
    t = req.headers.get('X-Auth-Token')
    if not t: t = req.args.get('token') # Support query param for downloads
    if not t: t = req.values.get('token') # Fallback to form values
    if t: t = t.strip() # Remove whitespace
    
    if t == ADMIN_PASSWORD: return "admin"
    if t == PASSWORD_VIEWER: return "viewer"
    return None

@app.route('/api/debug/run_billing', methods=['POST'])
def api_debug_run_billing():
    auth = check_auth(request)
    if not auth: return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.json or {}
    target_user = data.get('user') or data.get('target_user')
    force = data.get('force', False)
    notify_only = data.get('notify_only', False)
    
    # Pre-validation: Check if user exists in DB
    if target_user:
        db_v = load_db()
        found_in_db = False
        t_user_low = target_user.strip().lower()
        for c in db_v.get('clients', []):
            pu = c.get('credentials', {}).get('pppoe_user')
            pi = c.get('ip')
            cn = c.get('name')
            
            match_pu = (pu and pu.strip().lower() == t_user_low)
            match_pi = (pi and pi.strip() == target_user.strip())
            match_cn = (cn and cn.strip().lower() == t_user_low)
            
            if match_pu or match_pi or match_cn:
                found_in_db = True; break
                
        if not found_in_db:
            return jsonify({'status': 'error', 'message': f"User/IP/Nama '{target_user}' tidak ditemukan di database."})
    
    # Run in background to avoid timeout
    add_log("ADMIN", "system", "Menjalankan Patroli Billing secara manual.")
    threading.Thread(target=run_billing_check, kwargs={
        'target_user': target_user,
        'force': force,
        'notify_only': notify_only
    }).start()
    
    msg = "Proses penagihan (Notifikasi & Cek Isolir) berjalan di background."
    if notify_only:
        msg = "Proses Notifikasi WA berjalan di background (Tanpa Isolir)."
        
    return jsonify({'status': 'ok', 'message': msg})
@app.route('/api/version_check')
def api_version_check():
    import hashlib
    try:
        with open(__file__, 'rb') as f:
            md5 = hashlib.md5(f.read()).hexdigest()
    except: md5 = "error"
    return jsonify({
        "status": "active",
        "version": "V3.1 STABLE CLEAN",
        "file_path": os.path.abspath(__file__),
        "is_licensed": is_licensed(),
        "file_md5": md5
    })

@app.route('/')
def dashboard(): return render_template('dashboard.html')

@app.route('/maps')
def maps(): return render_template('index.html')

@app.route('/client')
def client_page(): return render_template('client.html')

@app.route('/hotspot')
def page_hotspot(): return render_template('hotspot.html')

@app.route('/pppoe')
def page_pppoe(): return render_template('pppoe.html')

@app.route('/billing')
def page_billing(): return render_template('billing.html')

@app.route('/network')
def page_network(): return render_template('network.html')

@app.route('/monitor')
def page_monitor(): return render_template('monitor.html')

@app.route('/finance')
def finance_page(): return render_template('finance.html')

@app.route('/about')
def page_about(): return render_template('about.html')

@app.route('/settings')
def settings_page(): return render_template('settings.html')

@app.route('/api/login', methods=['POST'])
def login():
    p = request.json.get('password')
    if p == ADMIN_PASSWORD: return jsonify({"status":"ok", "role":"admin"})
    if p == PASSWORD_VIEWER: return jsonify({"status":"ok", "role":"viewer"})
    return jsonify({"error":"Wrong password"}), 401

def get_system_uptime():
    try:
        boot_time = psutil.boot_time()
        uptime_seconds = time.time() - boot_time
        days = int(uptime_seconds // (24 * 3600))
        hours = int((uptime_seconds % (24 * 3600)) // 3600)
        minutes = int((uptime_seconds % 3600) // 60)
        if days > 0:
            return f"{days}d {hours}h {minutes}m"
        return f"{hours}h {minutes}m"
    except:
        return "-"

@app.route('/api/data')
def get_data_route():
    role = check_auth(request)
    if not role: return jsonify({"error":"Unauthorized"}), 401
    topo = load_db()
    
    # Inject Live Data for ALL Routers
    def _inject_live(node, r_id):
        if not node: return
        if r_id in MK_RES:
            res = MK_RES[r_id]
            if not res.get('error'):
                node['status'] = 'online'
                node['identity'] = res.get('identity', '-')
                node['_detected_wan'] = {'rx': res.get('wan_rx','0'), 'tx': res.get('wan_tx','0'), 'name': res.get('wan_name')}
                node['_detected_ports'] = {'lan': res.get('port_lan',0), 'sfp': res.get('port_sfp',0)}
            else:
                node['status'] = 'offline'
                node['identity'] = '-'
        else:
            # If no sync yet, default to offline for better UI feedback
            node['status'] = 'offline'

    _inject_live(topo.get('server'), "server_utama")
    for rtr in topo.get('extra_routers', []):
        _inject_live(rtr, rtr.get('id'))
    # License Info
    lic_info = {"status": "Unlicensed", "client": "-", "type": "FREE", "active": False}
    if is_licensed():
        try:
            with open(LICENSE_FILE, 'r') as f:
                k = f.read().strip()
            v, info = verify_license(k)
            if v:
                lic_info = {"status": "Active", "client": info.get('cli','-'), "type": "LICENSED (PRO)", "active": True}
        except: pass

    settings_data = load_settings()
    
    return jsonify({
        "topology": topo,
        "settings": settings_data,
        "system": get_system_stats_cached(), 
        "mikrotik_data": MK_RES, 
        "role": role,
        "license": lic_info
    })

def get_system_stats_cached():
    """Optimasi : Cache psutil stats for 5 seconds to reduce CPU load."""
    global _SYSTEM_STATS_CACHE
    now = time.time()
    if _SYSTEM_STATS_CACHE['expiry'] > now:
        return _SYSTEM_STATS_CACHE['data']
    
    try:
        cpu_val = psutil.cpu_percent()
        ram_val = psutil.virtual_memory().percent
        disk_val = psutil.disk_usage('/').percent
        
        stats = {
            "cpu": cpu_val, 
            "cpu_usage": cpu_val, # Compatible with Bot
            "ram": ram_val, 
            "ram_usage": ram_val, # Compatible with Bot
            "disk": disk_val, 
            "disk_usage": disk_val, # Compatible with Bot
            "temp": get_cpu_temp(),
            "uptime": get_system_uptime()
        }
        _SYSTEM_STATS_CACHE = {"data": stats, "expiry": now + 5}
        return stats
    except:
        return {}

@app.route('/api/save', methods=['POST'])
def save_route():
    if check_auth(request) != "admin": 
        return jsonify({"error":"Forbidden"}), 403
        
    incoming = request.json
    if not incoming: return jsonify({"status": "error", "msg": "No data received"}), 400
    
    # NEW: Auto-enable billing for new clients in incoming data if not specified
    if isinstance(incoming, dict) and 'clients' in incoming:
        updates_for_db = []
        current_db = load_db()
        current_clients = {str(c.get('id')): c for c in current_db.get('clients', [])}
        
        for c in incoming['clients']:
            cid = str(c.get('id'))
            if 'billing' not in c:
                # Default for new clients: Enabled
                c['billing'] = {'enabled': True}
                
            # Original behavior: Let save_db handle merging with preserve_live protection
            
    # save_db now handles merging status internally
    success = save_db(incoming, preserve_live=True)
    if not success:
        return jsonify({"status": "error", "msg": "Database write failed"}), 500

    fresh_topo = load_db()

    return jsonify({
        "status":"ok", 
        "msg": "Data saved (Live status preserved)",
        "topology": fresh_topo
    })

@app.route('/api/clients/import/template')
def api_clients_import_template():
    if not check_auth(request): return jsonify({"error":"Forbidden"}), 403
    if not OPENPYXL_READY:
        return jsonify({"status": "error", "msg": "openpyxl not installed on server"}), 500
    
    tpl_path = os.path.join(SCRIPT_DIR, 'template_import_client.xlsx')
    
    # Always regenerate to include new instructions
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Import Client'
    
    # Add Instructions
    instructions = [
        ['💡 PETUNJUK PENGISIAN TEMPLATE IMPORT CLIENT'],
        ['1. Tipe Koneksi: Isi dengan "PPPoE", "PPPoE Radius", atau "Statik"'],
        ['2. Username PPPoE: Wajib diisi untuk tipe PPPoE & PPPoE Radius (Samakan dengan secret di Mikrotik)'],
        ['3. IP Address: Wajib diisi untuk tipe Statik'],
        ['4. Nama Paket / Profile: Untuk PPPoE Radius & Statik isi manual, untuk PPPoE standar akan otomatis sync dari Mikrotik jika kosong'],
        ['5. Koordinat: Format harus "lat, lng" (Contoh: -7.123, 110.456) agar muncul di peta'],
        ['6. Billing: Isi "On" untuk aktifkan isolir otomatis, atau "Off" untuk nonaktifkan'],
        ['7. Induk: Isi dengan ID atau Nama ODP (Contoh: ODP-A-01). Kosongkan jika ingin menempel ke Router'],
        [''], # Blank row before headers
    ]
    for row in instructions:
        ws.append(row)
        
    # Headers
    headers = [
        'Pengelolah Mikrotik', 'Nama Client', 'Tipe Koneksi', 
        'Username PPPoE', 'IP Address', 'Nama Paket / Profile', 
        'WhatsApp', 'Koordinat', 'Billing', 'Induk'
    ]
    ws.append(headers)
    
    # Sample Rows
    ws.append(['server_utama', 'Contoh PPPoE', 'PPPoE', 'user_contoh', '-', 'Paket_10M', '628123456789', '-7.123, 110.456', 'On', ''])
    ws.append(['server_utama', 'Contoh Radius', 'PPPoE Radius', 'user_radius', '-', 'Paket_20M', '628123456789', '-7.124, 110.457', 'On', ''])
    ws.append(['server_utama', 'Contoh Statik', 'Statik', '-', '192.168.1.100', 'Paket_Statik', '628123456789', '-7.125, 110.458', 'Off', ''])
    
    # Simple Styling for Headers (Internal row index is len(instructions) + 1)
    header_row_idx = len(instructions) + 1
    for cell in ws[header_row_idx]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    wb.save(tpl_path)
    return send_file(tpl_path, as_attachment=True, download_name='template_import_client.xlsx')

@app.route('/api/clients/import', methods=['POST'])
def api_clients_import():
    if check_auth(request) != "admin": 
        return jsonify({"error":"Forbidden"}), 403
        
    if 'file' not in request.files:
        return jsonify({"status": "error", "msg": "No file uploaded"}), 400
        
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({"status": "error", "msg": "No selected file"}), 400

    if not openpyxl:
        return jsonify({"status": "error", "msg": "openpyxl not installed on server"}), 500

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        # Load Existing Data
        data = load_db()
        topology = data # in app.py load_db() returns the dict
        clients = topology.get('clients', [])
        routers = [{'id': 'server_utama', 'name': topology.get('server', {}).get('name', ''), 'identity': topology.get('server', {}).get('identity', '')}]
        for r in topology.get('extra_routers', []):
            routers.append({'id': r.get('id'), 'name': r.get('name', ''), 'identity': r.get('identity', '')})
            
        odps = topology.get('odps', [])
        
        new_clients_count = 0
        errors = []
        
        # Header Detection: Find the row containing "Nama Client"
        header_row_idx = 1
        headers = []
        for r_idx, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=20), start=1):
            row_vals = [str(cell.value).strip().lower() if cell.value else "" for cell in row_cells]
            if "nama client" in row_vals:
                header_row_idx = r_idx
                headers = row_vals
                break
        
        if not headers:
            return jsonify({"status": "error", "msg": "Header 'Nama Client' tidak ditemukan dalam 20 baris pertama"}), 400

        def find_idx(names):
            for n in names:
                if n.lower() in headers:
                    return headers.index(n.lower())
            return -1

        mapping = {
            'manager': find_idx(['pengelolah mikrotik', 'pengelola mikrotik', 'router']),
            'name': find_idx(['nama client', 'nama']),
            'mode': find_idx(['tipe koneksi', 'mode']),
            'ppp_user': find_idx(['username pppoe', 'pppoe user', 'user pppoe']),
            'ip': find_idx(['ip address', 'ip', 'alamat ip']),
            'packet': find_idx(['nama paket / profile', 'paket', 'profile', 'jenis client']),
            'wa': find_idx(['whatsapp', 'wa', 'no hp']),
            'coords': find_idx(['koordinat', 'coords', 'location']),
            'billing': find_idx(['billing', 'auto isolir']),
            'induk': find_idx(['induk', 'parent', 'odp']),
        }

        if mapping['name'] == -1:
            return jsonify({"status": "error", "msg": "Kolom 'Nama Client' tidak ditemukan pada baris header"}), 400

        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            name = str(row[mapping['name']]).strip() if row[mapping['name']] else None
            if not name: continue
            
            # 1. Resolve Connection Mode
            mode_val = str(row[mapping['mode']]).strip().lower() if mapping['mode'] != -1 and row[mapping['mode']] else "pppoe"
            if "radius" in mode_val: mode = "pppoe_radius"
            elif "statik" in mode_val or "static" in mode_val: mode = "static"
            else: mode = "pppoe"

            # 2. Resolve Manager
            mgr_val = str(row[mapping['manager']]).strip() if mapping['manager'] != -1 and row[mapping['manager']] else ""
            target_router_id = "server_utama"
            if mgr_val:
                for r in routers:
                    if mgr_val.lower() in [r['id'].lower(), r['name'].lower(), r['identity'].lower()]:
                        target_router_id = r['id']
                        break
            
            # 3. Resolve Parent (Induk)
            parent_val = str(row[mapping['induk']]).strip() if mapping['induk'] != -1 and row[mapping['induk']] else ""
            target_parent_id = "" # DEFAULT: Standalone (No Cable)
            if parent_val:
                found_parent = False
                for o in odps:
                    if parent_val.lower() in [o['id'].lower(), o['name'].lower()]:
                        target_parent_id = o['id']
                        found_parent = True
                        break
                if not found_parent:
                    for r in routers:
                        if parent_val.lower() in [r['id'].lower(), r['name'].lower(), r['identity'].lower()]:
                            target_parent_id = r['id']
                            found_parent = True
                            break
            
            # 4. Parse Coordinates
            coords_raw = str(row[mapping['coords']]).strip() if mapping['coords'] != -1 and row[mapping['coords']] else "0,0"
            try:
                if "," in coords_raw:
                    lat, lng = coords_raw.split(",")
                    coords = [float(lat.strip()), float(lng.strip())]
                else:
                    coords = [0, 0]
            except:
                coords = [0, 0]
            
            # 5. Billing status
            billing_idx = mapping['billing']
            billing_str = "on"
            if billing_idx != -1 and row[billing_idx]:
                billing_str = str(row[billing_idx]).strip().lower()
            billing_enabled = False if billing_str in ["off", "tidak", "false", "0"] else True
            
            # 6. Build Client Object
            new_id = f"client_{int(time.time() * 1000) + new_clients_count}"
            packet_name = str(row[mapping['packet']]).strip() if mapping['packet'] != -1 and row[mapping['packet']] else "Default"
            
            new_client = {
                "id": new_id,
                "name": name,
                "type": "client",
                "mode": mode,
                "managed_by": target_router_id,
                "parent_id": target_parent_id, # This field is actually 'parent' in the JSON, api usually uses 'parent'
                "parent": target_parent_id,
                "coordinates": coords,
                "packet_name": packet_name,
                "ip": str(row[mapping['ip']]).strip() if mapping['ip'] != -1 and row[mapping['ip']] else "-",
                "wa_number": str(row[mapping['wa']]).strip() if mapping['wa'] != -1 and row[mapping['wa']] else "",
                "billing": {"enabled": billing_enabled},
                "credentials": {
                    "pppoe_user": str(row[mapping['ppp_user']]).strip() if mapping['ppp_user'] != -1 and row[mapping['ppp_user']] else "",
                    "wifi_ssid": "",
                    "wifi_pass": "",
                    "pppoe_user_router": "",
                    "pppoe_pass": ""
                },
                "status": "offline"
            }
            
            clients.append(new_client)
            new_clients_count += 1
            time.sleep(0.001) # Ensure unique timestamp-based IDs if processing is fast

        # Save to DB
        topology['clients'] = clients
        save_db(topology, preserve_live=True)
        
        return jsonify({
            "status": "ok", 
            "msg": f"Berhasil mengimport {new_clients_count} client.",
            "count": new_clients_count
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "msg": f"Gagal memproses file: {str(e)}"}), 500

@app.route('/api/mikrotik/users/all')
def get_all_mk_users():
    if not check_auth(request): return jsonify([]), 401
    
    all_secrets = []
    # Collect secrets from all routers in MK_CACHE
    for r_id, cached in MK_CACHE.items():
        if isinstance(cached, dict) and "secrets" in cached:
            all_secrets.extend(cached["secrets"])
        elif isinstance(cached, list):
            # Fallback for old cache format if any
            all_secrets.extend(cached)
            
    return jsonify(all_secrets)

@app.route('/api/mikrotik/users/<router_id>')
def get_mk_users(router_id):
    if not check_auth(request): return jsonify([]), 401
    return jsonify(MK_CACHE.get(router_id if router_id != 'undefined' else 'server_utama', []))

@app.route('/api/bandwidth/<router_id>/<path:user>')
def get_bandwidth(router_id, user):
    if not check_auth(request): return jsonify({"error": "Auth"}), 401
    
    db = load_db()
    # Cari credentials router
    router_data = load_db()["server"] if router_id == 'server_utama' else next((r for r in load_db().get("extra_routers", []) if r["id"] == router_id), None)
    
    if not router_data or "login" not in router_data or not router_data["login"].get("host"):
        return jsonify({"rx":"0", "tx":"0"}), 200

    login_data = router_data["login"]
    host = login_data["host"]
    
    res = {"rx": "0", "tx": "0", "rx_load": 0, "tx_load": 0}

    # --- STATELESS MODE (STABIL & AMAN) ---
    # Kita kembali ke metode login-logout per request karena persistent connection
    # menyebabkan socket hanging (macet) di beberapa tipe router/network.
    conn = None
    try:
        conn = routeros_api.RouterOsApiPool(
            host, 
            username=login_data["user"], 
            password=login_data["pass"], 
            port=int(login_data.get("port", 8728)), 
            plaintext_login=True
        )
        api = conn.get_api()
        
        # Optimasi: Langsung sasar target
        queues = []
        q_res = api.get_resource('/queue/simple')
        
        queues = q_res.get(name=user)
        if not queues: queues = q_res.get(name=f"<{user}>")
        if not queues: queues = q_res.get(name=f"<pppoe-{user}>")
        if not queues and "." in user:
            queues = q_res.get(target=user)
            if not queues: queues = q_res.get(target=f"{user}/32")
        
        if queues:
            q = queues[0]
            rate = q.get('rate', '0/0').split('/'); limit = q.get('max-limit', '0/0').split('/')
            tx_curr = parse_size(rate[0]); rx_curr = parse_size(rate[1])
            tx_max = parse_size(limit[0]) if len(limit)>0 else 0; rx_max = parse_size(limit[1]) if len(limit)>1 else 0
            
            res.update({'tx': format_speed(tx_curr), 'rx': format_speed(rx_curr)})
            if tx_max > 0: res['tx_load'] = int((tx_curr / tx_max) * 100)
            if rx_max > 0: res['rx_load'] = int((rx_curr / rx_max) * 100)
    except: pass
    finally:
        if conn: conn.disconnect()

    return jsonify(res)

@app.route('/api/mikrotik/update_secret', methods=['POST'])
def update_mk_secret():
    if not check_auth(request): return jsonify({"error": "Auth"}), 401
    
    data = request.json
    router_id = data.get('router_id', 'server_utama')
    user = data.get('user')
    profile = data.get('profile')
    password = data.get('password')
    
    db = load_db()
    r_conf = load_db()["server"] if router_id == 'server_utama' else next((r for r in load_db().get("extra_routers", []) if r["id"] == router_id), None)
    if not r_conf: return jsonify({"status": "error", "msg": "Router Not Found"})
    
    lgn = r_conf.get('login', {})
    if not lgn.get('host'): return jsonify({"status": "error", "msg": "No Host"})
    
    conn = None
    try:
        conn = routeros_api.RouterOsApiPool(
            lgn['host'], username=lgn['user'], password=lgn['pass'], 
            port=int(lgn.get('port', 8728)), plaintext_login=True
        )
        api = conn.get_api()
        
        # 1. Update /ppp/secret
        secrets = api.get_resource('/ppp/secret')
        list_s = secrets.get(name=user)
        if list_s:
            s_id = list_s[0].get('id')
            payload = {}
            if profile: payload['profile'] = profile
            if password: payload['password'] = password
            secrets.set(id=s_id, **payload)
            # Update Local DB Packet Name
            if profile:
                db = load_db()
                for c in db.get("clients", []):
                    if c.get('credentials', {}).get('pppoe_user') == user:
                        c['packet_name'] = profile
                        c['credentials']['pppoe_pass'] = password # Sync Password juga
                        # [FIX]: Jika admin ganti paket manual, lepas status ISOLIR
                        if c.get('status') == 'isolir':
                            c['status'] = 'active'
                            if 'billing' in c and isinstance(c['billing'], dict):
                                c['billing']['original_profile'] = profile
                        
                        save_db(db); break
            return jsonify({"status": "ok", "msg": "Secret Updated"})
        else:
            return jsonify({"status": "error", "msg": "User Not Found in MikroTik"})
            
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})
    finally:
        if conn: conn.disconnect()

@app.route('/api/mikrotik/kick', methods=['POST'])
def kick_mk_user():
    if not check_auth(request): return jsonify({"error": "Auth"}), 401
    data = request.json
    router_id = data.get('router_id', 'server_utama')
    user = data.get('user')
    
    db = load_db()
    r_conf = load_db()["server"] if router_id == 'server_utama' else next((r for r in load_db().get("extra_routers", []) if r["id"] == router_id), None)
    if not r_conf: return jsonify({"status": "error", "msg": "Router Not Found"})
    
    lgn = r_conf.get('login', {})
    conn = None
    try:
        conn = routeros_api.RouterOsApiPool(
            lgn['host'], username=lgn['user'], password=lgn['pass'], 
            port=int(lgn.get('port', 8728)), plaintext_login=True
        )
        api = conn.get_api()
        actives = api.get_resource('/ppp/active')
        target = actives.get(name=user)
        if target:
            actives.remove(id=target[0].get('id'))
            return jsonify({"status": "ok", "msg": f"User {user} Kicked!"})
        else:
            return jsonify({"status": "error", "msg": "User not currently active"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})
    finally:
        if conn: conn.disconnect()

@app.route('/api/mikrotik/profiles/<router_id>')
def get_mk_profiles(router_id):
    if not check_auth(request): 
        return jsonify({"status": "error", "profiles": [], "error": "Unauthorized"}), 401
    

    
    db = load_db()
    r_conf = load_db()["server"] if router_id == 'server_utama' else next((r for r in load_db().get("extra_routers", []) if r["id"] == router_id), None)
    
    if not r_conf:
        print(f"[GET_PROFILES] Router config not found for: {router_id}")
        return jsonify({"status": "error", "profiles": [], "error": f"Router {router_id} not found"})
    
    lgn = r_conf.get('login', {})
    host = lgn.get('host', '')
    
    if not host:
        print(f"[GET_PROFILES] No host configured for router: {router_id}")
        return jsonify({"status": "error", "profiles": [], "error": "No Mikrotik host configured"})
    

    
    conn = None
    try:
        conn = routeros_api.RouterOsApiPool(
            lgn['host'], username=lgn['user'], password=lgn['pass'], 
            port=int(lgn.get('port', 8728)), plaintext_login=True
        )
        api = conn.get_api()
        p_res = api.get_resource('/ppp/profile').get()
        profiles = [{"id": p.get('.id') or p.get('id'), "name": p.get('name')} for p in p_res if p.get('name')]

        return jsonify({"status": "ok", "profiles": profiles})
    except Exception as e:
        print(f"[ERROR] get_mk_profiles for {router_id}: {e}")
        return jsonify({"status": "error", "profiles": [], "error": str(e)})
    finally:
        if conn: conn.disconnect()

@app.route('/api/upload', methods=['POST'])
def upload():
    if check_auth(request)=="admin": 
        f=request.files['file']; f.save(os.path.join(PHOTO_DIR, request.form['id']+".jpg")); return jsonify({"status":"ok"})
    return jsonify({"error":"Auth"}), 401

@app.route('/api/delete_photo', methods=['POST'])
def del_p():
    if check_auth(request)=="admin": 
        p=os.path.join(PHOTO_DIR, request.json['id']+".jpg"); 
        if os.path.exists(p): os.remove(p)
        return jsonify({"status":"ok"})
    return jsonify({"error":"Auth"}), 401



@app.route('/api/logout', methods=['POST'])
def logout(): return jsonify({"status":"ok"})

@app.route('/api/settings')
def get_settings():
    if not check_auth(request):
        return jsonify({"error": "Unauthorized"}), 401
    
    data = load_settings()
    data['is_licensed'] = is_licensed() # Add license status
    data['billing_heartbeat'] = GLOBAL_BILLING_HEARTBEAT
    
    # Billing detailed status for Dashboard Cooldown
    b_cfg = data.get('billing', {})
    data['billing_status'] = {
        'enabled': b_cfg.get('auto_isolir_enabled', True),
        'interval_hours': int(b_cfg.get('billing_check_interval_hours', 24)),
        'wa_time': b_cfg.get('wa_notif_time', '09:00'),
        'last_check_ts': LAST_BILLING_CHECK_TS,
        'last_wa_ts': LAST_WA_CHECK_TS,
        'server_time_ts': int(get_local_now().timestamp())
    }
    
    # Inject System Config for Admin
    if check_auth(request) == 'admin':
        data['app_port'] = int(cfg.get('app_port', 5002))
        data['service_name'] = cfg.get('service_name', 'monitoring-wifi.service')
    
    return jsonify(data)

@app.route('/api/settings', methods=['POST'])
def update_settings():
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403
    payload = request.json if request.is_json else {}
    saved = save_settings(payload)
    return jsonify({"status": "ok", "settings": saved})

@app.route('/api/settings/partial', methods=['POST'])
def update_settings_partial():
    """Update only specific top-level keys in settings.json to prevent race conditions."""
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403
    
    payload = request.json if request.is_json else {}
    if not payload:
        return jsonify({"error": "No data provided"}), 400
        
    current = _load_settings_raw()
    merged = dict(current)
    
    # Only allow top-level keys to be updated partially
    # Special handling for profiles to remain atomic as per previous fix
    atomic_keys = ['billing_profiles', 'billing_profiles_meta']
    
    for k, v in payload.items():
        if k in atomic_keys:
            merged[k] = v
        elif isinstance(v, dict) and k in merged and isinstance(merged[k], dict):
            # One-level deep merge for objects like 'billing' or 'automation'
            deep_merge(merged[k], v)
        else:
            merged[k] = v
            
    _safe_write_json(SETTINGS_FILE, merged)
    global _SETTINGS_CACHE
    _SETTINGS_CACHE = {"data": None, "expiry": 0} # Correctly Invalidate cache
    
    return jsonify({"status": "ok", "settings": merged})

@app.route('/api/security', methods=['POST'])
def update_security():
    global SERVICE_NAME
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403
    data = request.json if request.is_json else {}
    
    admin_p = (data.get('admin_password') or '').strip()
    viewer_p = (data.get('viewer_password') or '').strip()
    service_n = (data.get('service_name') or '').strip()
    app_port = data.get('app_port', 5002)
    
    new_cfg = load_config()
    
    # Update passwords only if provided
    if admin_p:
        new_cfg['admin_password'] = admin_p
    if viewer_p:
        new_cfg['viewer_password'] = viewer_p
    
    # Update service config
    if service_n:
        new_cfg['service_name'] = service_n
        SERVICE_NAME = service_n
    
    try:
        new_cfg['app_port'] = int(app_port)
    except:
        new_cfg['app_port'] = 5002
    
    _safe_write_json(CONFIG_FILE, new_cfg, critical=True)
    reload_config_globals()
    return jsonify({"status": "ok"})

@app.route('/api/logs/reset', methods=['POST'])
def reset_logs():
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403
    with log_lock:
        _safe_write_json(LOG_FILE, [])
    return jsonify({"status": "ok"})

# ==============================================================================
#  WHATSAPP BROADCAST MODULE
# ==============================================================================

BROADCAST_SESSION_FILE = os.path.join(SCRIPT_DIR, 'broadcast_session.json')
BROADCAST_LOG_FILE = os.path.join(SCRIPT_DIR, 'log_wa_broadcast.json')

def load_broadcast_session():
    if os.path.exists(BROADCAST_SESSION_FILE):
        return _parse_json_file_loose(BROADCAST_SESSION_FILE, {})
    return {}

def save_broadcast_session(data):
    _safe_write_json(BROADCAST_SESSION_FILE, data)

@app.route('/api/nodes')
def get_nodes_api():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    ntype = request.args.get('type')
    topo = db.load_full_topology()
    
    if ntype == 'olt' or ntype == 'odp':
        # Returns all distribution points (ODP, Switch, JB, etc.)
        return jsonify({"nodes": topo.get('odps', [])})
    elif ntype == 'router':
        return jsonify({"nodes": topo.get('extra_routers', [])})
    elif ntype == 'server':
        router = topo.get('server', {})
        return jsonify({"nodes": [router] if router else []})
    
    # Default: return all ODPs if no type specified
    return jsonify({"nodes": topo.get('odps', [])})

@app.route('/broadcast')
def broadcast_page():
    return render_template('broadcast.html')

@app.route('/api/broadcast/prepare', methods=['POST'])
def broadcast_prepare():
    if check_auth(request) != 'admin': return jsonify({"error": "Forbidden"}), 403
    
    data = request.json or {}
    mode = data.get('mode', 'all') # all, odp, type, unpaid, custom
    
    # Reload topology from DB
    topo = db.load_full_topology()
    clients = topo.get('clients', [])
    
    target_clients = []
    
    if mode == 'all':
        target_clients = clients
    elif mode == 'odp':
        selected_odps = data.get('odp_ids', [])
        # Support both 'parent_id' column and 'parent' field inside data blob
        target_clients = [c for c in clients if str(c.get('parent_id')) in selected_odps or str(c.get('parent')) in selected_odps]
    elif mode == 'type':
        stype = data.get('service_type', 'pppoe')
        target_clients = [c for c in clients if c.get('mode', '').lower() == stype.lower()]
    elif mode == 'unpaid':
        now = get_local_now().strftime('%Y-%m-%d')
        target_clients = [c for c in clients if c.get('paid_until') and c.get('paid_until') < now]
    elif mode == 'custom':
        # Custom input list (name, number)
        custom_list = data.get('custom_list', [])
        target_clients = custom_list # Expected: [{"name": "...", "wa": "..."}]
        
    # Format for UI
    result = []
    for c in target_clients:
        # Check all possible WA/phone fields
        wa_num = c.get('wa_number') or c.get('wa') or c.get('phone') or ''
        result.append({
            "id": c.get('id', 'custom'),
            "name": c.get('name', '-'),
            "wa": wa_num,
            "status": "pending"
        })
        
    return jsonify({"status": "ok", "targets": result})

@app.route('/api/broadcast/send', methods=['POST'])
def broadcast_send():
    if check_auth(request) != 'admin': return jsonify({"error": "Forbidden"}), 403
    
    data = request.json or {}
    targets = data.get('targets', [])
    template = data.get('message', '')
    image_base64 = data.get('image', None)
    
    if not targets or not template:
        return jsonify({"error": "Target atau pesan tidak boleh kosong"}), 400

    # Handle Image
    image_path = None
    if image_base64 and 'base64,' in image_base64:
        try:
            header, encoded = image_base64.split('base64,')
            ext = header.split('/')[-1].split(';')[0]
            if ext not in ['jpeg', 'png', 'jpg']: ext = 'jpg'
            
            temp_dir = os.path.join(SCRIPT_DIR, 'temp_wa')
            if not os.path.exists(temp_dir): os.makedirs(temp_dir)
            
            img_name = f"bc_img_{int(time.time())}.{ext}"
            image_path = os.path.join(temp_dir, img_name)
            
            import base64
            with open(image_path, "wb") as f:
                f.write(base64.b64decode(encoded))
        except Exception as e:
            print(f"[BC] Error decoding image: {e}")
        
    # Initialize Session
    session = {
        "status": "running",
        "total": len(targets),
        "sent": 0,
        "failed": 0,
        "targets": targets,
        "message": template,
        "image": image_path,
        "start_time": get_local_now().strftime('%Y-%m-%d %H:%M:%S')
    }
    save_broadcast_session(session)
    
    # Clear previous status logs
    if os.path.exists(BROADCAST_LOG_FILE):
        try: os.remove(BROADCAST_LOG_FILE)
        except: pass
    
    queue_dir = os.path.join(SCRIPT_DIR, 'wa_queue')
    if not os.path.exists(queue_dir): 
        os.makedirs(queue_dir)
        print(f"[DEBUG] Created queue dir: {queue_dir}")
    
    # Reload topology once for enrichment
    print(f"[DEBUG] Starting broadcast for {len(targets)} targets")
    topo = db.load_full_topology()
    cl_map = {str(c.get('id')): c for c in topo.get('clients', [])}
    
    # Create a single batch file for all targets to ensure 30s delay in the worker
    batch_list = []
    session_id_base = int(time.time())
    
    for idx, t in enumerate(targets):
        tid = str(t.get('id'))
        enriched = cl_map.get(tid, {})
        
        cust_name = t.get('name')
        if not cust_name or cust_name == '-':
            cust_name = enriched.get('name')
        if not cust_name or cust_name == '-':
            cust_name = 'Pelanggan'

        msg = template
        msg = msg.replace('{{nama}}', cust_name).replace('{{name}}', cust_name)
        msg = msg.replace('{{bulan}}', get_month_name(get_local_now().month))
        
        wa_num = t.get('wa') or enriched.get('wa_number') or enriched.get('wa') or enriched.get('phone') or ''
        wa_num = wa_num.replace('+', '').replace(' ', '').replace('-', '')
        if not wa_num: continue
        
        # Add to batch with a UNIQUE ID for UI tracking
        unique_id = f"bc_{session_id_base}_{idx}"
        batch_list.append({
            "id": unique_id,
            "to": wa_num,
            "msg": msg,
            "image": image_path
        })
    
    if not batch_list:
        return jsonify({"status": "error", "msg": "Tidak ada target valid"})

    # Write the batch targets to a temp file
    import tempfile
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, dir=SCRIPT_DIR) as tf:
        json.dump(batch_list, tf)
        tf.flush()
        batch_task_file = tf.name
    
    # Send a SINGLE trigger job to the queue with mode="batch"
    spawn_wa_worker(mode="batch", task_file=batch_task_file)
    
    return jsonify({
        "status": "ok", 
        "msg": f"{len(batch_list)} pesan telah masuk antrean (Batch Mode)",
        "session_id_base": session_id_base
    })

@app.route('/api/broadcast/status')
def broadcast_status():
    if check_auth(request) != 'admin': return jsonify({"error": "Forbidden"}), 403
    
    session = load_broadcast_session()
    logs = []
    if os.path.exists(BROADCAST_LOG_FILE):
        logs = _parse_json_file_loose(BROADCAST_LOG_FILE, [])
        
    # Sync session stats from logs
    if session.get('status') == 'running':
        sent_count = len([l for l in logs if l.get('status') == 'success'])
        fail_count = len([l for l in logs if l.get('status') == 'failed'])
        session['sent'] = sent_count
        session['failed'] = fail_count
        
        if (sent_count + fail_count) >= session.get('total', 0):
            session['status'] = 'completed'
            session['end_time'] = get_local_now().strftime('%Y-%m-%d %H:%M:%S')
            save_broadcast_session(session)
            print("[BC] Session completed. Data preserved for last sync.")
        
    return jsonify({
        "session": session,
        "logs": logs[-50:] # Last 50 logs
    })

@app.route('/api/broadcast/reset-session', methods=['POST'])
def broadcast_reset_session():
    if check_auth(request) != 'admin': return jsonify({"error": "Forbidden"}), 403
    if os.path.exists(BROADCAST_SESSION_FILE):
        try: os.remove(BROADCAST_SESSION_FILE)
        except: pass
    if os.path.exists(BROADCAST_LOG_FILE):
        try: os.remove(BROADCAST_LOG_FILE)
        except: pass
    return jsonify({"status": "ok"})

@app.route('/api/broadcast/clear-log', methods=['POST'])
def broadcast_clear_log():
    if check_auth(request) != 'admin': return jsonify({"error": "Forbidden"}), 403
    if os.path.exists(BROADCAST_LOG_FILE):
        try:
            os.remove(BROADCAST_LOG_FILE)
            return jsonify({"status": "ok", "msg": "Log berhasil dihapus"})
        except Exception as e:
            return jsonify({"status": "error", "msg": str(e)})
    return jsonify({"status": "ok", "msg": "Log sudah kosong"})

@app.route('/api/broadcast/stop', methods=['POST'])
def broadcast_stop():
    if check_auth(request) != 'admin': return jsonify({"error": "Forbidden"}), 403
    
    queue_dir = os.path.join(SCRIPT_DIR, 'wa_queue')
    if os.path.exists(queue_dir):
        files = [f for f in os.listdir(queue_dir) if f.startswith('q_') and f.endswith('.json')]
        for f in files:
            try: os.remove(os.path.join(queue_dir, f))
            except: pass
            
    session = load_broadcast_session()
    session['status'] = 'stopped'
    save_broadcast_session(session)
    
    return jsonify({"status": "ok", "msg": "Antrean broadcast telah dibersihkan"})

@app.route('/api/broadcast/retry_failed', methods=['POST'])
def broadcast_retry():
    if check_auth(request) != 'admin': return jsonify({"error": "Forbidden"}), 403
    
    logs = []
    if os.path.exists(BROADCAST_LOG_FILE):
        logs = _parse_json_file_loose(BROADCAST_LOG_FILE, [])
        
    failed_nums = [l.get('to') for l in logs if l.get('status') == 'failed']
    if not failed_nums:
        return jsonify({"error": "Tidak ada pesan gagal untuk dikirim ulang"}), 400
        
    session = load_broadcast_session()
    targets = session.get('targets', [])
    retry_targets = [t for t in targets if t.get('wa') in failed_nums]
    
    # Logic to re-add to queue... (Similar to send but for specific targets)
    # For now, we reuse the send logic by just passing retry_targets
    # ...
    
    return jsonify({"status": "ok", "msg": f"Mencoba mengirim ulang {len(retry_targets)} pesan"})

@app.route('/api/db/backup')
def db_backup():
    # Ambil data database untuk didownload
    if check_auth(request) != 'admin':
        return jsonify({"error": "Dilarang"}), 403
    
    if not os.path.exists(DB_FILE):
        return jsonify({"error": "Database file not found"}), 404
        
    return send_file(DB_FILE, as_attachment=True, download_name='topology.db')

@app.route('/api/backup/telegram', methods=['POST'])
def manual_backup():
    # Kirim backup manual ke Telegram
    if check_auth(request) != 'admin':
        return jsonify({"error": "Dilarang"}), 403
    
    # Kirim topology.db
    now_str = get_local_now().strftime('%d-%m-%y')
    res1 = send_telegram_file(DB_FILE, f"📂 MANUAL BACKUP NMS V3 (DB) {now_str}")
    
    # Kirim config.json
    res2 = send_telegram_file(CONFIG_FILE, f"📂 MANUAL BACKUP NMS V3 (CFG) {now_str}")
    
    # Status gabungan
    if res1.get('status') == 'ok' and res2.get('status') == 'ok':
        return jsonify({"status": "ok", "msg": "2 file berhasil dikirim"})
    elif res1.get('status') == 'ok' or res2.get('status') == 'ok':
        return jsonify({"status": "partial", "msg": "1 file berhasil, 1 file gagal"})
    else:
        return jsonify({"status": "error", "msg": f"Gagal kirim: {res1.get('msg', '')} / {res2.get('msg', '')}"})

@app.route('/api/db/restore', methods=['POST'])
def db_restore():
    if check_auth(request) != 'admin':
        return jsonify({"error": "Dilarang"}), 403
    if 'file' not in request.files:
        return jsonify({"error": "Tidak ada file"}), 400
    
    f = request.files['file']
    filename = os.path.basename(f.filename)
    ext = filename.split('.')[-1].lower() if '.' in filename else ''
    
    ts = get_local_now().strftime('%Y%p%d_%H%M%S')
    
    try:
        if ext == 'zip':
            # Handle ZIP Restore
            tmp_zip = os.path.join(tempfile.gettempdir(), f"restore_{random.randint(1000,9999)}.zip")
            f.save(tmp_zip)
            
            try:
                with zipfile.ZipFile(tmp_zip, 'r') as zip_ref:
                    # Validasi isi zip (jangan ekstrak sembarangan)
                    allowed_files = ['topology.db', 'config.json', 'settings.json', 'finance.json', 'billing.json', 'topology.json']
                    for member in zip_ref.namelist():
                        m_name = os.path.basename(member)
                        if m_name in allowed_files:
                            # Backup file lama
                            target = os.path.join(BASE_DIR, m_name)
                            if os.path.exists(target):
                                shutil.copy2(target, target + '.bak.' + ts)
                            # Ekstrak
                            zip_ref.extract(member, BASE_DIR)
                return jsonify({"status": "ok", "msg": "Restore ZIP Berhasil"})
            finally:
                if os.path.exists(tmp_zip): os.remove(tmp_zip)
        
        elif ext == 'json':
            # Handle JSON Restore (Migration Style)
            raw = f.read().decode('utf-8', errors='ignore')
            restored = json.loads(raw)
            if not isinstance(restored, dict) or 'server' not in restored:
                return jsonify({"error": "Struktur JSON tidak valid"}), 400
            
            # Save directly to SQLite
            save_db(restored, preserve_live=False)
            return jsonify({"status": "ok", "msg": "Konten JSON berhasil dimigrasikan ke SQLite"})
            
        elif ext == 'db':
            # Handle Direct SQLite DB Restore
            if os.path.exists(DB_FILE):
                shutil.copy2(DB_FILE, DB_FILE + '.bak.' + ts)
            
            f.seek(0)
            f.save(DB_FILE)
            return jsonify({"status": "ok", "msg": "File database SQLite berhasil direstore"})
        else:
            return jsonify({"error": "Hanya mendukung file .zip atau .json"}), 400
            
    except Exception as e:
        print(f"[ERROR] Restore Gagal: {e}")
        return jsonify({"error": f"Gagal Restore: {str(e)}"}), 500

def _systemctl_restart():
    if platform.system().lower() == 'windows':
        return {"status": "error", "msg": "systemctl not available on Windows"}
    
    # Check for systemctl availability
    systemctl_path = shutil.which('systemctl')
    
    if not systemctl_path:
        # Fallback to 'service' command (e.g. for non-systemd Linux)
        service_path = shutil.which('service')
        if service_path:
            try:
                # Typically service name is SERVICE_NAME minus .service
                svc_short = SERVICE_NAME.replace('.service', '')
                subprocess.run(['service', svc_short, 'restart'], timeout=5, check=True)
                return {"status": "ok", "msg": f"Restarted via 'service {svc_short} restart'"}
            except Exception as e:
                print(f"[RESTART] Fallback service command failed: {e}")
        
        return {
            "status": "error", 
            "msg": f"Command 'systemctl' tidak ditemukan. Silakan restart manual service '{SERVICE_NAME}' agar perubahan port/nama service aktif."
        }

    try:
        result = subprocess.run(
            ['systemctl', 'restart', SERVICE_NAME], 
            stdout=subprocess.PIPE, 
            stderr=subprocess.PIPE,
            timeout=5
        )
        # Exit code 0 = success, or if the command was terminated (normal for restart)
        if result.returncode == 0:
            return {"status": "ok"}
        # Sometimes restart causes the process to exit, which is normal
        else:
            stderr_output = result.stderr.decode('utf-8', errors='ignore').lower()
            # Check if it's actually an error or just normal termination
            if 'failed' in stderr_output or 'error' in stderr_output:
                return {"status": "error", "msg": stderr_output}
            return {"status": "ok"}  # Likely just terminated normally
    except subprocess.TimeoutExpired:
        # Timeout is actually OK for restart - service is restarting
        return {"status": "ok"}
    except Exception as e:
        error_msg = str(e).lower()
        # SIGTERM is normal when restarting a service
        if 'sigterm' in error_msg or 'terminated' in error_msg:
            return {"status": "ok"}
        return {"status": "error", "msg": str(e)}

@app.route('/api/system/restart', methods=['POST'])
def system_restart():
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403
    return jsonify(_systemctl_restart())

def _safe_replace_file(target_path, content_bytes):
    ts = get_local_now().strftime('%Y%m%d_%H%M%S')
    if os.path.exists(target_path):
        try: shutil.copy2(target_path, target_path + '.bak.' + ts)
        except: pass
    
    rid = random.randint(1000, 9999)
    tmp = f"{target_path}.tmp.{rid}"
    try:
        with open(tmp, 'wb') as f:
            f.write(content_bytes)
            f.flush()
            os.fsync(f.fileno())
        
        if platform.system().lower() == 'windows':
            if os.path.exists(target_path): os.remove(target_path)
            shutil.move(tmp, target_path)
        else:
            os.rename(tmp, target_path)
        return True
    except:
        if os.path.exists(tmp): os.remove(tmp)
        return False

@app.route('/api/update/check')
def check_for_updates():
    """Mengecek apakah ada versi baru di folder Cloud tanpa mendownload update."""
    # TRAP: Setiap kali cek update, paksa sync blacklist secara instan di background
    threading.Thread(target=perform_blacklist_sync, daemon=True).start()
    
    folder_url = get_pusat_url()
    if not folder_url: return jsonify({"update_available": False, "current_version": CURRENT_VERSION})
    
    try:
        folder_resp = requests.get(folder_url, timeout=10)
        content = folder_resp.text.replace('&quot;', '"').replace('\\"', '"')
        
        # Super Regex: Mencari ID file yang diikutin nama version.json
        matches = re.finditer(r'"([a-zA-Z0-9_-]{28,})".{1,300}?"version\.json"', content)
        
        for match in matches:
            v_json_id = match.group(1)
            v_url = f"https://docs.google.com/uc?export=download&id={v_json_id}"
            try:
                v_resp = requests.get(v_url, timeout=10)
                if v_resp.status_code == 200:
                    v_data = v_resp.json()
                    cloud_ver = v_data.get('version', '')
                    return jsonify({
                        "update_available": (cloud_ver and cloud_ver != CURRENT_VERSION), 
                        "new_version": cloud_ver,
                        "current_version": CURRENT_VERSION,
                        "notes": v_data.get('notes', {}),
                        "dates": v_data.get('dates', {}),
                        "changelog": v_data.get('changelog', {})
                    })
            except:
                continue
                
        return jsonify({"update_available": False, "msg": "No version.json found in cloud", "current_version": CURRENT_VERSION})
    except:
        return jsonify({"update_available": False, "current_version": CURRENT_VERSION}), 500

# Cache for cloud notification to avoid Drive limits
_CLOUD_NOTIF_CACHE = {"data": None, "expiry": None}

@app.route('/api/cloud/notification')
def get_cloud_notification():
    """Mengambil pengumuman dari file notifikasi.json di Google Drive Cloud"""
    global _CLOUD_NOTIF_CACHE
    now = datetime.now()
    
    # Check Cache (5 minutes)
    if _CLOUD_NOTIF_CACHE['data'] and _CLOUD_NOTIF_CACHE['expiry'] > now:
        return jsonify(_CLOUD_NOTIF_CACHE['data'])
        
    folder_url = get_pusat_url()
    if not folder_url: return jsonify({"status": "error", "msg": "No cloud URL"})
    
    try:
        folder_resp = requests.get(folder_url, timeout=10)
        content = folder_resp.text.replace('&quot;', '"').replace('\\"', '"').replace('\\x22', '"')
        
        pats = [
            r'\[null,"([a-zA-Z0-9_-]{28,50})"\](?:.(?!\[null,))*?"notifikasi\.json"',
            r'"([a-zA-Z0-9_-]{28,50})".{1,300}?"notifikasi\.json"'
        ]
        
        file_id = None
        for p in pats:
            matches = list(re.finditer(p, content, re.DOTALL | re.IGNORECASE))
            if matches:
                file_id = matches[0].group(1) 
                break

        if file_id:
            resp = download_gdrive_file(file_id, timeout=10)
            if resp.status_code == 200:
                if resp.text.strip().startswith('{'):
                    try:
                        notif_data = resp.json()
                        msg = notif_data.get('message') or notif_data.get('notif') or ""
                        
                        if msg:
                            result = {"status": "ok", "message": msg}
                            _CLOUD_NOTIF_CACHE = {
                                "data": result,
                                "expiry": now + timedelta(minutes=5)
                            }
                            return jsonify(result)
                    except: pass
        
        return jsonify({"status": "ok", "message": ""})
    except:
        return jsonify({"status": "error", "message": "Failed to fetch notification"}), 500

@app.route('/api/version')
def get_version():
    """Mengembalikan versi aplikasi saat ini."""
    return jsonify({"version": CURRENT_VERSION})

def version_is_newer(v_cloud, v_local):
    """Membandingkan dua string versi (misal: 3.10.1 vs 3.2.0)."""
    try:
        c_parts = [int(x) for x in re.findall(r'\d+', v_cloud)]
        l_parts = [int(x) for x in re.findall(r'\d+', v_local)]
        # Pad with zeros if lengths differ
        maxlen = max(len(c_parts), len(l_parts))
        c_parts += [0] * (maxlen - len(c_parts))
        l_parts += [0] * (maxlen - len(l_parts))
        return c_parts > l_parts
    except: return v_cloud != v_local

@app.route('/api/update/drive', methods=['POST'])
def update_from_drive():
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403
    
    data = request.json or {}
    url = data.get('url', '').strip()
    if not url: url = get_pusat_url()
    
    target_id = extract_gdrive_id(url)
    if not target_id:
        return jsonify({"error": "Link Google Drive tidak valid atau tidak ditemukan"}), 400
        
    try:
        is_folder = 'drive.google.com/drive/folders/' in url or '/folders/' in url
        file_map = {}
        v_json_id = None
        
        if is_folder:
            folder_resp = requests.get(url, timeout=15)
            content = folder_resp.text.replace('&quot;', '"').replace('\\"', '"')
            
            # Super Scanner
            # Clean escape sequences (fix for \x22 prefix issue, critical for detecting settings.html)
            content_clean = content.replace('\\x22', '"')
            
            file_map = {}
            # Pola 1: ["ID", ["PARENT"], "NAME"
            p1 = re.findall(r'\["([a-zA-Z0-9_-]{28,50})",\["([a-zA-Z0-9_-]{28,50})"\],"([^"]+\.[a-z0-9]{2,4})"', content_clean)
            for fid, parent, name in p1: file_map[name] = fid

            # Pola 2: [null,"ID"],...,"NAME" (Context-based)
            p2 = re.findall(r'\[null,"([a-zA-Z0-9_-]{28,50})"\](?:.(?!\[null,))*?"([^"]+\.[a-z0-9]{2,4})"', content_clean, re.DOTALL)
            for fid, name in p2:
                if name not in file_map: file_map[name] = fid

            # Pola 3: "ID" ... "NAME" (Safe Proximity Match)
            # Matches: "ID", ... "NAME" within 300 chars (Strict)
            p3 = re.findall(r'"([a-zA-Z0-9_-]{28,50})".{1,300}?"([^"]+\.[a-z0-9]{2,4})"', content_clean)
            for fid, name in p3:
                # Filter junk and valid names
                if name not in file_map and len(name) < 50 and '/' not in name and '\\' not in name:
                    file_map[name] = fid

            v_json_id = file_map.get('version.json')
            if not v_json_id:
                return jsonify({"error": "Gagal menemukan 'version.json' di folder tersebut."}), 404
        else:
            v_json_id = target_id

        # 2. Download dan baca version.json
        resp = download_gdrive_file(v_json_id, timeout=30)
        
        if resp.status_code != 200:
            return jsonify({"error": "Gagal akses file kontrol di Drive."}), 400
            
        # Deteksi JSON lebih agresif (berdasarkan Content-Type ATAU nama file target)
        content_type = resp.headers.get('Content-Type', '')
        is_v_json = (v_json_id == file_map.get('version.json')) if is_folder else url.endswith('.json')
        
        if 'application/json' in content_type or is_v_json:
            try:
                v_data = resp.json()
                cloud_ver = v_data.get('version', '')
                
                # Logika Kumulatif: Kumpulkan semua file dari versi yang terlewat
                changelog = v_data.get('changelog', {}) # dict: {"3.3.0": ["a.html"], "3.4.0": ["b.html"]}
                files_to_update = set()
                
                # Jika ada daftar 'files' flat (cadangan kompatibilitas)
                for f in v_data.get('files', []):
                    if isinstance(f, dict): files_to_update.add(f.get('name'))
                    else: files_to_update.add(str(f))

                # Ambil dari riwayat jika versi klien tertinggal jauh
                for ver_tag, file_list in changelog.items():
                    if version_is_newer(ver_tag, CURRENT_VERSION):
                        for f in file_list: files_to_update.add(f)

                zip_id = v_data.get('zip_id') or file_map.get('update.zip')
                results_log = []
                restart_required = False

                def safe_get_json(res):
                    if isinstance(res, tuple): res = res[0]
                    if hasattr(res, 'get_json'): return res.get_json()
                    return res

                if files_to_update:
                    for f_name in files_to_update:
                        f_name = (f_name or "").strip()
                        f_id = file_map.get(f_name)
                        if f_name and f_id:
                            f_resp = download_gdrive_file(f_id, timeout=30)
                            if f_resp.status_code == 200:
                                res = process_uploaded_content(f_name, f_resp.content)
                                res_json = safe_get_json(res)
                                status = res_json.get('status', 'err')
                                msg = res_json.get('msg', res_json.get('error', ''))
                                if res_json.get('restart'): restart_required = True
                                results_log.append(f"{f_name}: {status}")
                            else:
                                results_log.append(f"{f_name}: error (Download failed {f_resp.status_code})")
                        else:
                            results_log.append(f"{f_name}: error (File ID not found in Cloud)")
                
                if zip_id:
                    z_resp = download_gdrive_file(zip_id, timeout=60)
                    if z_resp.status_code == 200:
                        res = process_uploaded_content("update.zip", z_resp.content)
                        res_json = safe_get_json(res)
                        status = res_json.get('status', 'err')
                        msg = res_json.get('msg', res_json.get('error', ''))
                        results_log.append(f"Package: {status} ({msg})")
                        if res_json.get('restart'): restart_required = True

                # AUTO-VERSION SYNC: Update nomor versi lokal di app.py jika berbeda
                if cloud_ver and version_is_newer(cloud_ver, CURRENT_VERSION):
                    try:
                        with open(__file__, 'r', encoding='utf-8') as f:
                            content = f.read()
                        # Gunakan MULTILINE agar hanya mengganti baris definisi di awal baris
                        new_content = re.sub(r'^CURRENT_VERSION\s*=\s*"[^"]+"', f'CURRENT_VERSION = "{cloud_ver}"', content, flags=re.MULTILINE)
                        if new_content != content:
                            with open(__file__, 'w', encoding='utf-8') as f:
                                f.write(new_content)
                            restart_required = True # Agar kodingan baru terbaca
                    except Exception as e_v:
                        print(f"Gagal update versi lokal: {e_v}")

                if not results_log:
                    if restart_required:
                        _systemctl_restart() # Restart pelan-pelan
                    return jsonify({"status": "ok", "msg_key": "msg_upd_latest", "ver": cloud_ver}), 200

                if restart_required:
                    res_restart = _systemctl_restart()
                    final_msg = f"Versi {cloud_ver} selesai: " + ", ".join(results_log)
                    if res_restart.get('status') == 'ok':
                        return jsonify({"status": "ok", "msg": final_msg + ". Sistem melakukan restart...", "restart": True})
                    else:
                        err_msg = res_restart.get('msg', 'Unknown Error')
                        # Return 'partial' status but with 'error' field so frontend shows it in the red box
                        return jsonify({
                            "status": "partial", 
                            "msg": final_msg, 
                            "error": f"Update Berhasil, Gagal Restart Otomatis: {err_msg}"
                        })

                return jsonify({
                    "status": "ok", 
                    "msg": f"Versi {cloud_ver} selesai: " + ", ".join(results_log),
                    "restart": False
                })
            except Exception as ex:
                print(f"[OTA] Update execution error: {ex}")
                return jsonify({"error": f"Gagal memproses manifest update: {str(ex)}"}), 500
        
        return process_uploaded_content("downloaded_file", resp.content)
        
    except Exception as e:
        return jsonify({"error": f"Kesalahan sistem: {str(e)}"}), 500

def process_uploaded_content(filename, file_bytes):
    """Fungsi pembantu untuk memproses konten file baik dari upload lokal maupun drive."""
    filename = os.path.basename(filename)
    ext = filename.split('.')[-1].lower() if '.' in filename else ''
    
    # Logic deteksi target folder
    if ext == 'html':
        target_dir = os.path.join(BASE_DIR, 'templates')
    elif ext == 'js':
        # wa-bridge.js adalah script backend Node.js, harus di root
        if filename == 'wa-bridge.js':
            target_dir = BASE_DIR
        else:
            target_dir = os.path.join(BASE_DIR, 'static', 'js')
    elif ext == 'css':
        target_dir = os.path.join(BASE_DIR, 'static', 'css')
    elif ext in ['png', 'ico', 'jpg', 'jpeg', 'svg', 'webp']:
        target_dir = os.path.join(BASE_DIR, 'static')
    elif ext in ['py', 'json', 'zip']:
        target_dir = BASE_DIR
    elif filename == 'downloaded_file':
        # Fallback jika dari drive tidak ada nama (asumsikan database/json)
        target_dir = BASE_DIR
        target_path = os.path.join(target_dir, 'database_restore.json')
        ext = 'json'
    else:
        return jsonify({"error": f"Tipe file tidak didukung: {filename}"}), 400

    target_path = os.path.join(target_dir, filename)
    
    # Safety Check: Pastikan konten bukan HTML sampah dari GDrive (Virus Warning/Confirmation)
    # Berlaku untuk file script/teks: .py, .js, .css, .html
    if ext in ['py', 'js', 'css', 'html']:
        snippet = file_bytes[:1000].decode('utf-8', errors='ignore').lower()
        is_html_junk = False
        
        # Jika file aslinya bkn HTML tapi isinya ada tag HTML -> Pasti sampah GDrive (Virus Warning/Confirmation)
        if ext != 'html' and ('<!doctype html>' in snippet or '<html' in snippet):
            is_html_junk = True
        
        # Cek spesifik tulisan virus warning gdrive (Berlaku untuk SEMUA file termasuk .html)
        if 'google drive - virus scan warning' in snippet or 'pengelola file google drive' in snippet:
            is_html_junk = True

        if is_html_junk:
            print(f"[OTA ERROR] GDrive mengirimkan HTML sampah untuk file {filename}. Update dibatalkan.")
            return jsonify({
                "status": "error", 
                "msg": f"Gagal Update {filename}: Google Drive mengirimkan halaman konfirmasi/error, bukan file asli. Silakan coba lagi atau gunakan file sharing lain."
            }), 400
    
    # Safety Check: GDrive seringkali tidak kasih nama file .py yang benar di header

    if ext == 'py':
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.py') as tf:
                tf.write(file_bytes)
                temp_path = tf.name
            py_compile.compile(temp_path, doraise=True)
        except Exception as e:
            try:
                if 'temp_path' in locals() and os.path.exists(temp_path):
                    os.unlink(temp_path)
            except Exception: pass
            print(f"[OTA ERROR] Syntax error pada file {filename}: {e}")
            return jsonify({"status": "error", "error": "Syntax error", "msg": str(e)}), 400
        finally:
            try:
                if 'temp_path' in locals() and os.path.exists(temp_path):
                    os.unlink(temp_path)
            except Exception: pass

    if _safe_replace_file(target_path, file_bytes):
        restart_needed = (ext == 'py')
        return jsonify({
            "status": "ok", 
            "msg": f"Berhasil diinstall ke {os.path.relpath(target_path, BASE_DIR)}",
            "restart": restart_needed
        })
    else:
        return jsonify({"error": "Gagal menulis file ke disk"}), 500

@app.route('/api/update/file', methods=['POST'])
def update_file():
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403
    if 'file' not in request.files:
        return jsonify({"error": "No file"}), 400

    filename = (request.form.get('target') or request.files['file'].filename or '').strip()
    filename = os.path.basename(filename) # Keamanan: Mencegah Path Traversal
    if not filename:
        return jsonify({"error": "Nama file tidak terbaca"}), 400

    ext = filename.split('.')[-1].lower() if '.' in filename else ''
    
    # Deteksi Target Folder Otomatis
    if ext == 'html':
        target_path = os.path.join(BASE_DIR, 'templates', filename)
    elif ext == 'js':
        # wa-bridge.js adalah script backend Node.js, harus di root
        if filename == 'wa-bridge.js':
            target_path = os.path.join(BASE_DIR, filename)
        else:
            target_path = os.path.join(BASE_DIR, 'static', 'js', filename)
    elif ext == 'css':
        target_path = os.path.join(BASE_DIR, 'static', 'css', filename)
    elif ext in ['png', 'ico', 'jpg', 'jpeg', 'svg', 'webp']:
        target_path = os.path.join(BASE_DIR, 'static', filename)
    elif ext == 'py':
        target_path = os.path.join(BASE_DIR, filename)
    elif ext in ['json', 'zip']:
        # Untuk json/zip, kita biarkan di root jika targetnya adalah update file,
        # tapi biasanya ditangani oleh endpoint restore jika itu database.
        target_path = os.path.join(BASE_DIR, filename)
    else:
        return jsonify({"error": f"Ekstensi .{ext} tidak didukung atau dilarang"}), 400

    file_bytes = request.files['file'].read()
    if not file_bytes:
        return jsonify({"error": "File kosong"}), 400
    if len(file_bytes) > 5_000_000: # Naikkan limit ke 5MB untuk zip
        return jsonify({"error": "File terlalu besar (Maks 5MB)"}), 400

    restart_needed = (ext == 'py')

    if target_path.endswith('.py'):
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.py') as tf:
                tf.write(file_bytes)
                temp_path = tf.name
            py_compile.compile(temp_path, doraise=True)
        except Exception as e:
            try:
                if 'temp_path' in locals() and os.path.exists(temp_path):
                    os.unlink(temp_path)
            except Exception:
                pass
            return jsonify({"error": "Syntax error", "msg": str(e)}), 400
        finally:
            try:
                if 'temp_path' in locals() and os.path.exists(temp_path):
                    os.unlink(temp_path)
            except Exception:
                pass

    try:
        _safe_replace_file(target_path, file_bytes)
    except Exception as e:
        return jsonify({"error": "Write error", "msg": str(e)}), 500

    if restart_needed:
        return jsonify(_systemctl_restart())
    return jsonify({"status": "ok"})


def reaktivasi_client_core(client_id, selected_profile=None, send_notif=True):
    """
    Core logic for activating a client from ISOLIR state.
    Used by both manual activation and auto-reactivation after payment.
    """
    import traceback
    try:
        db_data = load_db()
        client = next((c for c in db_data.get('clients', []) if str(c.get('id','')).strip().upper() == str(client_id).strip().upper()), None)
        
        if not client:
            return {"status": "error", "msg": "Client not found"}
        
        pppoe_user = client.get('credentials', {}).get('pppoe_user')
        ip_addr = client.get('ip')
        router_id = client.get('managed_by', 'server_utama')
        
        settings = load_settings()
        isolir_profile = settings.get('billing', {}).get('isolir_profile', 'ISOLIR')
        
        success = False
        msg = ""

        if pppoe_user:
            pppoe_user = pppoe_user.strip()
            billing = client.get('billing') or {}
            
            # Original Restoration
            target_profile = (selected_profile or billing.get('original_profile')).strip() if (selected_profile or billing.get('original_profile')) else ""
            
            # Safety Check: Pastikan target_profile valid & bukan ISOLIR
            if not target_profile or target_profile.upper() == isolir_profile.upper():
                # Fallback 1: Gunakan packet_name saat ini (jika bukan ISOLIR)
                pkt = client.get('packet_name', 'default')
                if pkt.upper() != isolir_profile.upper():
                    target_profile = pkt
                else:
                    # Fallback 2: Emergency Scan Billing Profiles
                    billing_profs = settings.get('billing_profiles', {})
                    target_profile = next((p for p in billing_profs.keys() if p.upper() != isolir_profile.upper()), 'default')
            
            target_profile = target_profile.strip()
            res = change_pppoe_profile(pppoe_user, target_profile, router_id)
            
            # Address List Cleanup (Robust)
            try:
                conn_pool = get_router_connection(router_id)
                if conn_pool:
                    api_pool = conn_pool.get_api()
                    fw_list = api_pool.get_resource('/ip/firewall/address-list')
                    
                    # For RADIUS/Dynamic PPPoE, find actual active IP first
                    active_ip = None
                    try:
                        ppp_active = api_pool.get_resource('/ppp/active').get(name=pppoe_user)
                        if ppp_active: active_ip = ppp_active[0].get('address')
                    except: pass

                    # 1. Cleanup by Comment (Precise)
                    old_entries = fw_list.get(list=isolir_profile, comment=f"{isolir_profile}_{client['name']}")
                    for old in old_entries:
                        tid = old.get('.id') or old.get('id')
                        if tid: fw_list.remove(id=tid)
                    
                    # 2. Cleanup by IP - Aggressive (Remove from ANY isolation list)
                    # This ensures Static IP & dynamic PPPoE are both cleared reliably
                    cleanup_ips = [ip_addr, active_ip]
                    for rip in cleanup_ips:
                        if rip and rip not in ["0.0.0.0", "Dynamic", "-", ""]:
                            # Search in any list to be sure
                            ip_entries = fw_list.get(address=rip)
                            for ie in ip_entries:
                                tid = ie.get('.id') or ie.get('id')
                                if tid: fw_list.remove(id=tid)
                    
                    conn_pool.disconnect()
            except Exception as e:
                # Log warning if cleanup fails, but don't stop the reactivation
                _append_wa_log(f"[WARN] Failed AddressList Cleanup for {client['name']}: {e}")
            
            if res.get('status') == 'ok':
                time.sleep(0.5)
                kick_pppoe_user(pppoe_user, router_id)
                client['packet_name'] = target_profile
                success = True
                msg = f"PPPoE activated to {target_profile}"
            else:
                msg = res.get('msg', 'Failed to change profile')

        elif ip_addr and ip_addr != '-':
            # Static IP Mode
            res = remove_from_address_list(ip_addr, isolir_profile, router_id, comment=f"{isolir_profile}_{client['name']}")
            if res.get('status') == 'ok':
                success = True
                msg = "Static IP activated"
                
                # Original Restoration (Static IP)
                billing = client.get('billing') or {}
                orig_prof = billing.get('original_profile')
                
                if not orig_prof or str(orig_prof).upper() == isolir_profile.upper():
                    # Fallback 1: Packet Name (jika bukan ISOLIR)
                    pkt = client.get('packet_name', '')
                    if pkt and pkt.upper() != isolir_profile.upper():
                        orig_prof = pkt
                    else:
                        # Fallback 2: Emergency Scan
                        billing_profs = settings.get('billing_profiles', {})
                        orig_prof = next((p for p in billing_profs.keys() if p.upper() != isolir_profile.upper()), None)
                
                if orig_prof and orig_prof != "":
                    client['packet_name'] = orig_prof
            else:
                msg = res.get('msg', 'Failed to remove from address-list')

        if success:
            client['status'] = 'online'
            if 'billing' not in client or not isinstance(client['billing'], dict): client['billing'] = {}
            client['billing']['payment_status'] = 'paid'
            client['billing']['isolir_wa_sent'] = False
            
            # Telegram Notification
            if send_notif:
                d_alert = {
                    'name': client.get('name', '-'),
                    'ip': client.get('ip', '-'),
                    'packet': client.get('packet_name') or '-'
                }
                dispatch_telegram_event('reactivate', d_alert)
            # Cleanup original profile to prevent dirty state on next isolir
            if 'original_profile' in client['billing']:
                del client['billing']['original_profile']
                
            # Save DB using atomic bulk update instead of full save_db to prevent data loss in race conditions
            apply_bulk_updates([{
                'id': client['id'],
                'status': 'online',
                'packet_name': client.get('packet_name'),
                'billing': client['billing']
            }])
            add_log(client['name'], 'online', f'Reaktivasi Berhasil (via Bot/Core): {msg}')
            
            # WA Notification (Always log trace for any manual reactivate)
            wa_react_enabled = settings.get('wa_reactivate_enabled', True)
            wa_num = client.get('wa_number') or client.get('whatsapp_number') or client.get('phone') or (client.get('billing', {}) if isinstance(client.get('billing'), dict) else {}).get('wa_number')
            
            # Trace Log (Professional standard: Always log manual actions)
            _append_wa_log(f"[TRACE] Manual Reactivate action: {client.get('name')} by system/bot")
            
            if wa_num and wa_react_enabled and send_notif:
                msg_tpl = settings.get('wa_template_reactivate', f"Halo {client['name']}, layanan internet Anda sudah aktif kembali. Terima kasih.")
                
                # Dynamic Price for Template
                price_val = 0
                packet_name = client.get('packet_name', '')
                billing_profiles = settings.get('billing_profiles', {})
                for prof_name, prof_price in billing_profiles.items():
                    if prof_name.strip().lower() == packet_name.strip().lower():
                        price_val = prof_price; break
                
                # Format Expired for Template
                f_exp = client.get('paid_until', '-')
                if f_exp and '-' in f_exp:
                    try:
                        dt_p = datetime.strptime(f_exp, '%Y-%m-%d')
                        m_p = get_month_name(dt_p.month, settings.get('language', 'id'))
                        f_exp = f"{dt_p.day} {m_p} {dt_p.year}"
                    except: pass

                wa_msg = msg_tpl.replace("{name}", client.get('name', 'Pelanggan'))\
                               .replace("{id}", str(client.get('id', '')).replace('client_', ''))\
                               .replace("{expired}", f_exp)\
                               .replace("{price}", str(price_val))\
                               .replace("{amount}", str(price_val))

                try: 
                    # Trace Log (Debug)
                    _append_wa_log(f"[TRACE] Manual Reactivate WA: {client.get('name')} to {wa_num}")
                    
                    # Use mode="batch" and list format as wa-bridge requires
                    temp_wa = f"wa_react_{int(time.time())}_{random.randint(100,999)}.json"
                    temp_path = os.path.join(TEMP_FOLDER, temp_wa)
                    with open(temp_path, "w") as f:
                        json.dump([{"to": wa_num, "msg": wa_msg}], f)
                    spawn_wa_worker(mode="batch", task_file=temp_path)
                except Exception as e: 
                    print(f"Failed to queue WA reactivate: {e}")
                
            return {"status": "ok", "msg": msg}
        else:
            add_log(client['name'], 'error', f'Gagal Reaktivasi: {msg}')
            return {"status": "error", "msg": msg}

    except Exception as e:
        err_info = traceback.format_exc()
        with open("reaktivasi_error.txt", "w") as f:
            f.write(err_info)
        return {"status": "error", "msg": str(e)}

@app.route('/api/billing/client/<client_id>/activate', methods=['POST'])
def billing_activate_client(client_id):
    """Manual activation - restore from ISOLIR state"""
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json or {}
    selected_profile = data.get('profile')
    
    res = reaktivasi_client_core(client_id, selected_profile)
    return jsonify(res)


@app.route('/api/billing/client/<client_id>/isolir', methods=['POST'])
def isolir_client_core(client_id, processed_by="System"):
    """
    Core logic for isolating a client (manual or auto).
    Unifies Web UI and Telegram Bot logic to ensure reliability.
    """
    try:
        db_data = load_db()
        client = next((c for c in db_data.get('clients', []) if str(c.get('id','')).strip().upper() == str(client_id).strip().upper()), None)
        
        if not client:
            return {"status": "error", "msg": "Client not found"}
        
        pppoe_user = client.get('credentials', {}).get('pppoe_user')
        ip_addr = client.get('ip')
        
        # SMART ROUTER DETECTION: Prioritize managed_by, fallback to find_router helper
        router_id = client.get('managed_by')
        if not router_id:
            router_id = find_router_for_client(client['id'], db_data)
        
        success = False
        msg = ""
        
        # Setup Isolation Profile
        settings = load_settings()
        isolir_profile = settings.get('billing', {}).get('isolir_profile', 'ISOLIR')
        ensure_isolir_profile(router_id)

        if pppoe_user:
            pppoe_user = pppoe_user.strip()
            # PPPoE Mode: Check if Radius or Local
            is_radius = client.get('mode') == 'pppoe_radius'
            
            if is_radius:
                # PPPoE RADIUS Mode: Use Address List method (IP-based)
                target_ip = None
                try:
                    conn = get_router_connection(router_id)
                    if conn:
                        api = conn.get_api()
                        ppp_act = api.get_resource('/ppp/active').get(name=pppoe_user)
                        if ppp_act: target_ip = ppp_act[0].get('address')
                        conn.disconnect()
                except: pass
                
                if not target_ip: target_ip = client.get('ip')

                if target_ip and target_ip != '-':
                    res = add_to_address_list(target_ip, isolir_profile, router_id, comment=f"{isolir_profile}_{client['name']}")
                    if res.get('status') == 'ok':
                        time.sleep(0.5)
                        kick_pppoe_user(pppoe_user, router_id)
                        success = True
                        msg = "PPPoE RADIUS isolated via Address-List"
                    else:
                        msg = res.get('msg', 'Failed to add RADIUS IP to address-list')
                else:
                    msg = "No active session or valid IP found for RADIUS client"
            else:
                # Local PPPoE: Change profile
                current_profile_real = get_pppoe_current_profile(pppoe_user, router_id) or client.get('packet_name', 'default')
                res = change_pppoe_profile(pppoe_user, isolir_profile, router_id)
                if res.get('status') == 'ok':
                    time.sleep(0.5)
                    kick_pppoe_user(pppoe_user, router_id)
                    if 'billing' not in client or not isinstance(client['billing'], dict): 
                        client['billing'] = {}
                    
                    # Save original profile only if not already isolir
                    if isolir_profile.upper() not in current_profile_real.upper():
                        client['billing']['original_profile'] = current_profile_real
                    
                    success = True
                    msg = f"PPPoE Client isolated"
                else:
                    msg = res.get('msg', 'Failed to change segment profile')
        elif ip_addr and ip_addr != '-':
            # Static IP Mode: Add to ISOLIR list
            cur_prof = client.get('packet_name', 'default')
            if isolir_profile.upper() not in cur_prof.upper():
                if 'billing' not in client or not isinstance(client['billing'], dict): 
                    client['billing'] = {}
                client['billing']['original_profile'] = cur_prof
                
            res = add_to_address_list(ip_addr, isolir_profile, router_id, comment=f"{isolir_profile}_{client['name']}")
            if res.get('status') == 'ok':
                success = True
                msg = "Static IP Client isolated via Address-List"
            else:
                msg = res.get('msg', 'Failed to add to address-list')
        else:
            return {"status": "error", "msg": "Client has no PPPoE user or valid IP"}

        if success:
            if 'billing' not in client or not isinstance(client['billing'], dict): 
                client['billing'] = {}
            client['billing']['payment_status'] = 'overdue'
            client['billing']['isolir_date'] = get_local_now().strftime('%Y-%m-%d')
            client['status'] = 'isolir'
            client['packet_name'] = isolir_profile
            
            # Atomic Bulk Update
            apply_bulk_updates([{
                'id': client['id'],
                'status': 'isolir',
                'packet_name': isolir_profile,
                'billing': client['billing']
            }])
            
            add_log(client['name'], 'isolir', f'Manual isolir: {msg} (by {processed_by})')

            # Telegram Notification (Alert)
            d_alert = {
                'name': client.get('name', '-'),
                'ip': client.get('ip', '-'),
                'packet': client.get('packet_name') or '-'
            }
            dispatch_telegram_event('isolir', d_alert)

            # --- AUTO WA NOTIFICATION (Parity) ---
            try:
                wa_isolir_enabled = settings.get('wa_isolir_enabled', True)
                p_check = client.get('phone') or client.get('whatsapp_number') or client.get('wa_number') or client['billing'].get('wa_number')
                if p_check and wa_isolir_enabled:
                    p_clean = re.sub(r'\D', '', str(p_check))
                    if p_clean.startswith('0'): p_clean = "62" + p_clean[1:]
                    
                    msg_tpl = settings.get('wa_template_isolir') or "Yth. {name}, layanan internet Anda diisolir sementara."
                    
                    # Pricing and Status
                    price_val = 0
                    orig_p = client['billing'].get('original_profile', '')
                    for pn, pv in settings.get('billing_profiles', {}).items():
                        if pn.strip().lower() == orig_p.strip().lower():
                            price_val = pv; break
                    
                    m_arr_tot = 0
                    for ma in settings.get('manual_arrears', []):
                        if ma.get('client_name') == client.get('name'):
                            m_arr_tot += int(ma.get('amount', 0))
                    
                    c_bill_str = "{:,}".format(int(price_val + m_arr_tot)).replace(",", ".")
                    
                    final_wa = msg_tpl.replace("{name}", client.get('name', 'Pelanggan'))\
                                      .replace("{id}", str(client['id']).replace('client_', ''))\
                                      .replace("{amount}", c_bill_str)\
                                      .replace("{bill}", c_bill_str)\
                                      .replace("{month}", get_month_name(get_local_now().month, settings.get('language', 'id')))\
                                      .replace("{time}", get_local_now().strftime('%H:%M'))\
                                      .replace("{status}", "BELUM TERBAYAR")\
                                      .replace("{expired}", client.get('paid_until', 'N/A'))\
                                      .replace("{profile}", orig_p or isolir_profile)
                    
                    spawn_wa_worker(mode="test", target=p_clean, message=final_wa)
            except Exception as wa_e:
                print(f"[CORE_ISOLIR_WA_ERR] {wa_e}")

            return {"status": "ok", "msg": msg}
        else:
            return {"status": "error", "msg": msg}
    except Exception as e:
        import traceback
        print(f"[CORE_ISOLIR_ERR] {traceback.format_exc()}")
        return {"status": "error", "msg": str(e)}

@app.route('/api/billing/client/<client_id>/isolir', methods=['POST'])
def billing_isolir_client(client_id):
    """Refactored Manual isolir route"""
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    res = isolir_client_core(client_id, processed_by="Web UI")
    if res['status'] == 'ok':
        return jsonify(res)
    else:
        return jsonify(res), 500

@app.route('/api/finance/summary')
def fin_summary():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    # Ensuring carryover is generated if we entered a new month
    process_monthly_carryover()

    data = load_finance()
    txs = data.get('transactions', [])
    
    # Filter for current month using local time
    now = get_local_now()
    cur_month = now.strftime('%Y-%m')
    
    income_pure = 0    # Only real income this month (excluding carryover)
    expense_pure = 0   # Real expenses this month
    monthly_carry = 0  # Net balance carried from ALL previous months
    paid_clients = []
    
    for t in txs:
        t_date = t.get('date', '')
        
        # We only need to check current month transactions because 
        # the 'balance_carryover' for this month already summarizes everything before it.
        if t_date.startswith(cur_month):
            amt = int(t.get('amount', 0))
            t_type = t.get('type')
            t_cat = t.get('category', '')

            if t_cat == 'balance_carryover':
                # This entry IS the total balance from January to Last Month
                if t_type == 'income': monthly_carry += amt
                else: monthly_carry -= amt
            else:
                # Real transaction this month
                if t_type == 'income':
                    income_pure += amt
                    if t_cat == 'wifi_payment' and t.get('client_id'):
                        paid_clients.append(t.get('client_id'))
                elif t_type == 'expense':
                    expense_pure += amt
            
    # Total Global Cash = (Balance from the past) + (Activity this month)
    total_balance = monthly_carry + income_pure - expense_pure
    
    return jsonify({
        "income": income_pure,              # Card 1: Pure income this month
        "expense": expense_pure,            # Card 2: Total expense this month
        "balance": income_pure - expense_pure,   # Card 3: Monthly surplus/deficit
        "total_balance": total_balance,      # Card 4: Global Cash in hand (Physical Wallet)
        "month": cur_month,
        "paid_clients": list(set(paid_clients))
    })

@app.route('/api/finance/data')
def api_finance_data():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    return jsonify(load_finance())

def process_monthly_carryover():
    """
    Checks if current month has a carryover transaction.
    If not, calculate ALL previous balance and insert it.
    """
    with db_lock: # Reuse DB lock to prevent race conditions roughly
        data = load_finance()
        txs = data.get('transactions', [])
        
        now = get_local_now()
        cur_month_str = now.strftime('%Y-%m') # e.g. 2026-02
        
        # 1. Check if carryover exists for this month
        exists = any(t.get('category') == 'balance_carryover' and t.get('date', '').startswith(cur_month_str) for t in txs)
        if exists: return
        
        # 2. If valid previous data exists (don't run on brand new empty system unless needed?)
        # Actually run it.
        
        # Perhitungan Saldo Awal menggunakan logika "Waterflow" (Akuntansi Bertahap)
        # Mencari saldo pindahan terakhir sebelum bulan ini
        latest_carry = None
        for t in txs:
            if t.get('category') == 'balance_carryover':
                t_date = t.get('date', '')
                if t_date < cur_month_str + "-01":
                    if not latest_carry or t_date > latest_carry['date'] or (t_date == latest_carry['date'] and t['id'] > latest_carry['id']):
                        latest_carry = t

        prev_balance = 0
        has_history = False
        cutoff_date = ""
        cutoff_id = ""

        if latest_carry:
            has_history = True
            cutoff_date = latest_carry['date']
            cutoff_id = latest_carry['id']
            amt = int(latest_carry.get('amount', 0))
            if latest_carry.get('type') == 'income': prev_balance = amt
            else: prev_balance = -amt

        for t in txs:
            if t.get('category') == 'balance_carryover': continue
            
            t_date = t.get('date', '')
            if t_date < cur_month_str + "-01":
                # Tambah transaksi yang terjadi setelah atau pada hari yang sama dengan saldo pindahan terakhir
                if not latest_carry or t_date > cutoff_date or (t_date == cutoff_date and t['id'] > cutoff_id):
                    has_history = True
                    amt = int(t.get('amount', 0))
                    if t.get('type') == 'income': prev_balance += amt
                    else: prev_balance -= amt
            
        # 3. Create Carryover Transaction
        # Get Previous Month Name (Indonesian)
        # Python date math
        first_of_this_month = now.replace(day=1)
        last_month_obj = first_of_this_month - timedelta(days=1)
        months_id = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
        prev_month_name = months_id[last_month_obj.month - 1]
        
        note = f"Sisa Saldo Bulan {prev_month_name} {last_month_obj.year}"
        
        # Only carry over if positive? Or negative too? 
        # User said "keuangan masuk" (Income). If negative, it's technically expense/debt?
        # Let's map it: Positive -> Income. Negative -> Expense.
        
        tx_type = 'income'
        final_amt = prev_balance
        if prev_balance < 0:
            tx_type = 'expense'
            final_amt = abs(prev_balance)
            note = f"Minus Saldo Bulan {prev_month_name} {last_month_obj.year}"
            
        if final_amt == 0: return # Nothing to carry over
            
        new_tx = {
            "id": str(int(time.time() * 1000)),
            "date": cur_month_str + "-01", # 1st of current month
            "type": tx_type,
            "category": "balance_carryover",
            "amount": final_amt,
            "note": note,
            "user": "system"
        }
        
        data.setdefault('transactions', []).append(new_tx)
        
        # 4. CLEANUP (Keep Max 5 Months History strictly)
        # Current Month is kept. We keep 5 months back.
        # Logic: Go back 6 months to find cutoff.
        d = first_of_this_month
        for _ in range(6):
            d = d - timedelta(days=1)
            d = d.replace(day=1)
        cutoff_str = d.strftime('%Y-%m') # '%Y-%m-01' is safer comparison if string
        
        # Filter: Keep if date >= cutoff_str
        # Note: '2025-08' > '2025-07' works for strings
        original_len = len(data['transactions'])
        data['transactions'] = [t for t in data['transactions'] if t.get('date', '') >= cutoff_str]
        removed_count = original_len - len(data['transactions'])
        
        save_finance(data)
        print(f"[FINANCE] Generated Auto-Carryover for {cur_month_str}. Cleaned {removed_count} old transactions (Cutoff: {cutoff_str})")

def prettify_money(v): return "Rp " + "{:,}".format(v)

@app.route('/api/finance/history')
def fin_history():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    data = load_finance()
    # Return reverse chronological order (newest first)
    hist = sorted(data.get('transactions', []), key=lambda x: x.get('id', ''), reverse=True)
    return jsonify(hist)

@app.route('/api/finance/last_transaction/<client_id>')
def fin_last_tx(client_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    data = load_finance()
    # Find newest transaction for this client
    txs = [t for t in data.get('transactions', []) if str(t.get('client_id')) == str(client_id)]
    if not txs:
        return jsonify({"status": "empty", "msg": "Belum ada riwayat pembayaran."})
    
    # Sort descending by ID (timestamp)
    last = sorted(txs, key=lambda x: x.get('id', ''), reverse=True)[0]
    return jsonify({"status": "found", "data": last})

def bot_execute_payment(client_id, amount, processed_by="Bot Assistant"):
    """ SAFE: Dedicated helper for bot payment. Supports multi-month duration. """
    try:
        data = load_finance()
        db_data = load_db()
        client = next((c for c in db_data.get('clients', []) if str(c.get('id','')).strip().upper() == str(client_id).strip().upper()), None)
        if not client: return {"status": "error", "msg": "Client not found"}

        now = get_local_now()
        date_str = now.strftime('%Y-%m-%d')
        
        # --- SMART DURATION LOGIC ---
        billing_config = load_billing_config()
        
        # Read original profile to calculate true price
        bill_data = client.get('billing', {})
        p_name = getattr(bill_data, 'get', lambda k, d=None: d)('original_profile') if isinstance(bill_data, dict) else None
        if not p_name or p_name == "":
            p_name = client.get('packet_name', 'default')

        price_per_month = 0
        b_profs = billing_config.get('billing_profiles', {})
        for b_n, b_p in b_profs.items():
            if b_n.strip().lower() == p_name.strip().lower():
                price_per_month = int(b_p); break
        
        # [BUG FIX] Capture INITIAL manual arrears BEFORE smart billing adjustment
        # to avoid double-counting in the WhatsApp status label (e.g. 20k becomes 40k)
        initial_m_arr_total = 0
        m_arr_list_init = billing_config.get('manual_arrears', [])
        c_name_u = str(client.get('name', '')).strip().upper()
        for ma in m_arr_list_init:
            if str(ma.get('client_name', '')).strip().upper() == c_name_u:
                initial_m_arr_total += int(ma.get('amount', 0))
        
        duration = 1
        if price_per_month > 0:
            # Jika bayar kelipatan paket, otomatis tambah bulan
            potential_dur = amount // price_per_month
            if potential_dur >= 1:
                duration = potential_dur

        b_mode = billing_config.get('billing_mode', 'monthly')
        
        try: current_expiry = datetime.strptime(client.get('paid_until', ''), '%Y-%m-%d')
        except: current_expiry = now

        if b_mode == 'cyclic':
            new_expiry = max(now, current_expiry) + timedelta(days=30 * duration)
        else:
            b_day = billing_config.get('default_billing_day', 20)
            # Use client-specific b_day if mode is fixed
            if b_mode == 'fixed':
                b_day = client.get('billing', {}).get('billing_day') or b_day

            if current_expiry and current_expiry > datetime(2000, 1, 1):
                start_from_year, start_from_month = current_expiry.year, current_expiry.month
                if (now - current_expiry).days > 180: start_from_year, start_from_month = now.year, now.month
            else:
                if now.day > b_day:
                    idx = now.year * 12 + now.month
                    start_from_year, start_from_month = (idx // 12), (idx % 12) + 1
                    if start_from_month > 12: start_from_month = 1; start_from_year += 1
                else: start_from_year, start_from_month = now.year, now.month
            
            target_month = start_from_month + duration
            target_year = start_from_year + (target_month - 1) // 12
            target_month = (target_month - 1) % 12 + 1
            last_day = calendar.monthrange(target_year, target_month)[1]
            new_expiry = datetime(target_year, target_month, min(b_day, last_day))

        client['paid_until'] = new_expiry.strftime('%Y-%m-%d')
        
        # Record Transaction
        new_tx = {
            "id": str(int(time.time() * 1000)),
            "date": date_str, "type": "income", "category": "wifi_payment",
            "amount": amount, "note": f"Bayar via Telegram Bot (Nama: {client.get('name', 'Unknown')})",
            "client_id": client_id, "user": processed_by
        }
        data.setdefault('transactions', []).append(new_tx)
        save_finance(data)

        # --- SMART BILLING LOGIC: Handle Overpayments / Underpayments ---
        try:
            settings_tmp = load_settings()
            billing_config = load_billing_config()
            profiles = billing_config.get('billing_profiles', {})
            pkt_name = (client.get('packet_name') or "").strip().lower()
            
            p_price = 0
            for pn, pv in profiles.items():
                if pn.strip().lower() == pkt_name:
                    p_price = int(pv); break
            
            if p_price > 0:
                # Calculate expected cost for the duration bought
                expected_cost = p_price * duration
                diff = expected_cost - amount # Positive = debt, Negative = credit
                
                if diff != 0:
                    c_name_norm = str(client.get('name', '')).strip().upper()
                    m_arrs = settings_tmp.get('manual_arrears', [])
                    
                    target_ma = next((ma for ma in m_arrs if str(ma.get('client_name', '')).strip().upper() == c_name_norm), None)
                    if target_ma:
                        target_ma['amount'] = int(target_ma.get('amount', 0)) + diff
                        target_ma['desc'] = f"Updated via Bot ({date_str})"
                    else:
                        settings_tmp.setdefault('manual_arrears', []).append({
                            "id": str(int(time.time() * 1000)),
                            "client_name": client.get('name'),
                            "amount": diff,
                            "desc": f"Bot Payment Adjustment ({date_str})"
                        })
                    
                    # Cleanup 0 entries and save
                    settings_tmp['manual_arrears'] = [ma for ma in settings_tmp['manual_arrears'] if int(ma.get('amount', 0)) != 0]
                    save_settings(settings_tmp)
                    
                    # Add info to transaction note
                    if diff < 0: new_tx['note'] += f" (Kelebihan: {abs(diff)})"
                    else: new_tx['note'] += f" (Kekurangan: {diff})"
                    save_finance(data)
        except Exception as e:
            print(f"[BOT SMART BILLING ERROR] {e}")

        # Reactivation
        was_reactivated = False
        if client.get('status') == 'isolir':
            res_core = reaktivasi_client_core(client_id, send_notif=False)
            if res_core.get('status') == 'ok': was_reactivated = True
        
        # Bulk Save
        apply_bulk_updates([{
            'id': client['id'], 'paid_until': client['paid_until'],
            'billing': client['billing'], 'status': client.get('status'),
            'packet_name': client.get('packet_name')
        }])

        # --- 4. SEND WHATSAPP NOTIFICATION (Parity with Web UI) ---
        try:
            settings_wa = load_settings()
            wa_pay_enabled = settings_wa.get('wa_payment_notif_enabled')
            wa_react_enabled = settings_wa.get('wa_reactivate_enabled', True)

            if (was_reactivated and wa_react_enabled) or (not was_reactivated and wa_pay_enabled):
                p_check = client.get('phone') or client.get('whatsapp_number') or client.get('wa_number') or client['billing'].get('wa_number')
                if p_check:
                    p_clean = re.sub(r'\D', '', str(p_check))
                    if p_clean.startswith('0'): p_clean = "62" + p_clean[1:]
                    
                    if p_clean:
                        f_amt = "{:,}".format(int(amount)).replace(",", ".")
                        f_exp_raw = client.get('paid_until', '')
                        f_exp = f_exp_raw
                        if f_exp_raw and '-' in f_exp_raw:
                            try:
                                dt_p = datetime.strptime(f_exp_raw, '%Y-%m-%d')
                                m_p = get_month_name(dt_p.month, settings_wa.get('language', 'id'))
                                f_exp = f"{dt_p.day} {m_p} {dt_p.year}"
                            except: pass

                        if was_reactivated:
                            tpl = settings_wa.get('wa_template_reactivate') or "Halo {name}, pembayaran Rp {amount} telah diterima dan layanan Anda telah diaktifkan kembali hingga {expired}."
                        else:
                            tpl = settings_wa.get('wa_template_payment') or "Terima kasih, pembayaran wifi a.n {name} sebesar Rp {amount} pada {date} telah diterima."
                        
                        # Status Logic
                        p_name_notif = client.get('packet_name', 'default')
                        price_notif = price_per_month
                        
                        # Use captured INITIAL arrears to avoid double-counting
                        target_total = (price_notif * duration) + initial_m_arr_total
                        diff = int(amount) - target_total
                        
                        if target_total == 0: txt_status = "LUNAS" # Smart Logic: No bill yet, paid upfront
                        elif diff == 0: txt_status = "LUNAS"
                        elif diff < 0: txt_status = f"BELUM LUNAS (Kurang Rp " + "{:,}".format(abs(diff)).replace(",", ".") + ")"
                        else: txt_status = f"LUNAS (Lebih Rp " + "{:,}".format(diff).replace(",", ".") + ")"
                        
                        final_msg = tpl.replace('{name}', client.get('name', 'Pelanggan'))\
                                       .replace('{id}', str(client_id).replace('client_', ''))\
                                       .replace('{amount}', f_amt)\
                                       .replace('{price}', "{:,}".format(price_notif).replace(",", "."))\
                                       .replace('{bill}', "{:,}".format(target_total).replace(",", "."))\
                                       .replace('{month}', get_month_name(now.month, settings_wa.get('language', 'id')))\
                                       .replace('{date}', date_str)\
                                       .replace('{expired}', f_exp)\
                                       .replace('{profile}', p_name_notif)\
                                       .replace('{time}', datetime.now().strftime('%H:%M'))\
                                       .replace('{status}', txt_status)
                        
                        spawn_wa_worker(mode="test", target=p_clean, message=final_msg)
        except Exception as wa_e:
            print(f"[BOT_PAYMENT_WA_ERR] {wa_e}")
            
        return {"status": "ok", "new_expiry": client['paid_until'], "was_reactivated": was_reactivated, "duration": duration}
    except Exception as e:
        return {"status": "error", "msg": str(e)}

def bot_cancel_last_payment(client_id, sender_name):
    """ Smart Reversal: Find last payment, delete it, and revert expiry. """
    try:
        f_data = load_finance()
        db_data = load_db()
        client = next((c for c in db_data.get('clients', []) if str(c.get('id','')).strip().upper() == str(client_id).strip().upper()), None)
        if not client: return {"status": "error", "msg": "Client not found"}

        # 1. Find last payment transaction
        txs = [t for t in f_data.get('transactions', []) if str(t.get('client_id')) == str(client_id) and t.get('category') == 'wifi_payment']
        if not txs: return {"status": "error", "msg": "No payment history found"}
        
        last_tx = sorted(txs, key=lambda x: x.get('id', ''), reverse=True)[0]
        amount = int(last_tx.get('amount', 0))
        
        # 2. Determine duration to revert
        billing_config = load_billing_config()
        p_name = client.get('packet_name', 'default')
        price_per_month = 0
        b_profs = billing_config.get('billing_profiles', {})
        for b_n, b_p in b_profs.items():
            if b_n.strip().lower() == p_name.strip().lower():
                price_per_month = int(b_p); break
        
        revert_months = 1
        if price_per_month > 0:
            revert_months = amount // price_per_month
            if revert_months < 1: revert_months = 1

        # 3. Revert Expiry
        try:
            curr_exp = datetime.strptime(client.get('paid_until', ''), '%Y-%m-%d')
            # Mundur per bulan (target bulan sebelumnya)
            m = curr_exp.month - revert_months
            y = curr_exp.year
            while m <= 0: m += 12; y -= 1
            
            b_day = billing_config.get('default_billing_day', 20)
            last_day = calendar.monthrange(y, m)[1]
            new_exp = datetime(y, m, min(b_day, last_day))
            client['paid_until'] = new_exp.strftime('%Y-%m-%d')
        except: pass

        # 4. Remove Transaction
        f_data['transactions'] = [t for t in f_data['transactions'] if t.get('id') != last_tx.get('id')]
        save_finance(f_data)
        
        # 5. Log & Save DB
        add_log("BOT", "system", f"BATAL BAYAR: {client['name']} (Rp {amount}) oleh {sender_name}")
        
        # Reset payment status if needed
        if 'billing' in client:
            now_str = get_local_now().strftime('%Y-%m-%d')
            if client['paid_until'] < now_str: client['billing']['payment_status'] = 'overdue'
            else: client['billing']['payment_status'] = 'unpaid'

        apply_bulk_updates([{
            'id': client['id'], 'paid_until': client['paid_until'],
            'billing': client['billing'], 'status': client.get('status')
        }])
        
        return {"status": "ok", "new_expiry": client['paid_until'], "amount": amount}
    except Exception as e:
        return {"status": "error", "msg": str(e)}

def bot_add_manual_tx(tx_type, category_name, amount, note, sender_name):
    """ Record manual income/expense with category mapping and note enrichment. """
    try:
        data = load_finance()
        now = get_local_now()
        
        # Mapping Category
        cat_map = {
            "investasi": "cat_invest",
            "hibah": "cat_grant",
            "lain-lain": "cat_others",
            "invest": "cat_invest",
            "grant": "cat_grant",
            "others": "cat_others",
            # Expense Categories
            "alat": "cat_tools",
            "rawat": "cat_maint",
            "ops": "cat_ops",
            "gaji": "cat_salary",
            "fee": "cat_salary"
        }
        sys_cat = cat_map.get(category_name.lower(), "cat_others")
        
        # Enrich note with sender info so it's visible in web dashboard Description column
        enriched_note = f"{note} (Telegram Assistant Oleh: {sender_name})"
        
        new_tx = {
            "id": str(int(time.time() * 1000)),
            "date": now.strftime('%Y-%m-%d'),
            "type": tx_type,
            "category": sys_cat,
            "amount": amount,
            "note": enriched_note,
            "user": sender_name
        }
        data.setdefault('transactions', []).append(new_tx)
        save_finance(data)
        
        label = "PEMASUKAN" if tx_type == "income" else "PENGELUARAN"
        add_log("BOT", "finance", f"{label} MANUAL: {note} (Rp {amount}) oleh {sender_name}")
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "msg": str(e)}

@app.route('/api/finance/add', methods=['POST'])
@app.route('/api/finance/transaction', methods=['POST'])
def fin_add_tx():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    payload = request.json or {}
    type_ = payload.get('type', 'income') # Default income for legacy
    cat = payload.get('category', 'wifi_payment') # Default wifi for payment modal
    amt = int(payload.get('amount', 0))
    note = payload.get('note', '')
    date_ = payload.get('date') or get_local_now().strftime('%Y-%m-%d')
    client_id = payload.get('client_id')
    duration = int(payload.get('duration', 1)) 
    auto_reactivate = payload.get('auto_reactivate', True)
    
    if amt <= 0:
        return jsonify({"error": "Invalid amount"}), 400
        
    data = load_finance()
    tx_id = payload.get('id')
    is_update = False
    
    # Transactional Atomicity: Validate client BEFORE saving record
    client = None
    if (cat == 'wifi_payment' or cat == 'Pembayaran WiFi') and client_id:
        topology = db.load_full_topology()
        # Primary lookup: String-safe ID matching (Parity with Telegram Assistant)
        client = next((c for c in topology['clients'] if str(c.get('id','')).strip().upper() == str(client_id).strip().upper()), None)
        
        # Secondary lookup: Name/Username Fallback (Essential for imported clients)
        if not client:
            client = next((c for c in topology['clients'] if str(c.get('name', '')).strip().lower() == str(client_id).strip().lower()), None)
        
        if not client and not is_update:
            return jsonify({"error": f"Client not found: {client_id}"}), 404

    if tx_id:
        # Check if transaction exists
        for t in data.get('transactions', []):
            if str(t.get('id')) == str(tx_id):
                t.update({
                    "date": date_,
                    "type": type_,
                    "category": cat,
                    "amount": amt,
                    "note": note,
                    "client_id": client_id,
                    "user": check_auth(request)
                })
                new_tx = t
                is_update = True
                break
    
    if not is_update:
        new_tx = {
            "id": str(tx_id) if tx_id else str(int(time.time() * 1000)),
            "date": date_,
            "type": type_,
            "category": cat,
            "amount": amt,
            "note": note,
            "client_id": client_id,
            "user": check_auth(request)
        }
        data.setdefault('transactions', []).append(new_tx)
    
    save_finance(data)
    
    # AUTO-REACTIVATION & PAID_UNTIL UPDATE (ONLY FOR NEW TRANSACTIONS)
    if is_update:
        return jsonify({"status": "ok", "id": new_tx['id']})

    try:
        if (cat == 'wifi_payment' or cat == 'Pembayaran WiFi') and client:
            # 1. Update paid_until (Top Level)
            now = get_local_now()
            try:
                current_expiry = datetime.strptime(client.get('paid_until', ''), '%Y-%m-%d')
                has_existing_expiry = True
            except:
                current_expiry = now
                has_existing_expiry = False  # New client, no prior paid_until

            # --- PREMIUM PRICE VERIFICATION & SMART DURATION ---
            billing_config = load_billing_config()
            profiles = billing_config.get('billing_profiles', {})
            pkt_name = client.get('packet_name')
            
            price_per_month = 0
            if pkt_name:
                # Flexible lookup
                for b_n, b_p in profiles.items():
                    if b_n.strip().lower() == pkt_name.strip().lower():
                        price_per_month = int(b_p); break

            # [BUG FIX] Capture INITIAL manual arrears BEFORE smart billing adjustment
            # to avoid double-counting in the WhatsApp status label (e.g. 20k becomes 40k)
            initial_m_arr_total = 0
            m_arr_list_init = billing_config.get('manual_arrears', [])
            c_name_u = str(client.get('name', '')).strip().upper()
            for ma in m_arr_list_init:
                if str(ma.get('client_name', '')).strip().upper() == c_name_u:
                    initial_m_arr_total += int(ma.get('amount', 0))

            # SMART DURATION: If user pays multiples but duration is 1, auto-calculate
            if price_per_month > 0:
                potential_dur = amt // price_per_month
                if potential_dur > duration:
                    duration = potential_dur # Upgrade to higher duration if amount covers it
                    
            expected = price_per_month * duration
            if amt != expected:
                diff = amt - expected
                status_label = "Lebih" if diff > 0 else "Kurang"
                sign = "+" if diff > 0 else "-"
                # Format as IDR (100.000)
                fmt_diff = "{:,.0f}".format(abs(diff)).replace(',', '.')
                mismatch_tag = f" (!! MISMATCH: {status_label} {sign}Rp {fmt_diff})"
                new_tx['note'] = new_tx.get('note', '') + mismatch_tag
                # Note: This mismatch tag is for internal Finance logs and will be filtered in printReceipt.
            
            # --- AUTO ARREARS / SMART BILLING LOGIC Pre-check ---
            # 1. Sequential Manual Arrears Payoff
            settings_tmp = load_settings()
            m_arrs_tmp = settings_tmp.get('manual_arrears', [])
            
            effective_duration = duration
            remaining_amt_tmp = amt
            
            c_name_norm = str(client.get('name', '')).strip().upper()
            
            for ma in m_arrs_tmp:
                if str(ma.get('client_name', '')).strip().upper() == c_name_norm:
                    old_debt = int(ma.get('amount', 0))
                    if old_debt > 0 and remaining_amt_tmp > 0:
                        payoff = min(remaining_amt_tmp, old_debt)
                        ma['amount'] = old_debt - payoff
                        remaining_amt_tmp -= payoff
                        ma['desc'] = f"Hutang dibayar ({date_})"
                        
                        if remaining_amt_tmp <= 0:
                            effective_duration = 0
                            break
            
            # Cleanup: remove 0 amount manual arrears
            settings_tmp['manual_arrears'] = [ma for ma in m_arrs_tmp if int(ma.get('amount', 0)) != 0]

            # --- BILLING METHOD CALCULATION (3 MODES) ---
            billing_config = load_billing_config()
            b_mode = billing_config.get('billing_mode', 'monthly') # monthly=Global, cyclic=30-Day, fixed=Client-Specific
            
            # Resolve Billing Day (Priority depends on Mode)
            billing = client.get('billing') or {}
            payload_b_day = payload.get('billing_day')
            global_b_day = billing_config.get('default_billing_day', 20)

            if b_mode == 'monthly':
                # MONTHLY: Ignore stale client-specific day, use Global or Payload Override
                b_day = int(payload_b_day) if payload_b_day else global_b_day
            else:
                # FIXED: Prioritize Payload > Client Specific > Global
                if payload_b_day:
                    b_day = int(payload_b_day)
                else:
                    b_day = billing.get('billing_day') or global_b_day

            # Update billing day in object early to ensure consistency
            if b_mode == 'fixed' or payload_b_day:
                if 'billing' not in client or not isinstance(client['billing'], dict): client['billing'] = {}
                client['billing']['billing_day'] = b_day
                client['billing']['enabled'] = True
                billing = client['billing'] # re-sync

            if effective_duration <= 0:
                new_expiry = current_expiry
            elif b_mode == 'cyclic':
                # MODE 30 HARI (Cyclic): Accurate Addition of 30 days
                # Base from today or expiry, whichever is later
                base_date = max(now, current_expiry)
                new_expiry = base_date + timedelta(days=30 * effective_duration)
            else:
                # MODE GLOBAL (Monthly) or FIXED (Anniversary)
                # Ensure start_from variables are initialized
                # Use has_existing_expiry flag instead of date comparison
                # to correctly detect new clients (paid_until was None/empty)
                if has_existing_expiry:
                    # Continue from previous expiry to cover gaps
                    # If gap is too large (> 6 months), reset to 'now' to avoid huge backlog payments
                    if (now - current_expiry).days > 180:
                        start_from_year, start_from_month = now.year, now.month
                    else:
                        start_from_year, start_from_month = current_expiry.year, current_expiry.month
                else:
                    # New client or reset: check if we start this month or next
                    if now.day > b_day:
                        # After billing day: start from next month
                        start_from_year = now.year + (now.month // 12)
                        start_from_month = (now.month % 12) + 1
                    else:
                        # Before/on billing day: start from this month
                        start_from_year, start_from_month = now.year, now.month

                # Calculate final target date
                target_total_months = start_from_month + effective_duration
                target_year = start_from_year + (target_total_months - 1) // 12
                target_month = (target_total_months - 1) % 12 + 1
                
                import calendar
                last_day = calendar.monthrange(target_year, target_month)[1]
                actual_day = min(b_day, last_day)
                new_expiry = datetime(target_year, target_month, actual_day)

            # Final Assignment & Audit Log
            client['paid_until'] = new_expiry.strftime('%Y-%m-%d')
            
            # Update Finance Note for transparency
            if "Lunas s/d" not in new_tx['note']:
                new_tx['note'] += f" (Lunas s/d {client['paid_until']})"
                save_finance(data)

            # 2. Billing sub-object update (Ensure sync)
            if not client.get('billing'):
                client['billing'] = billing

            billing['last_payment_date'] = date_

            # Don't unconditionally reset status to 'paid' if they still have manual debt
            total_ma = 0
            for ma in settings_tmp.get('manual_arrears', []):
                if (ma.get('client_name') or "").strip().upper() == (client.get('name') or "").strip().upper():
                    total_ma += int(ma.get('amount') or 0)
            
            if total_ma <= 0:
                billing['payment_status'] = 'paid'
                billing['overdue_months'] = 0
            else:
                # If they still have debt, Keep status or set to 'partial'? 
                # For now just don't reset it to 'paid' if it was already debt/isolir
                if billing.get('payment_status') == 'paid':
                     # If it was paid but now they have manual debt (e.g. from adjustment), keep it paid but logically they have debt.
                     # But usually if they have debt, status should be 'debt'.
                     billing['payment_status'] = 'debt'
            
            # 3. Auto Activate if isolated
            was_reactivated = False
            if auto_reactivate and client.get('status') == 'isolir':
                saved_paid_until = client.get('paid_until')
                saved_billing = client.get('billing')
                
                res_core = reaktivasi_client_core(client_id, send_notif=False)
                if res_core.get('status') == 'ok':
                    was_reactivated = True
                    # Refresh client from DB to avoid overwriting state (e.g. Mikrotik info)
                    fresh_db = load_db(force_refresh=True)
                    client = next((c for c in fresh_db.get('clients', []) if str(c.get('id','')).strip().upper() == str(client_id).strip().upper()), client)
                    
                    # Restore calculated payment data that was lost during refresh
                    if saved_paid_until: client['paid_until'] = saved_paid_until
                    if saved_billing: client['billing'] = saved_billing
                
            # --- SMART BILLING LOGIC: Final Arrears Adjustment ---
            try:
                # price_per_month and duration were already calculated/verified at the start of fin_add_tx
                if price_per_month > 0:
                    # Calculate final difference after paying for the months (duration)
                    # Note: price_per_month * duration is the 'cost' of the expiry extension
                    # remaining_amt_tmp has already been used to payoff OLD debts (lines 6631-6638)
                    # So the surplus/deficit compared to 'expected' should be recorded.
                    
                    diff = (price_per_month * duration) - remaining_amt_tmp
                    
                    if diff != 0:
                        # Load FRESH settings to avoid Lost Update
                        ps = _load_settings_raw()
                        m_arrs = ps.get('manual_arrears', [])
                        
                        target_ma = next((ma for ma in m_arrs if str(ma.get('client_name', '')).strip().upper() == c_name_norm), None)
                        
                        if target_ma:
                            target_ma['amount'] = int(target_ma.get('amount', 0)) + diff
                            target_ma['desc'] = f"Updated via Smart Billing ({date_})"
                        else:
                            ps.setdefault('manual_arrears', []).append({
                                "id": str(int(time.time() * 1000)),
                                "client_name": client['name'],
                                "amount": diff,
                                "desc": f"Smart Billing Adjustment ({date_})"
                            })
                        
                        # Cleanup 0 entries
                        ps['manual_arrears'] = [ma for ma in ps['manual_arrears'] if int(ma.get('amount', 0)) != 0]
                        save_settings(ps)
                        
                        # Finance Note Update for transparency
                        if diff < 0: # Overpaid
                            new_tx['note'] += f" (Kelebihan: {abs(diff)})"
                        else: # Underpaid
                            new_tx['note'] += f" (Kekurangan: {diff})"
                        save_finance(data)
            except Exception as e:
                print(f"[SMART BILLING ERROR] {e}")

            # Save changes to SQLite Database (Targeted Atomic Update)
            apply_bulk_updates([{
                'id': client['id'],
                'paid_until': client.get('paid_until'),
                'billing': client.get('billing'),
                'status': client.get('status'),
                'packet_name': client.get('packet_name')
            }])

            # 4. SEND TELEGRAM NOTIFICATION (If reactivated)
            if was_reactivated:
                d_alert = {
                    'name': client.get('name', '-'),
                    'ip': client.get('ip', '-'),
                    'packet': client.get('packet_name') or '-'
                }
                dispatch_telegram_event('reactivate', d_alert)

            # 5. SEND WHATSAPP PAYMENT NOTIFICATION
            try:
                s_wa = load_settings()
                wa_pay_enabled = s_wa.get('wa_payment_notif_enabled')
                wa_react_enabled = s_wa.get('wa_reactivate_enabled', True)

                if (was_reactivated and wa_react_enabled) or (not was_reactivated and wa_pay_enabled):
                    # Check phone
                    p_check = client.get('phone') or client.get('whatsapp_number') or client.get('wa_number') or billing.get('wa_number')
                    if p_check:
                        # Basic cleanup
                        import re
                        p_clean = re.sub(r'\D', '', str(p_check))
                        if p_clean.startswith('0'): p_clean = "62" + p_clean[1:]
                        
                        if p_clean:
                            # Format price and expiry for template
                            f_amt = "{:,}".format(int(amt)).replace(",", ".")
                            f_exp = client.get('paid_until', '')

                            if was_reactivated:
                                tpl = s_wa.get('wa_template_reactivate')
                                if not tpl:
                                    tpl = "Halo {name}, pembayaran Rp {amount} telah diterima dan layanan Anda telah diaktifkan kembali hingga {expired}. Terima kasih."
                            else:
                                tpl = s_wa.get('wa_template_payment')
                                if not tpl:
                                    tpl = "Terima kasih, pembayaran wifi a.n {name} sebesar Rp {amount} pada {date} telah diterima."
                                
                            # Smart Expired Date for Payment Notif
                            f_exp = client.get('paid_until', '')
                            if f_exp and '-' in f_exp:
                                try:
                                    dt_p = datetime.strptime(f_exp, '%Y-%m-%d')
                                    m_p = get_month_name(dt_p.month, s_wa.get('language', 'id'))
                                    f_exp = f"{dt_p.day} {m_p} {dt_p.year}"
                                except: pass

                            f_id = str(client.get('id', '')).replace('client_', '')
                            
                            # SMART LOGIC: (Price x Duration) + Manual Arrears
                            p_name = client.get('packet_name', 'default')
                            price_per_month = 0
                            b_profs = s_wa.get('billing_profiles', {})
                            for b_n, b_p in b_profs.items():
                                if b_n.strip().lower() == p_name.strip().lower():
                                    price_per_month = int(b_p); break
                            
                            # Use captured INITIAL arrears to avoid double-counting
                            target_total = (price_per_month * duration) + initial_m_arr_total
                            diff = int(amt) - target_total
                            
                            if target_total == 0: txt_status = "LUNAS" # Smart Logic: No bill yet, paid upfront
                            elif diff == 0: txt_status = "LUNAS"
                            elif diff < 0: txt_status = f"BELUM LUNAS (Kurang Rp " + "{:,}".format(abs(diff)).replace(",", ".") + ")"
                            else: txt_status = f"LUNAS (Lebih Rp " + "{:,}".format(diff).replace(",", ".") + ")"
                            
                            curr_time = datetime.now().strftime('%H:%M')

                            final_price_str = "{:,}".format(int(price_per_month)).replace(",", ".")
                            final_bill_str = "{:,}".format(int(target_total)).replace(",", ".")

                            final_msg = tpl.replace('{name}', client.get('name', 'Pelanggan'))\
                                           .replace('{id}', f_id)\
                                           .replace('{amount}', f_amt)\
                                           .replace('{price}', final_price_str)\
                                           .replace('{bill}', final_bill_str)\
                                           .replace('{month}', get_month_name(datetime.now().month, s_wa.get('language', 'id')))\
                                           .replace('{date}', date_)\
                                           .replace('{expired}', f_exp)\
                                           .replace('{profile}', p_name)\
                                           .replace('{time}', curr_time)\
                                           .replace('{status}', txt_status)
                            
                            # Spawn worker in "test" mode (single message)
                            spawn_wa_worker(mode="test", target=p_clean, message=final_msg)
                        # Log removed
                        pass
            except Exception as wa_e:
                print(f"[PAYMENT_NOTIF_ERR] {wa_e}")

    except Exception as e:
        print(f"[FINANCE_HOOK] Error: {e}")

    # Logic: Return newest expiry for immediate UI refresh
    paid_until_rt = None
    if client:
        db_u = load_db(force_refresh=True)
        cl_upd = next((c for c in db_u.get('clients', []) if str(c.get('id','')).strip().upper() == str(client['id']).strip().upper()), {})
        paid_until_rt = cl_upd.get('paid_until')
    return jsonify({"status": "ok", "id": new_tx['id'], "paid_until": paid_until_rt})

@app.route('/api/billing/client/<client_id>/update_expiry', methods=['POST'])
def update_client_expiry(client_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    data = request.json
    new_expiry = data.get('paid_until') # Expected format YYYY-MM-DD (converted from DD-MM-YYYY in frontend)
    
    if not new_expiry:
        return jsonify({"error": "Expiry date is required"}), 400
        
    db_data = load_db()
    found = False
    client_name = "Unknown"
    for c in db_data.get('clients', []):
        if str(c.get('id','')).strip().upper() == str(client_id).strip().upper():
            client_name = c.get('name', client_id)
            # Targeted Update: Use apply_bulk_updates to bypass save_db protection
            # Auto-Enable & Set Anniversary from manual date edit
            p_until = new_expiry
            b_day = None
            b_enabled = None
            
            billing_config = load_billing_config()
            if billing_config.get('billing_mode') == 'fixed':
                try:
                    # Extract Day from YYYY-MM-DD
                    b_day = int(new_expiry.split('-')[2])
                    b_enabled = True
                except: pass

            client_update = {'id': client_id, 'paid_until': p_until}
            if b_day is not None:
                # Need to update specific billing fields in nested data
                # Load current client to merge billing sub-object
                topology = db.load_full_topology()
                client_obj = next((c for c in topology['clients'] if str(c['id']).strip().upper() == str(client_id).strip().upper()), None)
                if client_obj:
                    billing = client_obj.get('billing') or {}
                    billing['billing_day'] = b_day
                    billing['enabled'] = b_enabled
                    client_update['billing'] = billing

            apply_bulk_updates([client_update])
            
            # Record manual change in Finance history to satisfy sync_billing_from_finance
            try:
                fin_data = load_finance()
                new_tx = {
                    "id": str(int(time.time() * 1000)),
                    "date": get_local_now().strftime('%Y-%m-%d'),
                    "type": "income",
                    "category": "wifi_payment",
                    "amount": 0,
                    "note": f"Manual Expiry Update: Lunas s/d {new_expiry} (By Admin)",
                    "client_id": client_id,
                    "user": "SYSTEM"
                }
                fin_data.setdefault('transactions', []).append(new_tx)
                save_finance(fin_data)
            except Exception as fe:
                print(f"[BILLING] Failed to record manual update in finance: {fe}")

            found = True; break
            
    if found:
        add_log("SYSTEM", "system", f"Manual Expiry Update: {client_name} -> {new_expiry}")
        return jsonify({"status": "ok", "msg": f"Masa aktif berhasil diupdate ke {new_expiry}", "paid_until": new_expiry})
    return jsonify({"error": "Client not found"}), 404

@app.route('/api/finance/transaction/<tx_id>', methods=['DELETE'])
def fin_delete_tx(tx_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    data = load_finance()
    txs = data.get('transactions', [])
    
    # Find transaction to log it
    target_tx = next((t for t in txs if str(t.get('id')) == str(tx_id)), None)
    if not target_tx:
        return jsonify({"error": "Transaction not found"}), 404
        
    # Filter out the transaction
    new_txs = [t for t in txs if str(t.get('id')) != str(tx_id)]
    
    data['transactions'] = new_txs
    save_finance(data)
    
    # Log deletion
    desc = target_tx.get('note', '') or target_tx.get('category', 'tx')
    amt = target_tx.get('amount', 0)
    client_id = target_tx.get('client_id')
    add_log("SYSTEM", "system", f"Pembayaran Rp {amt} Dibatalkan. (Ket awal: {desc})")
    
    # --- SMART EXPIRY CHECK AFTER DELETE (Bug Fix: Phantom PAID status) ---
    if client_id:
        try:
            db_data = load_db()
            found_client = False
            for c in db_data.get('clients', []):
                if str(c.get('id','')).strip().upper() == str(client_id).strip().upper():
                    # Kita dapatkan klien yang transaksinya barusan dihapus
                    now_str = get_local_now().strftime('%Y-%m-%d')
                    expiry = c.get('paid_until', '')
                    
                    if 'billing' not in c:
                        c['billing'] = {}
                        
                    # 1. Jika tanggal expiry sudah lewat ATAU kosong, set status jadi nunggak
                    if not expiry or expiry < now_str:
                        c['billing']['payment_status'] = 'overdue'
                    # 2. Jika tanggal expiry hari ini atau ke depan, belum telat. Beri status belum terbayar 'unpaid' / biarkan sesuai rules.
                    else:
                        c['billing']['payment_status'] = 'unpaid'
                    
                    found_client = True
                    break
            
            if found_client:
                save_db(db_data, preserve_live=False)
                # Note: preserve_live=False assumes topology is mostly static or handled by thread
        except Exception as e:
            print(f"[DELETE TX HOOK ERR] {e}")
            
    return jsonify({"status": "ok", "msg": "Transaksi berhasil dihapus"})


@app.route('/api/telegram/test', methods=['POST'])
def test_telegram_connection():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json or {}
    token = data.get('token')
    chat_id = data.get('chat_id')
    
    if not token or not chat_id:
        return jsonify({"status": "error", "msg": "Token & Chat ID required"})
        
    try:
        msg = f"📢 NMS Connectivity Test\n\n✅ Bot Linked Successfully!\n✅ System: {SERVICE_NAME}"
        res = requests.post(f"https://api.telegram.org/bot{token}/sendMessage", 
                     data={'chat_id': chat_id, 'text': msg}, timeout=10)
        
        if res.status_code == 200:
            return jsonify({"status": "ok", "msg": "Test Message Sent!"})
        else:
            return jsonify({"status": "error", "msg": f"Telegram Error: {res.text}"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})

@app.route('/api/backup/download/<filename>')
def download_backup_file(filename):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    # Security: basic sanitization
    filename = os.path.basename(filename) 
    backup_dir = os.path.join(SCRIPT_DIR, 'backups')
    return send_from_directory(backup_dir, filename, as_attachment=True)

@app.route('/api/backup/list')
def list_backups_api():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    backup_dir = os.path.join(SCRIPT_DIR, 'backups')
    if not os.path.exists(backup_dir): return jsonify([])
    
    files = []
    for f in os.listdir(backup_dir):
        if f.endswith('.zip'):
            fp = os.path.join(backup_dir, f)
            files.append({
                "name": f,
                "size": round(os.path.getsize(fp) / 1024, 1), # KB
                "date": datetime.fromtimestamp(os.path.getmtime(fp)).strftime('%Y-%m-%d %H:%M')
            })
    
    # Sort new first
    files.sort(key=lambda x: x['date'], reverse=True)
    return jsonify(files)

@app.route('/api/backup/trigger', methods=['POST'])
def trigger_manual_backup():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    # Run in thread with force=True
    threading.Thread(target=auto_backup_logic, kwargs={'force': True}).start()
    return jsonify({"status": "ok", "msg": "Backup process started in background"})

@app.route('/api/backup/delete/<filename>', methods=['DELETE'])
def delete_backup_file(filename):
    if check_auth(request) != 'admin': return jsonify({"error": "Unauthorized"}), 401
    
    filename = os.path.basename(filename)
    backup_dir = os.path.join(SCRIPT_DIR, 'backups')
    fp = os.path.join(backup_dir, filename)
    
    if os.path.exists(fp):
        try:
            os.remove(fp)
            return jsonify({"status": "ok", "msg": f"File {filename} deleted"})
        except Exception as e:
            return jsonify({"status": "error", "msg": str(e)})
    return jsonify({"status": "error", "msg": "File not found"}), 404

@app.route('/api/finance/transaction/<tx_id>', methods=['PUT'])
def fin_edit_tx(tx_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    payload = request.json or {}
    amt = payload.get('amount')
    note = payload.get('note')
    
    data = load_finance()
    txs = data.get('transactions', [])
    
    found = False
    for t in txs:
        if str(t.get('id')) == str(tx_id):
            if amt is not None: t['amount'] = int(amt)
            if note is not None: t['note'] = note
            found = True
            break
            
    if not found:
        return jsonify({"error": "Transaction not found"}), 404
        
    save_finance(data)
    return jsonify({"status": "ok"})





@app.route('/api/client/<client_id>/bypass', methods=['POST'])
def toggle_client_bypass(client_id):
    if not check_auth(request):
        return jsonify({"error": "Unauthorized"}), 401
        
    db_data = load_db()
    data = request.json or {}
    target_state = data.get('bypass', None) 
    
    found = False
    new_state = False
    
    for c in db_data.get("clients", []):
        if str(c.get('id','')).strip().upper() == str(client_id).strip().upper():
            if target_state is not None:
                c['bypass_billing'] = bool(target_state)
            else:
                c['bypass_billing'] = not c.get('bypass_billing', False)
                
            new_state = c['bypass_billing']
            found = True
            break
            
    if not found:
        return jsonify({"error": "Client not found"}), 404
        
    save_db(db_data)
    return jsonify({"status": "ok", "bypass": new_state})


@app.route('/api/billing/client/<client_id>/enable', methods=['POST'])
def billing_enable(client_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    db_data = load_db()
    found = False
    for c in db_data.get("clients", []):
        if str(c.get('id','')).strip().upper() == str(client_id).strip().upper():
            if 'billing' not in c: c['billing'] = {}
            c['billing']['enabled'] = True
            found = True; break
    if found: 
        save_db(db_data)
        return jsonify({"status": "ok"})
    return jsonify({"error": "Not found"}), 404

@app.route('/api/billing/client/<client_id>/disable', methods=['POST'])
def billing_disable(client_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    db_data = load_db()
    found = False
    for c in db_data.get('clients', []):
        if str(c.get('id','')).strip().upper() == str(client_id).strip().upper():
            if 'billing' not in c: c['billing'] = {}
            c['billing']['enabled'] = False
            found = True; break
    if found:
        save_db(db_data)
        return jsonify({"status": "ok"})
    return jsonify({"error": "Not found"}), 404

@app.route('/api/billing/bulk_toggle', methods=['POST'])
def billing_bulk_toggle():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    data = request.json
    enabled = data.get('enabled', False)
    db_data = load_db()
    for c in db_data.get('clients', []):
        if 'billing' not in c: c['billing'] = {}
        c['billing']['enabled'] = enabled
    save_db(db_data)
    return jsonify({"status": "ok", "msg": f"Billing {'Enabled' if enabled else 'Disabled'} for all clients"})

@app.route('/api/billing/client/<client_id>/settings', methods=['POST'])
def billing_settings(client_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    db_data = load_db()
    data = request.json
    for c in db_data.get('clients', []):
        if str(c.get('id','')).strip().upper() == str(client_id).strip().upper():
            if 'billing' not in c: c['billing'] = {}
            if 'billing_day' in data: c['billing']['billing_day'] = int(data['billing_day'])
            if 'price' in data: c['billing']['price'] = int(data['price'])
            save_db(db_data); return jsonify({"status": "ok"})
    return jsonify({"error": "Not found"}), 404

@app.route('/api/billing/run_notify_manual', methods=['GET', 'POST'])
def api_billing_run_notify_manual():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    try:
        data = request.get_json(silent=True) or {}
        mode = data.get('template_mode', 'auto')
        run_billing_check(notify_only=True, force=True, template_mode=mode)
        return jsonify({"status": "ok", "msg": f"Proses pengiriman notifikasi manual dimulai (Mode: {mode})."})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)})

@app.route('/api/billing/today_schedule')
def api_billing_today_schedule():
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    import calendar
    db = load_db()
    now = get_local_now()
    today_str = now.strftime('%Y-%m-%d')
    curr_period = now.strftime('%m-%Y')
    
    cfg_billing = load_billing_config()
    default_day = cfg_billing.get('default_billing_day', 20)
    grace_period = cfg_billing.get('grace_period_days', 3)
    wa_pre_isolir_days = cfg_billing.get('wa_pre_isolir_days', 2)
    wa_end_month_active = cfg_billing.get('wa_end_month_enabled', True)
    wa_pre_isolir_active = cfg_billing.get('wa_pre_isolir_enabled', True)
    
    settings_root = _load_settings_raw()
    manual_arrears_list = settings_root.get('manual_arrears', [])

    schedule = []
    db_changed = False
    
    for client in db.get("clients", []):
        billing = client.get('billing', {})
        if not billing.get('enabled', False): continue
        if client.get('bypass_billing', False): continue
        
        # Proactive Sync from Finance - RESTORED
        if sync_billing_from_finance(client, db):
            db_changed = True
            billing = client.get('billing', {})
        
        paid_until_str = client.get('paid_until')
        payment_status = billing.get('payment_status', 'unpaid')
        
        # Dynamic Expiry Check for Dashboard at 00:00 before 09:00 engine
        if paid_until_str:
            try:
                exp_date = datetime.strptime(paid_until_str, '%Y-%m-%d').date()
                if exp_date <= now.date(): payment_status = 'unpaid'
            except: pass
            
        # Also check Manual Arrears
        manual_arr_val = sum(int(ma.get('amount', 0)) for ma in manual_arrears_list if ma.get('client_name') == client.get('name'))
        
        # SKIP TRULY PAID CLIENTS
        if payment_status == 'paid' and manual_arr_val == 0: continue
        
        track = billing.get('wa_sent_track', {})
        
        due_date_obj = None
        if paid_until_str:
            try:
                due_date_obj = datetime.strptime(paid_until_str, '%Y-%m-%d')
            except: pass
            
        if not due_date_obj:
            b_day = billing.get('billing_day') or default_day
            due_date_obj = calculate_due_date(now.year, now.month, b_day)
            
        days_overdue = (now - due_date_obj).days
        
        # Determine Potential Reasons for Today
        reasons = []
        
        # 1. REMOVED: H-3 Reminder (Per User Request)
        
        # 2. Due Date / EOM
        if wa_end_month_active:
             b_mode = cfg_billing.get('billing_mode', 'monthly')
             last_day_of_month = calendar.monthrange(now.year, now.month)[1]
             if b_mode == 'cyclic' and days_overdue == 0:
                 reasons.append({"type": "Jatuh Tempo", "key": "eom"})
             elif b_mode == 'fixed':
                 # Anniversary sync
                 b_day = billing.get('billing_day') or default_day
                 last_day_in_mo = calendar.monthrange(now.year, now.month)[1]
                 if now.day == b_day or (b_day > last_day_in_mo and now.day == last_day_in_mo):
                     reasons.append({"type": "Jatuh Tempo (Fixed)", "key": "eom"})
             elif b_mode == 'monthly' and now.day == last_day_of_month:
                 reasons.append({"type": "Akhir Bulan", "key": "eom"})
                 
        # 3. Pre-Isolation (Suppressed in Fixed Mode for Professional Standard)
        if wa_pre_isolir_active:
            trigger_day = (grace_period + 1) - wa_pre_isolir_days
            if days_overdue == trigger_day:
                reasons.append({"type": f"H-{wa_pre_isolir_days} Sebelum Isolir", "key": "pre_isolir"})
                
        # 4. Isolation
        # Theoretical isolation is when grace period is exactly hit today
        if days_overdue == grace_period + 1:
            reasons.append({"type": "Auto Isolir", "key": "isolir"})

        # PERSISTENCE LOGIC
        performed_today = []
        # REMOVED: H-3 Reminder Tracking (Per User Request)
        if track.get('eom') == curr_period and days_overdue >= 0: performed_today.append({"type": "Notif Tagihan", "key": "eom"})
        if track.get('pre_isolir') == curr_period and days_overdue >= ((grace_period + 1) - wa_pre_isolir_days): 
            performed_today.append({"type": "Peringatan Isolir", "key": "pre_isolir"})
        
        # Specific check for Today's Isolation
        if billing.get('isolir_date') == today_str:
            performed_today.append({"type": "Auto Isolir", "key": "isolir"})

        # Combine items
        all_matches = reasons + performed_today
        seen_keys = set()
        unique_matches = []
        for am in all_matches:
            if am['key'] not in seen_keys:
                unique_matches.append(am)
                seen_keys.add(am['key'])

        phone_num = client.get('phone') or client.get('whatsapp_number') or client.get('wa_number') or billing.get('wa_number') or '-'
        
        for m in unique_matches:
            status = "Menunggu"
            
            # CHECK FOR COMPLETION / RESOLUTION
            if payment_status == 'paid' and manual_arr_val == 0:
                status = "Selesai / Lunas"
            elif client.get('status') == 'online' and m['key'] == 'isolir':
                status = "Selesai / Lunas"
            elif track.get(m['key']) == curr_period:
                status = "Terkirim"
            
            # Specific check for Terisolir status
            if status != "Selesai / Lunas" and m['key'] == 'isolir':
                if client.get('status') == 'isolir' or track.get('isolir_wa_sent'):
                    status = "Terisolir"
            
            schedule.append({
                "name": client.get('name', 'N/A'),
                "phone": phone_num,
                "reason": m['type'],
                "reason_key": m['key'],
                "status": status,
                "date": today_str
            })
            
    if db_changed:
        save_db(db)
            
    return jsonify({"date": now.strftime('%d %B %Y'), "schedule": schedule})

@app.route('/api/logs')
def get_logs_route():
    if not check_auth(request): return jsonify({"error":"Unauthorized"}), 401
    logs = []
    if os.path.exists(LOG_FILE):
        try:
            with open(LOG_FILE, 'r') as f: logs = json.load(f)
        except: pass
    return jsonify(logs)

@app.route('/api/logo', methods=['POST', 'DELETE'])
def handle_logo():
    """Handle logo upload and deletion"""
    # Check auth and role
    role = check_auth(request)
    if not role:
        return jsonify({"status": "error", "msg": "Unauthorized"}), 401
    if role != 'admin':
        return jsonify({"status": "error", "msg": "Admin only"}), 403
    
    if request.method == 'POST':
        # Upload logo
        if 'file' not in request.files:
            return jsonify({"status": "error", "msg": "No file"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"status": "error", "msg": "No file selected"}), 400
        
        # Validate file type
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'svg', 'webp'}
        ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if ext not in allowed_extensions:
            return jsonify({"status": "error", "msg": "Invalid file type. Allowed: PNG, JPG, GIF, SVG"}), 400
        
        # Check file size (max 500KB)
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        if file_size > 500 * 1024:
            return jsonify({"status": "error", "msg": "File too large. Max 500KB"}), 400
        
        # Save logo (always as logo.png for simplicity)
        try:
            logo_path = os.path.join(SCRIPT_DIR, 'static', 'logo.png')
            tmp_logo = logo_path + ".tmp"
            file.save(tmp_logo)
            if os.path.exists(logo_path): os.remove(logo_path)
            os.rename(tmp_logo, logo_path)
            return jsonify({"status": "ok"})
        except Exception as e:
            return jsonify({"status": "error", "msg": str(e)}), 500
    
    elif request.method == 'DELETE':
        # Delete logo
        try:
            logo_path = os.path.join(SCRIPT_DIR, 'static', 'logo.png')
            if os.path.exists(logo_path):
                os.remove(logo_path)
            return jsonify({"status": "ok"})
        except Exception as e:
            return jsonify({"status": "error", "msg": str(e)}), 500

@app.route('/api/qris', methods=['POST', 'DELETE'])
def handle_qris():
    """Handle QRIS upload and deletion"""
    if check_auth(request) != 'admin':
        return jsonify({"error": "Forbidden"}), 403

    try:
        # Ensure photos directory exists
        photos_dir = os.path.join(SCRIPT_DIR, 'static', 'photos')
        if not os.path.exists(photos_dir):
            os.makedirs(photos_dir)
            
        qris_path = os.path.join(photos_dir, 'qris.jpg')

        if request.method == 'POST':
            if 'file' not in request.files:
                return jsonify({"status": "error", "message": "No file part"}), 400
            file = request.files['file']
            if file.filename == '':
                return jsonify({"status": "error", "message": "No selected file"}), 400
            
            # Simple validation
            ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
            if ext not in ['jpg', 'jpeg', 'png']:
                return jsonify({"status": "error", "message": "Format harus JPG/PNG"}), 400

            # Save directly as qris.jpg (using atomic pattern)
            try:
                tmp_qris = qris_path + ".tmp"
                file.save(tmp_qris)
                if os.path.exists(qris_path): os.remove(qris_path)
                os.rename(tmp_qris, qris_path)
                return jsonify({"status": "ok"})
            except Exception as e:
                return jsonify({"status": "error", "message": f"Save failed: {str(e)}"}), 500
        
        elif request.method == 'DELETE':
            if os.path.exists(qris_path):
                os.remove(qris_path)
            return jsonify({"status": "ok"})
            
    except Exception as e:
        print(f"[QRIS] Error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


def get_router_connection(router_id):
    """Helper to get routeros_api connection"""
    try:
        if not routeros_api: return None # Library missing
        
        data = load_db()
        if router_id == 'server_utama':
            r_data = data.get("server")
        else:
            r_data = next((r for r in data.get("extra_routers", []) if r["id"] == router_id), None)
        
        if not r_data or not r_data.get("login"): return None
        login = r_data["login"]
        
        conn = routeros_api.RouterOsApiPool(
            login.get("host"), 
            username=login.get("user"), 
            password=login.get("pass"),
            port=int(login.get("port", 8728)),
            plaintext_login=True
        )
        return conn
    except Exception as e:
        print(f"[ERROR] Router Conn: {e}")
        return None

@app.route('/api/hotspot/profiles/<router_id>', methods=['GET'])
def get_hotspot_profiles(router_id):
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed or Lib Missing"}), 500
    
    try:
        api = conn.get_api()
        
        # 1. Get ALL Server Profiles first (to get DNS names)
        # Equivalent to /ip/hotspot/profile/print
        srv_profiles_res = api.get_resource('/ip/hotspot/profile').get()
        # Mapping: Profile Name -> DNS Name
        prof_dns_map = {}
        for p in srv_profiles_res:
            nm = p.get('name')
            if nm:
                dns = p.get('dns-name')
                if not dns: dns = p.get('hotspot-address')
                prof_dns_map[nm] = dns or ""

        # 2. Get ACTUAL Hotspot Servers
        # Equivalent to /ip/hotspot/print
        hotspots_res = api.get_resource('/ip/hotspot').get()
        
        servers = []
        server_dns_map = {}
        for hs in hotspots_res:
            name = hs.get('name')
            if name:
                servers.append(name)
                prof_name = hs.get('profile')
                server_dns_map[name] = prof_dns_map.get(prof_name, "")

        servers.sort()
        
        # 3. Get User Profiles
        usr_profiles_res = api.get_resource('/ip/hotspot/user/profile').get()
        usr_profiles = sorted(list(set([x.get('name') for x in usr_profiles_res if x.get('name')])))
        
        # 4. Get Address Pools
        pools_res = api.get_resource('/ip/pool').get()
        pools = sorted([x.get('name') for x in pools_res if x.get('name')])

        # 5. Get Queues & Queue Types
        queues_res = api.get_resource('/queue/simple').get()
        queues = sorted([x.get('name') for x in queues_res if x.get('name')])
        
        qtypes_res = api.get_resource('/queue/type').get()
        qtypes = sorted([x.get('name') for x in qtypes_res if x.get('name')])

        # 6. Get Firewall Address Lists
        addr_lists_res = api.get_resource('/ip/firewall/address-list').get()
        addr_lists = sorted(list(set([x.get('list') for x in addr_lists_res if x.get('list')])))

        return jsonify({
            "status": "ok",
            "server_profiles": servers,
            "server_dns_map": server_dns_map,
            "user_profiles": usr_profiles,
            "pools": pools,
            "queues": queues,
            "queue_types": qtypes,
            "address_lists": addr_lists
        })
    except Exception as e:
        print(f"[HOTSPOT] Profile Error: {e}")
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/profile/list/<router_id>', methods=['GET'])
def get_hotspot_profile_list(router_id):
    """Get detailed list of user profiles"""
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify([]), 500
    
    try:
        api = conn.get_api()
        res = api.get_resource('/ip/hotspot/user/profile').get()
        
        profiles = []
        for r in res:
            pid = r.get('.id') or r.get('id')
            profiles.append({
                ".id": pid,
                "id": pid,
                "profile_id": pid,
                "name": r.get('name'),
                "address-pool": r.get('address-pool', 'none'),
                "session-timeout": r.get('session-timeout', ''),
                "idle-timeout": r.get('idle-timeout', 'none'),
                "keepalive-timeout": r.get('keepalive-timeout', '00:02:00'),
                "status-autorefresh": r.get('status-autorefresh', '00:01:00'),
                "shared-users": r.get('shared-users', '1'),
                "rate-limit": r.get('rate-limit', ''),
                "add-mac-cookie": r.get('add-mac-cookie', 'false'),
                "mac-cookie-timeout": r.get('mac-cookie-timeout', '3d 00:00:00'),
                "address-list": r.get('address-list', ''),
                "parent-queue": r.get('parent-queue', 'none'),
                "queue-type": r.get('queue-type', 'default-small'),
                "insert-queue-before": r.get('insert-queue-before', 'first'),
                "status": r.get('status', '')
            })
        return jsonify(profiles)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/profile/add', methods=['POST'])
def add_hotspot_profile():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    name = data.get('name')
    rate_limit = data.get('rate_limit')
    shared_users = data.get('shared_users')
    session_timeout = data.get('session_timeout')
    idle_timeout = data.get('idle_timeout')
    keepalive_timeout = data.get('keepalive_timeout')
    status_autorefresh = data.get('status_autorefresh')
    address_pool = data.get('address_pool')
    add_mac_cookie = data.get('add_mac_cookie')
    mac_cookie_timeout = data.get('mac_cookie_timeout')
    address_list = data.get('address_list')
    
    # Queue fields
    parent_queue = data.get('parent_queue')
    queue_type = data.get('queue_type')
    insert_queue_before = data.get('insert_queue_before')

    if not name: return jsonify({"status": "error", "msg": "Name Required"}), 400
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res_prof = api.get_resource('/ip/hotspot/user/profile')
        
        payload = {'name': name}
        if rate_limit: payload['rate-limit'] = rate_limit
        if shared_users: payload['shared-users'] = str(shared_users)
        if session_timeout: payload['session-timeout'] = session_timeout
        if idle_timeout: payload['idle-timeout'] = idle_timeout
        if keepalive_timeout: payload['keepalive-timeout'] = keepalive_timeout
        if status_autorefresh: payload['status-autorefresh'] = status_autorefresh
        if address_pool: payload['address-pool'] = address_pool
        if add_mac_cookie is not None: payload['add-mac-cookie'] = 'yes' if add_mac_cookie else 'no'
        if mac_cookie_timeout: payload['mac-cookie-timeout'] = mac_cookie_timeout
        if address_list: payload['address-list'] = address_list

        # Queue fields
        if parent_queue: payload['parent-queue'] = parent_queue
        if queue_type: payload['queue-type'] = queue_type
        if insert_queue_before: payload['insert-queue-before'] = insert_queue_before
        
        res_prof.add(**payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/profile/remove', methods=['POST'])
def remove_hotspot_profile():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin': return jsonify({"error": "Admin only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    pid = data.get('id')
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res_prof = api.get_resource('/ip/hotspot/user/profile')
        res_prof.remove(id=pid)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/generate', methods=['POST'])
def generate_hotspot_users():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician':
        return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    qty = int(data.get('qty', 1))
    server = data.get('server')
    if not server or server == 'all': server = 'all'
    profile = data.get('profile')
    timelimit = data.get('timelimit')
    datalimit = data.get('datalimit')
    comment = data.get('comment', 'generated-nms')
    prefix = data.get('prefix', '')
    mode = data.get('mode', 'up') 
    char_type = data.get('char_type', 'rand')
    char_len = int(data.get('char_len', 6))
    
    import random, string
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "message": "Connection Failed"}), 500
    
    generated = []
    errors = []
    
    try:
        api = conn.get_api()
        res_user = api.get_resource('/ip/hotspot/user')
        
        # Char sets
        chars = string.digits
        if char_type == 'abc': chars = string.ascii_lowercase
        elif char_type == 'ABC': chars = string.ascii_uppercase
        elif char_type == 'mix': chars = string.ascii_lowercase + string.digits
        elif char_type == 'MIX': chars = string.ascii_letters + string.digits

        for i in range(qty):
            user_code = ''.join(random.choices(chars, k=char_len))
            username = f"{prefix}{user_code}"
            
            if mode == 'vc':
                password = ''.join(random.choices(chars, k=char_len))
            else:
                password = username
            
            payload = {
                'name': username,
                'password': password,
                'profile': profile,
                'comment': comment
            }
            
            if server and server != 'all': payload['server'] = server
            if timelimit: payload['limit-uptime'] = timelimit
            
            pbytes = parse_hotspot_limit_bytes(datalimit)
            if pbytes is not None: payload['limit-bytes-total'] = pbytes
            
            # Silently handle or log elsewhere
            pass
            try:
                res_user.add(**payload)
                # Log removed
                pass
                generated.append({
                    "name": username, 
                    "password": password, 
                    "profile": profile or "-", 
                    "limit_uptime": timelimit or "-",
                    "limit_bytes": datalimit or "-"
                })
            except Exception as e:
                # Log removed
                pass
                errors.append(f"Fail {username}: {e}")
                
        # Log removed
        pass
        
        if errors and not generated:
            return jsonify({
                "status": "error",
                "message": "Fail: " + (errors[0] if errors else "Unknown"),
                "errors": errors
            }), 400

        return jsonify({
            "status": "ok",
            "generated": len(generated),
            "errors": len(errors),
            "error_details": errors if errors else [],
            "users": generated
        })
    except Exception as e:
         print(f"[HS-GEN] FATAL ERROR: {e}")
         return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/users/<router_id>', methods=['GET'])
def get_hotspot_users(router_id):
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    comment_filter = request.args.get('comment')
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify([]), 500
    
    try:
        api = conn.get_api()
        res_user = api.get_resource('/ip/hotspot/user')
        
        if comment_filter:
            # Try exact match first
            res = res_user.get(comment=comment_filter)
            # If empty and high data volume, we should be careful. 
            # But we'll try basic filter if exact match fails.
            if not res: 
                # Limit fetch if no filter is possible or use manual filter defensively
                all_u = res_user.get()
                res = [u for u in all_u if comment_filter in u.get('comment', '')]
                # If too many items, limit to prevent UI crash? 
                # User requested 30 in UI, but API should return all or a reasonable max.
                if len(res) > 2000: res = res[:2000] # Safety cap
        else:
            # Fetch all but cap at 5000 for safety on branch nodes
            res = res_user.get()
            if len(res) > 5000: res = res[:5000] 
            
        users = []
        for r in res:
             mid = r.get('.id') or r.get('id')
             
             users.append({
                "id": mid, 
                "name": r.get('name'),
                "password": r.get('password'),
                "profile": r.get('profile'),
                "server": r.get('server', 'all'),
                "comment": r.get('comment', ''),
                "limit_uptime": r.get('limit-uptime', ''),
                "limit_bytes": r.get('limit-bytes-total') or r.get('limit-bytes-out') or '',
                "disabled": r.get('disabled') == 'true'
            })
        return jsonify(users)
    except Exception as e:
        print(f"[HS-USERS] Error: {e}")
        return jsonify([])
    finally:
        if conn: conn.disconnect()

@app.route('/api/hotspot/delete', methods=['POST'])
def delete_hotspot_users():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    ids = data.get('ids', [])
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    success = 0
    last_error = ""
    try:
        api = conn.get_api()
        res_user = api.get_resource('/ip/hotspot/user')
        
        for mid in ids:
            try:
                res_user.remove(id=mid)
                success += 1
            except Exception as e:
                # Capture the error to return it if needed
                last_error = str(e)
            
        return jsonify({"status": "ok", "deleted": success, "last_error": last_error})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/active/<router_id>', methods=['GET'])
def get_active_hotspot(router_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify([]), 500
    
    try:
        api = conn.get_api()
        # Ensure we always get a list and handle potential latency better
        res = api.get_resource('/ip/hotspot/active').get()
        
        # Normalize data for frontend to ensure keys exist and dashes are handled
        active = []
        for r in res:
            active.append({
                ".id": r.get('.id') or r.get('id'),
                "server": r.get('server', 'all'),
                "user": r.get('user', r.get('name', 'unknown')),
                "address": r.get('address', '-'),
                "uptime": r.get('uptime', '00:00:00'),
                "bytes-in": r.get('bytes-in', '0'),
                "bytes-out": r.get('bytes-out', '0')
            })
            
        return jsonify(active)
    except Exception as e:
        print(f"[HS-ACTIVE] Error: {e}")
        return jsonify([])
    finally:
        if conn: conn.disconnect()

@app.route('/api/hotspot/kick', methods=['POST'])
def kick_hotspot_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json
    router_id = data.get('router_id')
    mid = data.get('id')
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        api.get_resource('/ip/hotspot/active').remove(id=mid)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})
    finally:
        conn.disconnect()

# ==============================================================================
#  PPPoE MANAGEMENT API
# ==============================================================================

@app.route('/api/pppoe/profiles/<router_id>', methods=['GET'])
def get_pppoe_profiles(router_id):
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        # Get PPP Profiles
        profiles_res = api.get_resource('/ppp/profile').get()
        profiles = []
        for p in profiles_res:
            if not p.get('name'): continue
            # Split combined queue types if necessary (handle ROS v6 legacy or manual combined inputs)
            q_raw = p.get('queue-type', '')
            q_up = q_raw
            q_dn = p.get('download-queue-type', '')
            
            if '/' in q_raw and not q_dn:
                parts = q_raw.split('/')
                q_up = parts[0]
                q_dn = parts[1] if len(parts) > 1 else ''

            profiles.append({
                "id": p.get('.id') or p.get('id'),
                "name": p.get('name'),
                "local_address": p.get('local-address', ''),
                "local-address": p.get('local-address', ''),
                "remote_address": p.get('remote-address', ''),
                "remote-address": p.get('remote-address', ''),
                "dns_server": p.get('dns-server', ''),
                "dns-server": p.get('dns-server', ''),
                "rate_limit": p.get('rate-limit', ''),
                "rate-limit": p.get('rate-limit', ''),
                "only_one": p.get('only-one', 'default'),
                "only-one": p.get('only-one', 'default'),
                "insert_queue_before": p.get('insert-queue-before', ''),
                "insert-queue-before": p.get('insert-queue-before', ''),
                "parent_queue": p.get('parent-queue', ''),
                "parent-queue": p.get('parent-queue', ''),
                "queue_type": q_up,
                "queue-type": q_up,
                "download_queue_type": q_dn,
                "download-queue-type": q_dn,
                "raw": p
            })
        
        # Sort by name
        profiles = sorted(profiles, key=lambda x: x['name'])

        # Fetch router identity for display in frontend dropdown
        try:
            identity_res = api.get_resource('/system/identity').get()
            router_identity = identity_res[0].get('name', '') if identity_res else ''
        except Exception:
            router_identity = ''

        return jsonify({
            "status": "ok",
            "profiles": profiles,
            "router_identity": router_identity
        })
    except Exception as e:
        print(f"[PPPoE] Profile Error: {e}")
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/pppoe/profile_options/<router_id>', methods=['GET'])
def get_pppoe_profile_options(router_id):
    """Return IP pools, queue names, and queue types for the profile creation form."""
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401

    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500

    try:
        api = conn.get_api()
        # IP Pools
        try:
            pools_raw = api.get_resource('/ip/pool').get()
            pools = [p.get('name') for p in pools_raw if p.get('name')]
        except Exception:
            pools = []
        # Queue simple list (for insert-before / parent)
        try:
            qs_raw = api.get_resource('/queue/simple').get()
            queues_simple = ['first', 'bottom'] + [q.get('name') for q in qs_raw if q.get('name')]
        except Exception:
            queues_simple = ['first', 'bottom']
        # Queue tree list (for parent)
        try:
            qt_raw = api.get_resource('/queue/tree').get()
            queues_tree = [q.get('name') for q in qt_raw if q.get('name')]
        except Exception:
            queues_tree = []
        # Queue types
        try:
            qtype_raw = api.get_resource('/queue/type').get()
            queue_types = [q.get('name') for q in qtype_raw if q.get('name')]
        except Exception:
            queue_types = ['default', 'ethernet-default', 'wireless-default', 'pcq-upload-default', 'pcq-download-default']

        # Fetch router identity for display in frontend dropdown
        router_identity = ''
        try:
            identity_res = api.get_resource('/system/identity').get()
            router_identity = identity_res[0].get('name', '') if identity_res else ''
        except Exception: pass

        return jsonify({
            "status": "ok",
            "pools": pools,
            "queues_simple": queues_simple,
            "queues_tree": queues_tree,
            "queue_types": queue_types,
            "router_identity": router_identity
        })
    except Exception as e:
        print(f"[PPPoE] profile_options Error: {e}")
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()


@app.route('/api/pppoe/profile/create', methods=['POST'])
def create_pppoe_profile():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role not in ('admin', 'technician'): return jsonify({"error": "Admin/Tech only"}), 403

    data = request.json
    router_id = data.get('router_id')
    name = data.get('name', '').strip()
    if not name:
        return jsonify({"status": "error", "msg": "Profile name required"}), 400

    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500

    try:
        api = conn.get_api()
        payload = {'name': name}
        q_up = (data.get('queue_type') or data.get('queue_type_upload') or '').strip()
        q_dn = (data.get('queue_type_download') or '').strip()
        
        optional_str = {
            'local-address': data.get('local_address', '').strip(),
            'remote-address': data.get('remote_address', '').strip(),
            'dns-server': data.get('dns_server', '').strip(),
            'rate-limit': data.get('rate_limit', '').strip(),
            'insert-queue-before': data.get('insert_queue_before', '').strip(),
            'parent-queue': data.get('parent_queue', '').strip(),
        }
        
        if q_up and q_dn:
            optional_str['queue-type'] = f"{q_up}/{q_dn}"
        elif q_up:
            optional_str['queue-type'] = q_up
        elif q_dn:
            optional_str['queue-type'] = f"default/{q_dn}"
        for k, v in optional_str.items():
            if v:
                payload[k] = v
        only_one = data.get('only_one', 'default')
        if only_one and only_one != 'default':
            payload['only-one'] = only_one

        api.get_resource('/ppp/profile').add(**payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()


@app.route('/api/pppoe/profile/update', methods=['POST'])
def update_pppoe_profile():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role not in ('admin', 'technician'): return jsonify({"error": "Admin/Tech only"}), 403

    data = request.json
    router_id = data.get('router_id')
    old_name = data.get('old_name', '').strip()
    new_name = data.get('name', '').strip()
    
    if not old_name:
        return jsonify({"status": "error", "msg": "Target profile name required"}), 400

    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500

    try:
        api = conn.get_api()
        res = api.get_resource('/ppp/profile')
        
        # Find the .id by name
        existing = res.get(name=old_name)
        if not existing:
            # Fallback for some Mikrotik/API versions where filtered get fails
            all_profs = res.get()
            match = [p for p in all_profs if p.get('name') == old_name]
            if match:
                existing = match
            else:
                return jsonify({"status": "error", "msg": f"Profile '{old_name}' not found"}), 404
            
        item = existing[0]
        target_id = item.get('.id') or item.get('id') or item.get('*ID')
        
        # If still no ID, maybe it's a 'default' profile that uses its name as ID in some ROS versions
        # or it's a special case. But normally we need an ID for .set()
        if not target_id:
            if old_name.lower() in ['default', 'default-encryption']:
                # For default profiles, sometimes the name itself works as ID in some API implementations
                target_id = old_name
            else:
                available_keys = list(item.keys())
                return jsonify({"status": "error", "msg": f"Could not determine ID for '{old_name}'. Keys found: {available_keys}"}), 500
        
        payload = {}
        if new_name and new_name != old_name:
            payload['name'] = new_name
            
        q_up = str(data.get('queue_type') or data.get('queue_type_upload') or '').strip()
        q_dn = str(data.get('queue_type_download') or '').strip()
            
        mapping = {
            'local-address': data.get('local_address'),
            'remote-address': data.get('remote_address'),
            'dns-server': data.get('dns_server'),
            'rate-limit': data.get('rate_limit'),
            'only-one': data.get('only_one'),
            'insert-queue-before': data.get('insert_queue_before'),
            'parent-queue': data.get('parent_queue'),
        }
        
        if q_up and q_dn: mapping['queue-type'] = f"{q_up}/{q_dn}"
        elif q_up: mapping['queue-type'] = q_up
        elif q_dn: mapping['queue-type'] = f"default/{q_dn}"
        
        for mk_key, val in mapping.items():
            if val is not None:
                val = str(val).strip()
                # Skip empty values to avoid 'ambiguous value of pool' error in Mikrotik
                if not val:
                    continue
                if mk_key == 'only-one' and val == 'default':
                    continue
                payload[mk_key] = val

        try:
            res.set(id=target_id, **payload)
        except Exception as set_err:
            # If set by ID fails, try set by name as fallback for some ROS versions/profiles
            try:
                # Some API implementations allow set by name if it's a unique key
                res.set(name=old_name, **payload)
            except:
                raise set_err # Re-raise original error if fallback also fails

        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()


@app.route('/api/pppoe/profile/delete', methods=['POST'])
def delete_pppoe_profile():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role not in ('admin', 'technician'): return jsonify({"error": "Admin/Tech only"}), 403

    data = request.json
    router_id = data.get('router_id')
    name = data.get('name', '').strip()
    if not name:
        return jsonify({"status": "error", "msg": "Profile name required"}), 400

    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500

    try:
        api = conn.get_api()
        res = api.get_resource('/ppp/profile')
        items = res.get(name=name)
        if not items:
            return jsonify({"status": "error", "msg": f"Profile '{name}' not found"}), 404
        mid = items[0].get('.id') or items[0].get('id')
        res.remove(id=mid)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()


@app.route('/api/pppoe/users/<router_id>', methods=['GET'])
def get_pppoe_users(router_id):
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    comment_filter = request.args.get('comment')
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify([]), 500
    
    try:
        api = conn.get_api()
        res_user = api.get_resource('/ppp/secret')
        
        if comment_filter:
            res = res_user.get(comment=comment_filter)
            if not res:
                all_u = res_user.get()
                res = [u for u in all_u if comment_filter in u.get('comment', '')]
        else:
            res = res_user.get()
            
        # Limit safety
        if len(res) > 5000: res = res[:5000]
            
        users = []
        for r in res:
             mid = r.get('.id') or r.get('id')
             users.append({
                "id": mid,
                "name": r.get('name'),
                "password": r.get('password'),
                "profile": r.get('profile'),
                "service": r.get('service', 'any'),
                "comment": r.get('comment', ''),
                "local_address": r.get('local-address', ''),
                "remote_address": r.get('remote-address', ''),
                "last_logged_out": r.get('last-logged-out', ''),
                "last_disconnect_reason": r.get('last-disconnect-reason', ''),
                "disabled": r.get('disabled') == 'true'
            })
        # Fetch router identity for display in frontend dropdown
        router_identity = ''
        try:
            identity_res = api.get_resource('/system/identity').get()
            router_identity = identity_res[0].get('name', '') if identity_res else ''
        except Exception: pass

        return jsonify({
            "status": "ok",
            "users": users,
            "router_identity": router_identity
        })
    except Exception as e:
        print(f"[PPPoE-USERS] Error: {e}")
        return jsonify({"status": "error", "msg": str(e), "users": []})
    finally:
        conn.disconnect()

@app.route('/api/pppoe/create', methods=['POST'])
def create_pppoe_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    name = data.get('name')
    password = data.get('password')
    profile = data.get('profile', 'default')
    service = data.get('service', 'pppoe')
    comment = data.get('comment', 'created-via-nms')
    local_address = data.get('local_address', '').strip()
    remote_address = data.get('remote_address', '').strip()
    
    if not name or not password:
        return jsonify({"status": "error", "msg": "Name and Password required"}), 400
        
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res_secret = api.get_resource('/ppp/secret')
        
        payload = {
            'name': name,
            'password': password,
            'profile': profile,
            'service': service,
            'comment': comment
        }
        if local_address:
            payload['local-address'] = local_address
        if remote_address:
            payload['remote-address'] = remote_address
        if not profile: profile = 'default'
        if not service: service = 'pppoe'
        payload['profile'] = profile
        payload['service'] = service
        
        res_secret.add(**payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/pppoe/update', methods=['POST'])
def update_pppoe_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    name = data.get('name')
    password = data.get('password')
    profile = data.get('profile')
    service = data.get('service')
    comment = data.get('comment')
    local_address = data.get('local_address', '').strip()
    remote_address = data.get('remote_address', '').strip()
    disabled = data.get('disabled') # Accept boolean or None
    
    if not name:
        return jsonify({"status": "error", "msg": "Name required"}), 400
        
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res_secret = api.get_resource('/ppp/secret')
        
        # Find the secret ID
        items = res_secret.get(name=name)
        if not items:
            return jsonify({"status": "error", "msg": "User not found"}), 404
        
        item = items[0]
        secret_id = item.get('.id') or item.get('id') or item.get('*ID')
        
        if not secret_id:
            available_keys = list(item.keys())
            return jsonify({"status": "error", "msg": f"Could not determine ID for user '{name}'. Keys found: {available_keys}"}), 500
        
        payload = {}
        if password is not None: payload['password'] = password
        if profile: payload['profile'] = profile
        if service: payload['service'] = service
        if comment is not None: payload['comment'] = comment
        if local_address: payload['local-address'] = local_address
        if remote_address: payload['remote-address'] = remote_address
        if disabled is not None: payload['disabled'] = 'yes' if disabled else 'no'
        
        res_secret.set(id=secret_id, **payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/pppoe/toggle', methods=['POST'])
def toggle_pppoe_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    name = data.get('name')
    
    if not name or not router_id:
        return jsonify({"status": "error", "msg": "Name and Router ID required"}), 400
        
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"status": "error", "msg": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res_secret = api.get_resource('/ppp/secret')
        items = res_secret.get(name=name)
        if not items:
            # Fallback check if name was passed as ID
            items = res_secret.get(**{'.id': name})
            if not items: items = res_secret.get(id=name)
            
        if not items:
            return jsonify({"status": "error", "msg": "User not found"}), 404
        
        item = items[0]
        secret_id = item.get('.id') or item.get('id') or item.get('*ID')
        if not secret_id:
            return jsonify({"status": "error", "msg": f"Could not determine ID for user '{name}'"}), 500
            
        current_disabled = item.get('disabled') == 'true'
        new_disabled = not current_disabled
        
        res_secret.set(id=secret_id, disabled='yes' if new_disabled else 'no')
        
        # If disabling, also kick the current session
        if new_disabled:
            try: kick_pppoe_user(name, router_id)
            except: pass
            
        return jsonify({"status": "ok", "disabled": new_disabled})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        conn.disconnect()


@app.route('/api/pppoe/delete', methods=['POST'])
def delete_pppoe_users():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    ids = data.get('ids', [])
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    success = 0
    last_error = ""
    try:
        api = conn.get_api()
        res_secret = api.get_resource('/ppp/secret')
        
        for mid in ids:
            try:
                # Robust ID lookup before delete
                target_user = None
                try: target_user = res_secret.get(**{'.id': mid})
                except: pass
                
                if not target_user:
                    try: target_user = res_secret.get(id=mid)
                    except: pass
                
                if target_user:
                    actual_id = target_user[0].get('.id') or target_user[0].get('id') or target_user[0].get('*ID')
                    res_secret.remove(id=actual_id)
                    success += 1
                else:
                    # Final attempt by name
                    try:
                        res_secret.remove(name=mid)
                        success += 1
                    except:
                        last_error = f"ID/Name {mid} not found on router"
            except Exception as e:
                last_error = str(e)
            
        return jsonify({"status": "ok", "deleted": success, "last_error": last_error})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/pppoe/active/<router_id>', methods=['GET'])
def get_active_pppoe(router_id):
    if not check_auth(request): return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify([]), 500
    
    try:
        api = conn.get_api()
        # Equivalent to /ppp/active/print
        res = api.get_resource('/ppp/active').get()
        
        active = []
        for r in res:
            active.append({
                "id": r.get('.id') or r.get('id'),
                "name": r.get('name', 'unknown'),
                "service": r.get('service', 'pppoe'),
                "address": r.get('address', '-'),
                "uptime": r.get('uptime', '00:00:00'),
                "caller_id": r.get('caller-id', '-')
            })
            
        # Fetch router identity for display in frontend dropdown
        router_identity = ''
        try:
            identity_res = api.get_resource('/system/identity').get()
            router_identity = identity_res[0].get('name', '') if identity_res else ''
        except Exception: pass

        return jsonify({
            "status": "ok",
            "active": active,
            "router_identity": router_identity
        })
    except Exception as e:
        print(f"[PPPoE-ACTIVE] Error: {e}")
        return jsonify({"status": "error", "msg": str(e), "active": []})
    finally:
        conn.disconnect()

@app.route('/api/pppoe/kick', methods=['POST'])
def kick_pppoe_active_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json
    router_id = data.get('router_id')
    mid = data.get('id')
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res_active = api.get_resource('/ppp/active')
        
        target = None
        try: target = res_active.get(**{'.id': mid})
        except: pass
        if not target:
            try: target = res_active.get(id=mid)
            except: pass
            
        if target:
            actual_id = target[0].get('.id') or target[0].get('id') or target[0].get('*ID')
            res_active.remove(id=actual_id)
            return jsonify({"status": "ok"})
        else:
             return jsonify({"status": "error", "msg": "Active session not found"})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)})
    finally:
        conn.disconnect()


@app.route('/api/hotspot/user/create', methods=['POST'])
def create_hotspot_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    name = data.get('name')
    password = data.get('password', '')
    profile = data.get('profile', 'default')
    server = data.get('server', 'all')
    comment = data.get('comment', '')
    limit_uptime = data.get('limit_uptime')
    limit_bytes = data.get('limit_bytes')
    
    if not name: return jsonify({"error": "Name required"}), 400
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res = api.get_resource('/ip/hotspot/user')
        payload = {
            "name": name,
            "password": password,
            "profile": profile,
            "server": server or 'all',
            "comment": comment
        }
        if limit_uptime: payload['limit-uptime'] = limit_uptime
        pbytes = parse_hotspot_limit_bytes(limit_bytes)
        if pbytes is not None: payload['limit-bytes-total'] = pbytes
        
        res.add(**payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/user/update', methods=['POST'])
def update_hotspot_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    user_id = data.get('id')
    name = data.get('name')
    password = data.get('password')
    profile = data.get('profile')
    comment = data.get('comment')
    limit_uptime = data.get('limit_uptime')
    limit_bytes = data.get('limit_bytes')
    
    if not user_id: return jsonify({"error": "ID required"}), 400
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res = api.get_resource('/ip/hotspot/user')
        
        payload = {}
        if name: payload['name'] = name
        if password is not None: payload['password'] = password
        if profile: payload['profile'] = profile
        if data.get('server') is not None: payload['server'] = data.get('server') or 'all'
        if comment is not None: payload['comment'] = comment
        if limit_uptime is not None: payload['limit-uptime'] = limit_uptime if limit_uptime else "0s"
        pbytes = parse_hotspot_limit_bytes(limit_bytes)
        if pbytes is not None: payload['limit-bytes-total'] = pbytes
        
        res.set(id=user_id, **payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/user/toggle', methods=['POST'])
def toggle_hotspot_user():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json
    router_id = data.get('router_id')
    user_id = data.get('id')
    disabled = data.get('disabled')
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res = api.get_resource('/ip/hotspot/user')
        res.set(id=user_id, disabled='yes' if disabled else 'no')
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

@app.route('/api/hotspot/profile/update', methods=['POST'])
def update_hotspot_profile():
    role = check_auth(request)
    if not role: return jsonify({"error": "Unauthorized"}), 401
    if role != 'admin' and role != 'technician': return jsonify({"error": "Admin/Tech only"}), 403
    
    data = request.json
    router_id = data.get('router_id')
    profile_id = data.get('id')
    rate_limit = data.get('rate_limit')
    shared_users = data.get('shared_users')
    session_timeout = data.get('session_timeout')
    idle_timeout = data.get('idle_timeout')
    keepalive_timeout = data.get('keepalive_timeout')
    status_autorefresh = data.get('status_autorefresh')
    address_pool = data.get('address_pool')
    add_mac_cookie = data.get('add_mac_cookie')
    mac_cookie_timeout = data.get('mac_cookie_timeout')
    address_list = data.get('address_list')
    
    # Queue fields
    parent_queue = data.get('parent_queue')
    queue_type = data.get('queue_type')
    insert_queue_before = data.get('insert_queue_before')
    
    if not profile_id: return jsonify({"error": "ID required"}), 400
    
    conn = get_router_connection(router_id)
    if not conn: return jsonify({"error": "Connection Failed"}), 500
    
    try:
        api = conn.get_api()
        res = api.get_resource('/ip/hotspot/user/profile')
        payload = {}
        if rate_limit is not None: payload['rate-limit'] = rate_limit
        if shared_users is not None: payload['shared-users'] = str(shared_users)
        if session_timeout is not None: payload['session-timeout'] = session_timeout
        if idle_timeout is not None: payload['idle-timeout'] = idle_timeout
        if keepalive_timeout is not None: payload['keepalive-timeout'] = keepalive_timeout
        if status_autorefresh is not None: payload['status-autorefresh'] = status_autorefresh
        if address_pool is not None: payload['address-pool'] = address_pool
        if add_mac_cookie is not None: payload['add-mac-cookie'] = 'yes' if add_mac_cookie else 'no'
        if mac_cookie_timeout is not None: payload['mac-cookie-timeout'] = mac_cookie_timeout
        if address_list is not None: payload['address-list'] = address_list

        # Queue fields
        if parent_queue: payload['parent-queue'] = parent_queue
        if queue_type: payload['queue-type'] = queue_type
        if insert_queue_before: payload['insert-queue-before'] = insert_queue_before
        
        res.set(id=profile_id, **payload)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.disconnect()

# --- START BACKGROUND THREADS (Gunicorn Compatible) ---
threading.Thread(target=telegram_listener_loop, daemon=True).start()

if __name__ == '__main__':
    # Flask Native Start
    port = int(cfg.get('app_port', 5002))
    app.run(host='0.0.0.0', port=port, threaded=True)
